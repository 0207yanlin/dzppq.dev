# -*- coding: utf-8 -*-
"""Generate a rebuilt DZPPQ meta report from the latest match database."""

from __future__ import annotations

import argparse
import html
import itertools
import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote


ROOT = Path(__file__).resolve().parents[4]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.card_rules import (  # noqa: E402
    normalize_card_label,
    resolve_jsb_xj_card_labels,
    split_card_prefix,
)
from src.match_db import ensure_match_schema, parse_match_batch  # noqa: E402

DEFAULT_LATEST_DB = ROOT / "data" / "match_latest.db"
DEFAULT_JSON = ROOT / "data" / "latest_meta_analysis.json"
DEFAULT_MD = ROOT / "data" / "latest_meta_analysis_report.md"
DEFAULT_INTERACTIVE_HTML = ROOT / "data" / "环境分析详情.html"
DEFAULT_XLSX = ROOT / "data" / "latest_meta_analysis_equipment.xlsx"
DEFAULT_HERO_EQUIPMENT_DIR = ROOT / "data" / "hero-equipment"
CARD_HTML_SUFFIXES = {
    "彩": "cai",
    "黄": "yellow",
    "蓝": "blue",
    "白": "white",
}
LEGACY_HTML_FILENAMES = [
    "环境分析一图流.html",
    "阵容推荐详情.html",
    "啾啾阵容依赖.html",
    "啾啾佩戴推荐.html",
    "英雄出装推荐.html",
    "版本陷阱阵容.html",
    "彩卡强度排行.html",
    "黄卡强度排行.html",
    "蓝卡强度排行.html",
    "白卡强度排行.html",
    "双人阵容配合.html",
    "低费主C追三难度.html",
    "主羁绊强度排行.html",
]
INTERACTIVE_PANELS: list[tuple[str, str, str]] = [
    ("compositions", "阵容推荐详情", "composition_recommendations"),
    ("primary-bond", "主羁绊强度排行", "primary_bond_strength"),
    ("equipment", "英雄出装推荐", "equipment"),
    ("super-equipment", "超级装备强度", "super_equipment"),
    ("food-equipment", "美食社装备强度", "food_equipment"),
    ("cards-cai", "彩卡强度排行", "cards_cai"),
    ("cards-yellow", "黄卡强度排行", "cards_yellow"),
    ("cards-blue", "蓝卡强度排行", "cards_blue"),
    ("cards-white", "白卡强度排行", "cards_white"),
    ("duo", "双人阵容配合", "duo_compositions"),
    ("low-cost", "低费主C追三难度", "low_cost_carries"),
    ("jiujiu-comps", "啾啾阵容依赖", "jiujiu_comps"),
    ("jiujiu-wearers", "啾啾佩戴推荐", "jiujiu_wearers"),
    ("traps", "版本陷阱阵容", "trap_compositions"),
]
SUPER_EQUIPMENT_NAMES = frozenset(
    {
        "巫术玩偶",
        "小鲨包",
        "金咸鱼",
        "幸运猫猫",
        "碰碰气球",
        "炸炸魔术箱",
        "发财树",
        "核桃火箭",
        "鲱鱼罐头",
    }
)
FOOD_SPECIAL_EQUIPMENT_NAMES = frozenset({"杏仁豆腐", "椒盐酥糖", "岛好锅"})
FOOD_HARVEST_PREFIXES = ("美味", "绝味", "暗黑")
SPECIAL_EQUIPMENT_RELIABLE_MIN = 8
SPECIAL_EQUIPMENT_WEARER_MIN = 3
DEFAULT_RECENCY_HALF_LIFE_DAYS = 2.0
MIN_RECENCY_WEIGHT = 0.0
TREND_RECENT_BATCHES = 2
TREND_PRIOR_BATCHES = 2
TREND_MIN_SAMPLES_PER_WINDOW = 4
TREND_AVG_RANK_THRESHOLD = 0.15
TREND_TOP4_THRESHOLD = 5.0
TREND_PICK_RATE_THRESHOLD = 0.5
RECOMMENDATION_MIN_RAW_N = 10
RECOMMENDATION_MIN_WEIGHTED_N = 5.0
RECOMMENDATION_MIN_EFFECTIVE_N = 8.0
RECOMMENDATION_MIN_BATCHES = 2
RECOMMENDATION_MIN_CLUSTER_PURITY = 0.70
RECOMMENDATION_TOP4_LOWER_BOUND_MAX_GAP = 0.15
RECOMMENDATION_MIN_OBSERVED_WINS = 1
MATURE_STAGE_MIN_RELIABLE_N_EFF = 5.0
RECOMMENDATION_CRITERION_LABELS: dict[str, str] = {
    "raw_n": "原始n",
    "weighted_n": "加权n",
    "n_eff": "有效n",
    "batch_coverage": "批次覆盖",
    "cluster_purity": "聚类纯度",
    "observed_wins": "吃鸡数",
    "top4_vs_play_style_baseline": "前四稳健下界",
    "normal_cost_ceiling": "常规成型",
}
PRIMARY_BOND_SOURCE_LABELS: dict[str, str] = {
    "study_override": "4学习独占",
    "food_harvest": "收菜归美食社",
    "qualified_bond": "普通羁绊第二档门",
    "high_cost_pdd": "高费拼多多兜底",
}
CEILING_CONDITION_LABELS: dict[str, str] = {
    "level_9_or_higher": "等级>=9",
    "four_five_cost_count": "4/5费>=4张",
    "four_five_cost_share": "4/5费占比>=50%",
    "key_high_cost_two_star": "关键4/5费两星",
    "main_carry_equipment_complete": "主C三件装备完整",
}
CEILING_REPRESENTATIVE_BOARD_LIMIT = 3
HERO_EQUIPMENT_DETAIL_MIN_APPEARANCES = 11  # Strict appearances > 10.
INVERSION_REASON_LABELS: dict[str, str] = {
    "avg_rank_regression": "均分劣于参考成熟盘",
    "top4_regression": "前四率劣于参考成熟盘",
}
MATURE_STAGE_MAX_AVG_RANK_REGRESSION = 0.75
MATURE_STAGE_MAX_TOP4_REGRESSION = 15.0
COMPOSITION_RANK_PRIOR_STRENGTH = 12.0
COMPOSITION_RATE_PRIOR_STRENGTH = 12.0
BETA_LOWER_BOUND_Z = 1.2815515655446004  # One-sided 90% normal approximation.
StatItem = tuple[str, int] | tuple[str, int, float]
CARD_TEMPLATE_DIR = ROOT / "assets" / "templates" / "cards"
MERGED_TEMPLATE_EXPANSIONS: dict[str, list[str]] = {
    "黄·吸吸宝pro快速成型": ["黄·快速成型", "黄·吸吸宝pro"],
    "蓝·重质拍档支援": ["蓝·拍档支援", "蓝·重质也重量pro"],
    "蓝·一起刷刷刷+天降啾啾pro": ["蓝·一起刷刷刷", "蓝·天降啾啾pro"],
    "黄·巨神兵+迅迅迅捷双剑": ["黄·巨神兵", "黄·迅迅迅捷双剑"],
}
LEGACY_CARD_TEMPLATE_NAMES = frozenset(
    {
        "法力专注",
        "蓝·开攒",
        "蓝·大亨",
    }
)

CARD_GRANTED_HEROES = {"暴龙虾饺"}
PLAY_STYLES = ("赌狗", "高费")
CARD_PREFIX_TYPES = ("彩", "黄", "蓝", "白", "其他")
CARD_MERGE_NOTES: dict[str, str] = {
    "蓝": (
        "蓝卡同图标规则："
        "福袋，有钱同享 -> 福袋有钱；"
        "最佳拍档，最强支援 -> 拍档支援；"
        "最后的波纹，利己主义 -> 波纹利己；"
        "开攒，大亨 -> 开攒大亨。"
        "一起刷刷刷与天降啾啾pro虽共用图标，已按最终阵容啾啾装备数量分别统计。"
    ),
    "彩": (
        "以下卡牌因图标完全相同做了合并处理："
        "法师礼包，战士礼包，射手礼包 -> 法师战士射手礼包。"
    ),
    "黄": (
        "以下卡牌因图标完全相同做了合并处理："
        "大力，巫术，守护 -> 大力巫术守护。"
        "巨神兵与迅迅迅捷双剑虽共用图标，已按最终阵容巨神兵之斧/迅捷双剑数量分别统计："
        "仅斧 -> 巨神兵，仅剑 -> 迅迅迅捷双剑，都有则数量占优；"
        "数量相同则按本次数据库明确样本比例并以固定种子可复现分配。"
    ),
}

HERO_ALIASES = {
    "双面教师林野·前排": "双面教师林野",
    "双面教师林野·后排": "双面教师林野",
}


@dataclass
class Equipment:
    raw_name: str
    name: str
    selected: bool


@dataclass
class Hero:
    id: int
    name: str
    canonical_name: str
    slot_index: int
    tier: int | None
    stars: int
    equipment_count: int
    equipments: list[Equipment] = field(default_factory=list)
    traits: list[str] = field(default_factory=list)
    carry_score: float = 0.0

    @property
    def selected_equipment_count(self) -> int:
        return sum(1 for equipment in self.equipments if equipment.selected)


@dataclass
class PlayerFeature:
    player_id: int
    match_id: int
    rank: int
    row_index: int
    partner_player: int | None
    heroes: list[Hero]
    cards: list[str]
    trait_counts: Counter[str]
    jiujiu_bonus: Counter[str]
    trait_totals: Counter[str]
    active_traits: dict[str, int]
    main_bond: str | None
    main_carry: Hero | None
    secondary_carry: Hero | None
    hero_set: set[str]
    level: int
    carry_candidates: list[Hero] = field(default_factory=list)
    family_id: int | None = None
    team_rank: int | None = None
    team_best_rank: int | None = None
    match_batch: str | None = None
    sample_weight: float = 1.0
    # Identity is intentionally separate from active_traits/main_bond.  The
    # latter are factual board state; identity describes how the board is
    # played and is used only by composition grouping/naming.
    archetype: str = "未分类"
    archetype_signals: list[dict[str, Any]] = field(default_factory=list)
    trait_investment: dict[str, Any] = field(default_factory=dict)
    high_cost_structure: dict[str, Any] = field(default_factory=dict)


@dataclass
class RankStats:
    appearances: int = 0
    weighted_appearances: float = 0.0
    weight_square_sum: float = 0.0
    rank_sum: float = 0.0
    wins: float = 0.0
    top2: float = 0.0
    top4: float = 0.0

    def add(self, rank: int, weight: float = 1.0) -> None:
        self.appearances += 1
        self.weighted_appearances += weight
        self.weight_square_sum += weight * weight
        self.rank_sum += rank * weight
        if rank == 1:
            self.wins += weight
        if rank <= 2:
            self.top2 += weight
        if rank <= 4:
            self.top4 += weight

    def to_dict(self, baseline_rank: float | None = None, prior: int = 0) -> dict[str, Any]:
        n = max(self.weighted_appearances, 1e-9)
        n_eff = (
            self.weighted_appearances * self.weighted_appearances / self.weight_square_sum
            if self.weight_square_sum > 1e-12
            else 0.0
        )
        row = {
            "appearances": self.appearances,
            "weighted_appearances": round(self.weighted_appearances, 2),
            "n_eff": round(n_eff, 2),
            "avg_rank": round(self.rank_sum / n, 2),
            "win_rate": round(self.wins * 100.0 / n, 1),
            "top2_rate": round(self.top2 * 100.0 / n, 1),
            "top4_rate": round(self.top4 * 100.0 / n, 1),
        }
        if baseline_rank is not None and prior > 0:
            adjusted = (self.rank_sum + baseline_rank * prior) / (self.weighted_appearances + prior)
            row["adjusted_avg_rank"] = round(adjusted, 2)
        return row


def load_game_config() -> tuple[dict[str, list[Any]], dict[str, list[int]]]:
    from config_s2 import dict_bond, dict_character

    return dict_character, dict_bond


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def unpack_stat_item(item: StatItem) -> tuple[str, int, float]:
    if len(item) == 2:
        return item[0], item[1], 1.0
    return item[0], item[1], item[2]


def infer_batch_date(batch: str | None, reference_date: date | None = None) -> date | None:
    """Map MMDD to its most recent occurrence, avoiding Dec/Jan inversions."""
    if not batch or len(batch) != 4 or not batch.isdigit():
        return None
    reference = reference_date or datetime.now(timezone.utc).date()
    try:
        candidate = date(reference.year, int(batch[:2]), int(batch[2:]))
    except ValueError:
        return None
    if candidate > reference + timedelta(days=1):
        candidate = candidate.replace(year=candidate.year - 1)
    return candidate


def batch_ordinal(batch: str | None, reference_date: date | None = None) -> int:
    inferred = infer_batch_date(batch, reference_date)
    return inferred.toordinal() if inferred else 0


def ordered_batches(
    features: Iterable[PlayerFeature],
    reference_date: date | None = None,
) -> list[str]:
    batches = {feature.match_batch for feature in features if feature.match_batch}
    return sorted(batches, key=lambda batch: batch_ordinal(batch, reference_date))


def compute_sample_weights(
    features: list[PlayerFeature],
    *,
    half_life_days: float = DEFAULT_RECENCY_HALF_LIFE_DAYS,
    min_weight: float = MIN_RECENCY_WEIGHT,
    reference_date: date | None = None,
) -> None:
    batches = ordered_batches(features, reference_date)
    if not batches:
        for feature in features:
            feature.sample_weight = 1.0
        return
    max_ord = batch_ordinal(batches[-1], reference_date)
    decay = 0.6931471805599453 / max(half_life_days, 1e-6)
    for feature in features:
        if not feature.match_batch:
            feature.sample_weight = min_weight
            continue
        days_ago = max(
            max_ord - batch_ordinal(feature.match_batch, reference_date),
            0,
        )
        feature.sample_weight = max(min_weight, 2.718281828459045 ** (-decay * days_ago))


def weighted_totals(features: list[PlayerFeature]) -> tuple[float, float, dict[int, float]]:
    total_weight = sum(feature.sample_weight for feature in features) or 1.0
    match_weights: dict[int, float] = {}
    for feature in features:
        match_weights[feature.match_id] = feature.sample_weight
    total_match_weight = sum(match_weights.values()) or 1.0
    return total_weight, total_match_weight, match_weights


def recency_overview(
    features: list[PlayerFeature],
    half_life_days: float,
    reference_date: date | None = None,
) -> dict[str, Any]:
    batch_counts: Counter[str] = Counter()
    batch_weight: Counter[str] = Counter()
    for feature in features:
        batch = feature.match_batch or "unknown"
        batch_counts[batch] += 1
        batch_weight[batch] += feature.sample_weight
    batches = ordered_batches(features, reference_date)
    return {
        "source": "matches.path -> screenshots.MMDD",
        "half_life_days": half_life_days,
        "min_weight": MIN_RECENCY_WEIGHT,
        "weight_floor_policy": "valid dated batches have no permanent floor",
        "cross_year_inference": "MMDD is mapped to its most recent occurrence relative to analysis date",
        "latest_batch": batches[-1] if batches else None,
        "batch_range": [batches[0], batches[-1]] if batches else [],
        "batch_counts": dict(
            sorted(
                batch_counts.items(),
                key=lambda item: batch_ordinal(item[0], reference_date),
            )
        ),
        "batch_weighted_counts": {
            batch: round(batch_weight[batch], 2)
            for batch in sorted(
                batch_weight,
                key=lambda value: batch_ordinal(value, reference_date),
            )
        },
        "effective_sample_weight": round(sum(feature.sample_weight for feature in features), 2),
    }


def parse_equipment_count(value: Any) -> int:
    if value is None or value == "-":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def normalize_hero_name(name: str) -> str:
    return HERO_ALIASES.get(name, name)


def normalize_equipment_name(name: str) -> tuple[str, bool]:
    if name.startswith("核选"):
        return name[len("核选") :], True
    return name, False


def is_super_equipment(name: str) -> bool:
    normalized, _ = normalize_equipment_name(name)
    return normalized in SUPER_EQUIPMENT_NAMES


def is_food_equipment(name: str) -> bool:
    normalized, _ = normalize_equipment_name(name)
    return normalized.startswith(FOOD_HARVEST_PREFIXES) or normalized in FOOD_SPECIAL_EQUIPMENT_NAMES


def equipment_kind(name: str) -> str:
    normalized, _ = normalize_equipment_name(name)
    if normalized in SUPER_EQUIPMENT_NAMES:
        return "super"
    if normalized.startswith(FOOD_HARVEST_PREFIXES) or normalized in FOOD_SPECIAL_EQUIPMENT_NAMES:
        return "food"
    if normalized.endswith("啾啾"):
        return "jiujiu"
    return "normal"


def card_prefix_type(card_name: str) -> str:
    prefix, _ = split_card_prefix(card_name)
    if prefix:
        return prefix
    return "其他"


def load_report_card_catalog() -> dict[str, list[str]]:
    by_prefix: dict[str, set[str]] = {
        prefix: set() for prefix in CARD_PREFIX_TYPES if prefix != "其他"
    }
    for path in sorted(CARD_TEMPLATE_DIR.glob("*.jpg")):
        if path.name.startswith("player"):
            continue
        raw_name = path.stem
        if raw_name in LEGACY_CARD_TEMPLATE_NAMES:
            continue
        canonical = normalize_card_label(raw_name)
        expansions = MERGED_TEMPLATE_EXPANSIONS.get(raw_name) or MERGED_TEMPLATE_EXPANSIONS.get(
            canonical
        )
        names = expansions if expansions else [canonical]
        for name in names:
            prefix = card_prefix_type(name)
            if prefix in by_prefix:
                by_prefix[prefix].add(name)
    return {prefix: sorted(names) for prefix, names in by_prefix.items()}


def empty_card_row(key: str) -> dict[str, Any]:
    return {
        "key": key,
        "appearances": 0,
        "avg_rank": None,
        "win_rate": None,
        "top4_rate": None,
        "adjusted_avg_rank": None,
    }


def aggregate_key_stats(items: list[StatItem], min_apps: int, baseline: float) -> list[dict[str, Any]]:
    stats: dict[str, RankStats] = defaultdict(RankStats)
    for item in items:
        key, rank, weight = unpack_stat_item(item)
        if key:
            stats[key].add(rank, weight)
    rows = []
    for key, stat in stats.items():
        if stat.appearances >= min_apps:
            rows.append({"key": key, **stat.to_dict(baseline_rank=baseline, prior=8)})
    rows.sort(key=lambda row: (row["adjusted_avg_rank"], row["avg_rank"], -row["top4_rate"]))
    return rows


def aggregate_single_cards_by_catalog(
    items: list[StatItem],
    baseline: float,
    catalog: dict[str, list[str]],
    *,
    sample_first: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    stats: dict[str, RankStats] = defaultdict(RankStats)
    for item in items:
        key, rank, weight = unpack_stat_item(item)
        if key:
            stats[key].add(rank, weight)

    annotated: list[dict[str, Any]] = []
    by_prefix: dict[str, list[dict[str, Any]]] = {}
    for prefix_type in CARD_PREFIX_TYPES:
        if prefix_type == "其他":
            continue
        group_rows: list[dict[str, Any]] = []
        for key in catalog.get(prefix_type, []):
            stat = stats.get(key)
            if stat and stat.appearances > 0:
                group_rows.append({"key": key, **stat.to_dict(baseline_rank=baseline, prior=8)})
            else:
                group_rows.append(empty_card_row(key))

        with_data = [row for row in group_rows if row["appearances"] > 0]
        without_data = [row for row in group_rows if row["appearances"] == 0]
        if sample_first:
            with_data.sort(
                key=lambda row: (
                    -row["appearances"],
                    row["adjusted_avg_rank"],
                    row["avg_rank"],
                    -row["top4_rate"],
                )
            )
        else:
            with_data.sort(
                key=lambda row: (row["adjusted_avg_rank"], row["avg_rank"], -row["top4_rate"])
            )
        without_data.sort(key=lambda row: row["key"])
        ordered_rows = with_data + without_data

        ranked_rows: list[dict[str, Any]] = []
        rank = 1
        for row in ordered_rows:
            ranked_row = {
                **row,
                "prefix_type": prefix_type,
                "prefix_rank": rank if row["appearances"] > 0 else None,
            }
            if row["appearances"] > 0:
                rank += 1
            ranked_rows.append(ranked_row)
            annotated.append(ranked_row)
        if ranked_rows:
            by_prefix[prefix_type] = ranked_rows
    return annotated, by_prefix


def aggregate_key_stats_by_prefix(
    items: list[StatItem],
    min_apps: int,
    baseline: float,
    *,
    sample_first: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    rows = aggregate_key_stats(items, min_apps, baseline)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[card_prefix_type(row["key"])].append(row)

    annotated: list[dict[str, Any]] = []
    by_prefix: dict[str, list[dict[str, Any]]] = {}
    for prefix_type in CARD_PREFIX_TYPES:
        group_rows = grouped.get(prefix_type, [])
        if sample_first:
            group_rows.sort(
                key=lambda row: (
                    -row["appearances"],
                    row["adjusted_avg_rank"],
                    row["avg_rank"],
                    -row["top4_rate"],
                )
            )
        else:
            group_rows.sort(
                key=lambda row: (row["adjusted_avg_rank"], row["avg_rank"], -row["top4_rate"])
            )
        ranked_rows: list[dict[str, Any]] = []
        for rank, row in enumerate(group_rows, start=1):
            ranked_row = {**row, "prefix_type": prefix_type, "prefix_rank": rank}
            ranked_rows.append(ranked_row)
            annotated.append(ranked_row)
        if ranked_rows:
            by_prefix[prefix_type] = ranked_rows
    return annotated, by_prefix


def add_avg_appearances_per_match(
    rows: list[dict[str, Any]],
    total_matches: int,
) -> list[dict[str, Any]]:
    denominator = max(total_matches, 1)
    for row in rows:
        row["avg_appearances_per_match"] = round(row["appearances"] / denominator, 2)
    return rows


def add_avg_appearances_to_prefix_groups(
    groups: dict[str, list[dict[str, Any]]],
    total_matches: int,
) -> dict[str, list[dict[str, Any]]]:
    for rows in groups.values():
        add_avg_appearances_per_match(rows, total_matches)
    return groups


def jiujiu_trait(equipment_name: str) -> str | None:
    normalized, _ = normalize_equipment_name(equipment_name)
    if normalized.endswith("啾啾"):
        return normalized[: -len("啾啾")]
    return None


def active_tier(count: int, thresholds: list[int]) -> int:
    tier = 0
    for threshold in sorted(thresholds):
        if count >= threshold:
            tier = threshold
    return tier


def confidence_label(n: int, unknown_rate: float = 0.0) -> str:
    if n >= 30 and unknown_rate <= 0.03:
        return "高"
    if n >= 10 and unknown_rate <= 0.08:
        return "中"
    return "低"


def effective_sample_size(weights: Iterable[float]) -> float:
    """Return Kish's effective sample size for arbitrary non-negative weights."""
    values = [max(float(weight), 0.0) for weight in weights]
    total = sum(values)
    squares = sum(weight * weight for weight in values)
    return total * total / squares if squares > 1e-12 else 0.0


def beta_posterior_summary(
    observed_rate: float,
    n_eff: float,
    baseline_rate: float,
    *,
    prior_strength: float = COMPOSITION_RATE_PRIOR_STRENGTH,
) -> dict[str, float]:
    """Shrink a rate to its play-style prior and expose a conservative bound."""
    rate = min(max(observed_rate, 0.0), 1.0)
    prior_rate = min(max(baseline_rate, 0.0), 1.0)
    effective_n = max(n_eff, 0.0)
    alpha = max(prior_rate * prior_strength + rate * effective_n, 1e-9)
    beta = max((1.0 - prior_rate) * prior_strength + (1.0 - rate) * effective_n, 1e-9)
    posterior_mean = alpha / (alpha + beta)
    posterior_variance = alpha * beta / (
        (alpha + beta) ** 2 * (alpha + beta + 1.0)
    )
    lower_bound = max(0.0, posterior_mean - BETA_LOWER_BOUND_Z * posterior_variance**0.5)
    return {
        "posterior_mean": round(posterior_mean, 4),
        "lower_bound": round(lower_bound, 4),
        "prior_rate": round(prior_rate, 4),
        "prior_strength": prior_strength,
    }


def composition_baselines(features: list[PlayerFeature]) -> dict[str, dict[str, float]]:
    """Build weighted empirical-Bayes priors separately for each play style."""
    grouped: dict[str, RankStats] = defaultdict(RankStats)
    for feature in features:
        grouped["all"].add(feature.rank, feature.sample_weight)
        grouped[classify_play_style(feature)].add(feature.rank, feature.sample_weight)

    def baseline(stats: RankStats) -> dict[str, float]:
        if not stats.appearances:
            return {"avg_rank": 4.5, "top4_rate": 0.5, "win_rate": 0.125}
        weight = max(stats.weighted_appearances, 1e-9)
        return {
            "avg_rank": stats.rank_sum / weight,
            "top4_rate": stats.top4 / weight,
            "win_rate": stats.wins / weight,
        }

    all_baseline = baseline(grouped["all"])
    return {
        style: baseline(grouped[style]) if grouped[style].appearances else all_baseline
        for style in (*PLAY_STYLES, "all")
    }


def build_confidence_evidence(
    row: dict[str, Any],
    members: list[PlayerFeature],
    baseline: dict[str, float] | None = None,
    *,
    allow_high_cost_ceiling: bool = False,
) -> dict[str, Any]:
    """Keep recommendation eligibility separate from composition discovery."""
    stats = row["stats"]
    raw_n = int(stats.get("appearances", len(members)))
    weighted_n = float(stats.get("weighted_appearances", 0.0))
    n_eff = float(stats.get("n_eff", effective_sample_size(member.sample_weight for member in members)))
    batch_count = len({member.match_batch for member in members if member.match_batch})
    archetypes = row.get("archetype_distribution", [])
    purity = (
        max((float(item.get("share", 0.0)) for item in archetypes), default=0.0) / 100.0
    )
    prior = baseline or {"avg_rank": 4.5, "top4_rate": 0.5, "win_rate": 0.125}
    observed_top4 = float(stats.get("top4_rate", 0.0)) / 100.0
    observed_win = float(stats.get("win_rate", 0.0)) / 100.0
    observed_avg = float(stats.get("avg_rank", prior["avg_rank"]))
    rank_prior = COMPOSITION_RANK_PRIOR_STRENGTH
    shrunk_avg_rank = (
        observed_avg * n_eff + prior["avg_rank"] * rank_prior
    ) / max(n_eff + rank_prior, 1e-9)
    top4_summary = beta_posterior_summary(observed_top4, n_eff, prior["top4_rate"])
    win_summary = beta_posterior_summary(observed_win, n_eff, prior["win_rate"])
    observed_wins = sum(member.rank == 1 for member in members)
    top4_floor = max(
        0.0, prior["top4_rate"] - RECOMMENDATION_TOP4_LOWER_BOUND_MAX_GAP
    )
    criteria = {
        "raw_n": {"value": raw_n, "minimum": RECOMMENDATION_MIN_RAW_N, "met": raw_n >= RECOMMENDATION_MIN_RAW_N},
        "weighted_n": {"value": round(weighted_n, 2), "minimum": RECOMMENDATION_MIN_WEIGHTED_N, "met": weighted_n >= RECOMMENDATION_MIN_WEIGHTED_N},
        "n_eff": {"value": round(n_eff, 2), "minimum": RECOMMENDATION_MIN_EFFECTIVE_N, "met": n_eff >= RECOMMENDATION_MIN_EFFECTIVE_N},
        "batch_coverage": {"value": batch_count, "minimum": RECOMMENDATION_MIN_BATCHES, "met": batch_count >= RECOMMENDATION_MIN_BATCHES},
        "cluster_purity": {"value": round(purity, 3), "minimum": RECOMMENDATION_MIN_CLUSTER_PURITY, "met": purity >= RECOMMENDATION_MIN_CLUSTER_PURITY},
        "observed_wins": {
            "value": observed_wins,
            "minimum": RECOMMENDATION_MIN_OBSERVED_WINS,
            "met": observed_wins >= RECOMMENDATION_MIN_OBSERVED_WINS,
        },
        "top4_vs_play_style_baseline": {
            "value": top4_summary["lower_bound"],
            "minimum": round(top4_floor, 4),
            "baseline": round(prior["top4_rate"], 4),
            "maximum_gap": RECOMMENDATION_TOP4_LOWER_BOUND_MAX_GAP,
            "metric": "shrunk_top4_90pct_lower_bound",
            "met": top4_summary["lower_bound"] >= top4_floor,
        },
        "normal_cost_ceiling": {
            "value": (
                True
                if allow_high_cost_ceiling
                else not bool(row.get("high_cost_three_star_dependency"))
            ),
            "required": True,
            "met": (
                True
                if allow_high_cost_ceiling
                else not bool(row.get("high_cost_three_star_dependency"))
            ),
            "ceiling_exception": allow_high_cost_ceiling,
        },
    }
    failure_reasons = [
        {
            "criterion": name,
            "value": criterion.get("value"),
            "required": criterion.get("minimum", criterion.get("required")),
        }
        for name, criterion in criteria.items()
        if not criterion["met"]
    ]
    return {
        "discovery_min_apps": 5,
        "recommendation_eligible": all(item["met"] for item in criteria.values()),
        "recommendation_criteria": criteria,
        "recommendation_failure_reasons": failure_reasons,
        "raw_n": raw_n,
        "weighted_n": round(weighted_n, 2),
        "n_eff": round(n_eff, 2),
        "batch_coverage": batch_count,
        "cluster_purity": round(purity, 3),
        "baseline": {
            "play_style": row.get("play_style", "高费"),
            "avg_rank": round(prior["avg_rank"], 4),
            "top4_rate": round(prior["top4_rate"], 4),
            "win_rate": round(prior["win_rate"], 4),
        },
        "shrunk_metrics": {
            "avg_rank": round(shrunk_avg_rank, 4),
            "rank_prior_strength": rank_prior,
            "top4": top4_summary,
            "win": win_summary,
        },
    }


def level_label(level: int) -> int:
    if level >= 9:
        return 9
    if level >= 8:
        return 8
    return 7


def is_lineup_hero(hero_name: str) -> bool:
    return hero_name not in CARD_GRANTED_HEROES


def first_card(feature: PlayerFeature) -> str | None:
    return feature.cards[0] if feature.cards else None


def team_rank_value(feature: PlayerFeature) -> int:
    return feature.team_rank if feature.team_rank is not None else feature.rank


def is_low_cost_hero(hero: Hero | None) -> bool:
    return bool(hero and hero.tier is not None and hero.tier <= 3)


def is_low_cost_three_star(hero: Hero) -> bool:
    return is_low_cost_hero(hero) and hero.stars >= 3 and is_lineup_hero(hero.name)


def food_harvest_evidence(heroes: list[Hero]) -> list[dict[str, Any]]:
    """Return normalized equipment evidence for the food-club harvest mechanic."""
    evidence = []
    for hero in heroes:
        for equipment in hero.equipments:
            normalized, _ = normalize_equipment_name(equipment.raw_name)
            if normalized.startswith(FOOD_HARVEST_PREFIXES):
                evidence.append(
                    {
                        "hero_name": hero.name,
                        "equipment_name": normalized,
                        "raw_equipment_name": equipment.raw_name,
                    }
                )
    return sorted(evidence, key=lambda item: (item["hero_name"], item["equipment_name"]))


def analyze_trait_investment(
    heroes: list[Hero],
    active_traits: dict[str, int],
    trait_counts: Counter[str],
    dict_bond: dict[str, list[int]],
    main_carry: Hero | None,
) -> dict[str, Any]:
    """Score real trait commitment without changing factual active traits."""
    lineup = [hero for hero in heroes if is_lineup_hero(hero.name)]
    lineup_count = max(len(lineup), 1)
    low_cost_three_stars = [hero for hero in lineup if is_low_cost_three_star(hero)]
    main_carry_low_cost_three_star = bool(
        main_carry and is_low_cost_three_star(main_carry)
    )
    # Three or more shallow activations on a full board are a scattered
    # structure, not three independent main-trait investments.  A low-cost
    # 3-star main carry is the deliberate exception: it is concrete board
    # investment and can validate that carry's otherwise shallow trait.
    scattered_structure = len(active_traits) >= 3 and all(
        trait_counts[trait] / lineup_count <= 0.5 for trait in active_traits
    )
    rows = []
    for trait, tier in sorted(active_traits.items()):
        thresholds = sorted(dict_bond.get(trait, []))
        if not thresholds:
            continue
        depth_index = max((index for index, value in enumerate(thresholds) if value <= tier), default=0)
        depth = (depth_index + 1) / len(thresholds)
        coverage = trait_counts[trait] / lineup_count
        carry_aligned = bool(main_carry and trait in main_carry.traits)
        low_cost_three_star_carry_aligned = bool(
            main_carry_low_cost_three_star and carry_aligned
        )
        # Depth is the strongest signal; coverage and carry alignment stop an
        # incidental first breakpoint from masquerading as a main investment.
        # A 3-star 1/2/3-cost main carry is an observable commitment, while
        # scattered first-tier boards receive a small counterweight.
        score = (
            depth * 0.5
            + coverage * 0.3
            + (0.2 if carry_aligned else 0.0)
            + (0.15 if low_cost_three_star_carry_aligned else 0.0)
            - (0.1 if scattered_structure and not low_cost_three_star_carry_aligned else 0.0)
        )
        stable = (
            score >= 0.55
            and (
                depth >= 0.5
                or (coverage >= 0.5 and carry_aligned)
                or low_cost_three_star_carry_aligned
            )
        )
        rows.append(
            {
                "trait": trait,
                "active_tier": tier,
                "count": trait_counts[trait],
                "depth": round(depth, 3),
                "coverage": round(coverage, 3),
                "carry_aligned": carry_aligned,
                "low_cost_three_star_carry_aligned": low_cost_three_star_carry_aligned,
                "score": round(max(score, 0.0), 3),
                "stable": stable,
            }
        )
    rows.sort(key=lambda item: (item["stable"], item["score"], item["trait"]), reverse=True)
    stable_rows = [item for item in rows if item["stable"]]
    return {
        "traits": rows,
        "stable_traits": [item["trait"] for item in stable_rows],
        "dominant_trait": stable_rows[0]["trait"] if stable_rows else None,
        "scattered_active_traits": len(active_traits),
        "low_cost_three_star_count": len(low_cost_three_stars),
        "main_carry_low_cost_three_star": main_carry_low_cost_three_star,
        "scattered_structure": scattered_structure,
    }


def classify_archetype(
    heroes: list[Hero],
    food_evidence: list[dict[str, Any]],
    trait_investment: dict[str, Any],
    play_style: str,
    main_carry: Hero | None = None,
) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    """Classify gameplay identity, retaining the evidence needed to audit it."""
    lineup = [hero for hero in heroes if is_lineup_hero(hero.name)]
    high_cost = [hero for hero in lineup if hero.tier is not None and hero.tier >= 4]
    low_cost_three_star = [hero for hero in lineup if is_low_cost_three_star(hero)]
    high_cost_floor = max(2, round(len(lineup) * 0.3))
    main_carry_tier = main_carry.tier if main_carry else None
    main_carry_stars = main_carry.stars if main_carry else None
    main_carry_is_high_cost_two_star = bool(
        main_carry
        and main_carry.tier is not None
        and main_carry.tier >= 4
        and main_carry.stars >= 2
        and is_lineup_hero(main_carry.name)
    )
    structure = {
        "four_five_cost_count": len(high_cost),
        "four_five_cost_share": round(len(high_cost) / max(len(lineup), 1), 3),
        "low_cost_three_star_count": len(low_cost_three_star),
        "main_carry_tier": main_carry_tier,
        "main_carry_stars": main_carry_stars,
        "main_carry_is_high_cost_two_star": main_carry_is_high_cost_two_star,
        "scattered_active_traits": trait_investment["scattered_active_traits"],
        "stable_traits": trait_investment["stable_traits"],
        "high_cost_threshold": high_cost_floor,
    }
    if food_evidence:
        return (
            "美食社收菜",
            [{"type": "美食装备", "strength": "强", "equipment": item} for item in food_evidence],
            structure,
        )
    # True high-cost PDD: no low-cost 3-star units, enough 4/5-cost presence,
    # no stable deep trait investment, and a 2-star+ 4/5-cost main carry.
    if (
        play_style == "高费"
        and not low_cost_three_star
        and main_carry_is_high_cost_two_star
        and len(high_cost) >= high_cost_floor
        and not trait_investment["stable_traits"]
    ):
        return (
            "高费拼多多",
            [
                {
                    "type": "高费散羁绊",
                    "strength": "强",
                    "four_five_cost_count": len(high_cost),
                    "threshold": high_cost_floor,
                    "stable_traits": [],
                    "low_cost_three_star_count": 0,
                    "main_carry_tier": main_carry_tier,
                    "main_carry_stars": main_carry_stars,
                }
            ],
            structure,
        )
    dominant = trait_investment["dominant_trait"]
    if dominant:
        return (
            f"羁绊运营:{dominant}",
            [{"type": "羁绊投入", "strength": "中", "trait": dominant}],
            structure,
        )
    return (
        "拼多多",
        [{"type": "散羁绊", "strength": "中", "active_trait_count": len(trait_investment["traits"])}],
        structure,
    )


def classify_play_style(feature: PlayerFeature) -> str:
    """Assign a board to 赌狗 or 高费.

    Any lineup 1/2/3-cost 3-star unit forces 赌狗. Only boards without low-cost
    3-stars can enter 高费; level-7 boards with a low-cost main carry are also
    赌狗 even when that carry is not yet 3-star.
    """
    lineup_count = len(unique_heroes_by_slot(feature))
    main_carry = feature.main_carry
    has_low_cost_three_star = any(
        is_low_cost_three_star(hero) for hero in feature.heroes if is_lineup_hero(hero.name)
    )

    if lineup_count <= 6:
        return "赌狗"
    if has_low_cost_three_star:
        return "赌狗"
    if feature.level == 7 and is_low_cost_hero(main_carry):
        return "赌狗"
    return "高费"


def has_low_cost_three_star_carry_requirement(row: dict[str, Any]) -> bool:
    """True when mature carry advice requires a 1/2/3-cost unit at 3 stars."""
    for item in row.get("carry_requirements", []):
        tier = item.get("tier")
        if tier is None:
            continue
        if int(tier) <= 3 and int(item.get("recommended_min_stars", 0) or 0) >= 3:
            return True
    return False


def resolve_strategy_play_style(row: dict[str, Any]) -> str:
    """Final recommendation bucket follows mature-stage style plus carry gates."""
    mature = row.get("mature_stage") or {}
    style = mature.get("play_style") or row.get("play_style") or "高费"
    requirements = mature.get("carry_requirements") or row.get("carry_requirements") or []
    probe = {"carry_requirements": requirements}
    if has_low_cost_three_star_carry_requirement(probe):
        return "赌狗"
    if style not in PLAY_STYLES:
        return "高费"
    return style


def three_star_lineup_count(feature: PlayerFeature) -> int:
    return len(
        {
            hero.name
            for hero in feature.heroes
            if is_lineup_hero(hero.name) and hero.stars >= 3
        }
    )


def play_style_summary(members: list[PlayerFeature]) -> tuple[str, list[dict[str, Any]]]:
    counts = Counter(classify_play_style(member) for member in members)
    total = len(members) or 1
    breakdown = [
        {
            "play_style": style,
            "appearances": counts.get(style, 0),
            "share": round(counts.get(style, 0) * 100.0 / total, 1),
        }
        for style in PLAY_STYLES
        if counts.get(style, 0) > 0
    ]
    primary = max(PLAY_STYLES, key=lambda style: (counts.get(style, 0), -PLAY_STYLES.index(style)))
    return primary, breakdown


def find_latest_db(explicit: Path | None) -> Path:
    if explicit is not None:
        db_path = explicit if explicit.is_absolute() else ROOT / explicit
        if not db_path.exists():
            raise SystemExit(f"DB not found: {db_path}")
        return db_path

    if DEFAULT_LATEST_DB.exists():
        return DEFAULT_LATEST_DB

    candidates = sorted(
        (ROOT / "data").glob("matches_*.db"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise SystemExit(
            "No data/match_latest.db or data/matches_*.db found. "
            "Build or provide the latest DB with --db."
        )
    return candidates[0]


def find_bot_player_ids(conn: sqlite3.Connection) -> set[int]:
    rows = conn.execute(
        """
        SELECT p7.id AS p7_id, p8.id AS p8_id
        FROM players p7
        JOIN players p8 ON p8.match_id = p7.match_id AND p8.rank = 8
        WHERE p7.rank = 7
          AND (
            p7.partner_player = 8
            OR p8.partner_player = 7
            OR EXISTS (
              SELECT 1
              FROM pairs pair
              WHERE pair.match_id = p7.match_id
                AND (
                  (pair.player_a = 7 AND pair.player_b = 8)
                  OR (pair.player_a = 8 AND pair.player_b = 7)
                )
            )
          )
        """
    ).fetchall()
    bot_ids: set[int] = set()
    for row in rows:
        bot_ids.add(int(row["p7_id"]))
        bot_ids.add(int(row["p8_id"]))
    return bot_ids


def db_count(conn: sqlite3.Connection, table: str) -> int:
    return int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])


def data_quality(conn: sqlite3.Connection, bot_ids: set[int]) -> dict[str, Any]:
    unknown_heroes = conn.execute(
        "SELECT COUNT(*) FROM heroes WHERE hero_name = 'unknown'"
    ).fetchone()[0]
    unknown_cards = conn.execute(
        "SELECT COUNT(*) FROM cards WHERE card_name = 'unknown'"
    ).fetchone()[0]
    unknown_equipment = conn.execute(
        "SELECT COUNT(*) FROM hero_equipments WHERE equipment_name = 'unknown'"
    ).fetchone()[0]
    card_granted_heroes = conn.execute(
        "SELECT COUNT(*) FROM heroes WHERE hero_name IN ({})".format(
            ",".join("?" for _ in CARD_GRANTED_HEROES)
        ),
        tuple(CARD_GRANTED_HEROES),
    ).fetchone()[0]
    return {
        "matches": db_count(conn, "matches"),
        "players": db_count(conn, "players"),
        "heroes": db_count(conn, "heroes"),
        "hero_equipments": db_count(conn, "hero_equipments"),
        "cards": db_count(conn, "cards"),
        "unknown_heroes": int(unknown_heroes),
        "unknown_cards": int(unknown_cards),
        "unknown_equipment": int(unknown_equipment),
        "card_granted_heroes": int(card_granted_heroes),
        "bot_player_records_excluded": len(bot_ids),
        "seven_eight_bot_matches": len(bot_ids) // 2,
    }


def validate_config(
    conn: sqlite3.Connection,
    dict_character: dict[str, list[Any]],
    dict_bond: dict[str, list[int]],
) -> dict[str, Any]:
    db_heroes = [
        row[0]
        for row in conn.execute(
            "SELECT DISTINCT hero_name FROM heroes WHERE hero_name != 'unknown'"
        )
    ]
    missing = [
        hero
        for hero in sorted(db_heroes)
        if normalize_hero_name(hero) not in dict_character
        and hero not in CARD_GRANTED_HEROES
    ]
    jiujiu_items = [
        row[0]
        for row in conn.execute(
            """
            SELECT DISTINCT equipment_name
            FROM hero_equipments
            WHERE equipment_name LIKE '%啾啾%' AND equipment_name != 'unknown'
            """
        )
    ]
    unmapped_jiujiu = [
        name for name in jiujiu_items if jiujiu_trait(name) not in dict_bond
    ]
    return {
        "db_hero_count": len(db_heroes),
        "missing_config_heroes": missing,
        "card_granted_heroes": sorted(CARD_GRANTED_HEROES & set(db_heroes)),
        "config_heroes_not_seen": sorted(
            name
            for name in dict_character
            if name not in {normalize_hero_name(hero) for hero in db_heroes}
        ),
        "jiujiu_equipment_seen": sorted(jiujiu_items),
        "jiujiu_unmapped": sorted(unmapped_jiujiu),
    }


def load_player_features(
    conn: sqlite3.Connection,
    bot_ids: set[int],
    dict_character: dict[str, list[Any]],
    dict_bond: dict[str, list[int]],
) -> list[PlayerFeature]:
    ensure_match_schema(conn)
    conn.row_factory = sqlite3.Row
    match_meta = {
        int(row["id"]): {
            "path": row["path"],
            "match_date": row["match_date"],
        }
        for row in conn.execute("SELECT id, path, match_date FROM matches").fetchall()
    }
    player_rows = conn.execute("SELECT * FROM players ORDER BY match_id, rank").fetchall()
    kept_player_ids = {int(row["id"]) for row in player_rows if int(row["id"]) not in bot_ids}

    heroes_by_player: dict[int, list[Hero]] = defaultdict(list)
    hero_by_id: dict[int, Hero] = {}
    hero_rows = conn.execute(
        """
        SELECT h.*, he.equipment_name
        FROM heroes h
        LEFT JOIN hero_equipments he ON he.hero_id = h.id
        WHERE h.player_id IN ({})
        ORDER BY h.player_id, h.slot_index, he.item_index
        """.format(",".join("?" for _ in kept_player_ids) or "NULL"),
        tuple(kept_player_ids),
    ).fetchall()

    for row in hero_rows:
        player_id = int(row["player_id"])
        raw_name = str(row["hero_name"])
        if raw_name == "unknown":
            continue
        hero_id = int(row["id"])
        if hero_id not in hero_by_id:
            canonical = normalize_hero_name(raw_name)
            config_entry = dict_character.get(canonical)
            config_tier = int(config_entry[0]) if config_entry else row["tier"]
            traits = [str(trait) for trait in config_entry[1:]] if config_entry else []
            hero = Hero(
                id=hero_id,
                name=raw_name,
                canonical_name=canonical,
                slot_index=int(row["slot_index"]),
                tier=int(config_tier) if config_tier is not None else None,
                stars=int(row["stars"] or 0),
                equipment_count=parse_equipment_count(row["equipment_count"]),
                traits=traits,
            )
            hero_by_id[hero_id] = hero
            heroes_by_player[player_id].append(hero)
        equipment_name = row["equipment_name"]
        if equipment_name and equipment_name != "unknown":
            normalized_name, selected = normalize_equipment_name(str(equipment_name))
            hero_by_id[hero_id].equipments.append(
                Equipment(raw_name=str(equipment_name), name=normalized_name, selected=selected)
            )

    cards_by_player: dict[int, list[str]] = defaultdict(list)
    card_rows = conn.execute(
        """
        SELECT player_id, card_name, slot_index
        FROM cards
        WHERE player_id IN ({})
        ORDER BY player_id, slot_index
        """.format(",".join("?" for _ in kept_player_ids) or "NULL"),
        tuple(kept_player_ids),
    ).fetchall()
    resolve_items: list[dict[str, Any]] = []
    resolve_player_ids: list[int] = []
    for row in card_rows:
        card_name = str(row["card_name"])
        if card_name == "unknown":
            continue
        player_id = int(row["player_id"])
        slot_index = int(row["slot_index"])
        hero_context = [
            {
                "stars": hero.stars,
                "equipments": [equipment.raw_name for equipment in hero.equipments],
            }
            for hero in heroes_by_player.get(player_id, [])
        ]
        resolve_items.append(
            {
                "label": card_name,
                "slot_index": slot_index,
                "heroes": hero_context,
            }
        )
        resolve_player_ids.append(player_id)
    for player_id, resolved_name in zip(
        resolve_player_ids,
        resolve_jsb_xj_card_labels(resolve_items),
        strict=True,
    ):
        cards_by_player[player_id].append(resolved_name)

    features: list[PlayerFeature] = []
    for player in player_rows:
        player_id = int(player["id"])
        if player_id in bot_ids:
            continue
        heroes = heroes_by_player.get(player_id, [])
        for hero in heroes:
            tier_score = hero.tier or 0
            hero.carry_score = (
                hero.equipment_count * 30
                + hero.selected_equipment_count * 12
                + hero.stars * 10
                + tier_score * 2
                + max(0, 8 - hero.slot_index) * 1.5
            )

        trait_counts: Counter[str] = Counter()
        jiujiu_bonus: Counter[str] = Counter()
        for hero in heroes:
            for trait in hero.traits:
                if trait in dict_bond:
                    trait_counts[trait] += 1
            for equipment in hero.equipments:
                trait = jiujiu_trait(equipment.raw_name)
                if trait in dict_bond:
                    jiujiu_bonus[trait] += 1
        trait_totals = trait_counts + jiujiu_bonus
        active_traits = {
            trait: active_tier(count, dict_bond[trait])
            for trait, count in trait_totals.items()
            if trait in dict_bond and active_tier(count, dict_bond[trait]) > 0
        }
        main_bond = None
        if active_traits:
            main_bond = max(
                active_traits,
                key=lambda trait: (active_traits[trait], trait_totals[trait], trait),
            )
        carries = sorted(
            heroes,
            key=lambda hero: (hero.carry_score, -hero.slot_index, hero.name),
            reverse=True,
        )
        carry_candidates = carries[:3]
        match_id = int(player["match_id"])
        meta = match_meta.get(match_id, {})
        match_batch = meta.get("match_date") or parse_match_batch(meta.get("path"))
        feature = PlayerFeature(
                player_id=player_id,
                match_id=match_id,
                rank=int(player["rank"]),
                row_index=int(player["row_index"]),
                partner_player=player["partner_player"],
                heroes=heroes,
                cards=cards_by_player.get(player_id, []),
                trait_counts=trait_counts,
                jiujiu_bonus=jiujiu_bonus,
                trait_totals=trait_totals,
                active_traits=active_traits,
                main_bond=main_bond,
                main_carry=carry_candidates[0] if carry_candidates else None,
                secondary_carry=carry_candidates[1] if len(carry_candidates) > 1 else None,
                carry_candidates=carry_candidates,
                hero_set={hero.name for hero in heroes if is_lineup_hero(hero.name)},
                level=level_label(sum(1 for hero in heroes if is_lineup_hero(hero.name))),
                match_batch=match_batch,
            )
        food_evidence = food_harvest_evidence(heroes)
        trait_investment = analyze_trait_investment(
            heroes,
            active_traits,
            trait_counts,
            dict_bond,
            feature.main_carry,
        )
        feature.trait_investment = trait_investment
        feature.archetype, feature.archetype_signals, feature.high_cost_structure = classify_archetype(
            heroes,
            food_evidence,
            trait_investment,
            classify_play_style(feature),
            feature.main_carry,
        )
        features.append(feature)
    assign_team_ranks(features)
    return features


def primary_bond_business_selections(
    feature: PlayerFeature,
    bond_thresholds: dict[str, list[int]] | None = None,
) -> list[dict[str, Any]]:
    """Classify a final board for the business-level primary-bond ranking.

    This is deliberately separate from ``PlayerFeature.main_bond``: the latter
    remains the factual leading active trait used by composition analysis.

    Priority:
    1. Study club at configured tier-4 threshold (exclusive; covers food harvest
       and every other business category)
    2. Food-harvest boards -> 美食社
    3. Qualified factual bonds at the second configured threshold (ties retained)
    4. High-cost PDD fallback
    """
    if bond_thresholds is None:
        _, bond_thresholds = load_game_config()

    study_thresholds = bond_thresholds.get("学习社", [])
    study_count = int(feature.trait_totals.get("学习社", 0))
    study_tier4_min = study_thresholds[2] if len(study_thresholds) > 2 else 4
    if study_count >= study_tier4_min:
        return [
            {
                "bond": "学习社",
                "category": "学习社",
                "source": "study_override",
                "activation_count": study_count,
                "active_tier": active_tier(study_count, study_thresholds),
            }
        ]

    has_food_harvest = (
        feature.archetype == "美食社收菜"
        or bool(food_harvest_evidence(feature.heroes))
    )
    if has_food_harvest:
        activation_count = int(feature.trait_totals.get("美食社", 0))
        return [
            {
                "bond": "美食社",
                "category": "美食社",
                "source": "food_harvest",
                "activation_count": activation_count,
                "active_tier": active_tier(
                    activation_count, bond_thresholds.get("美食社", [])
                ),
            }
        ]

    qualified: list[dict[str, Any]] = []
    for trait in feature.active_traits:
        thresholds = bond_thresholds.get(trait, [])
        activation_count = int(feature.trait_totals.get(trait, 0))
        if len(thresholds) < 2 or activation_count < thresholds[1]:
            continue
        qualified.append(
            {
                "bond": trait,
                "category": trait,
                "source": "qualified_bond",
                "activation_count": activation_count,
                "active_tier": active_tier(activation_count, thresholds),
            }
        )
    if qualified:
        max_count = max(item["activation_count"] for item in qualified)
        return [
            item for item in qualified
            if item["activation_count"] == max_count
        ]

    if feature.archetype == "高费拼多多":
        return [
            {
                "bond": "高费拼多多",
                "category": "高费拼多多",
                "source": "high_cost_pdd",
                "activation_count": 0,
                "active_tier": 0,
            }
        ]
    return []


def primary_bonds_by_count(feature: PlayerFeature) -> list[tuple[str, int, int]]:
    """Return business primary bonds as (category, count, active tier).

    Study club at the configured tier-4 threshold exclusively maps to 学习社.
    Otherwise food-harvest boards map to 美食社. Normal factual traits must
    reach their configured second threshold; ties at the highest qualifying
    activation count are all retained. High-cost PDD is a final fallback.
    """
    return [
        (item["category"], item["activation_count"], item["active_tier"])
        for item in primary_bond_business_selections(feature)
    ]


def assign_team_ranks(features: list[PlayerFeature]) -> None:
    by_match_rank = {
        (feature.match_id, feature.rank): feature
        for feature in features
    }
    features_by_match: dict[int, list[PlayerFeature]] = defaultdict(list)
    for feature in features:
        features_by_match[feature.match_id].append(feature)

    for match_id, match_features in features_by_match.items():
        seen: set[int] = set()
        teams: list[list[PlayerFeature]] = []
        for feature in sorted(match_features, key=lambda item: item.rank):
            if feature.player_id in seen:
                continue
            members = [feature]
            seen.add(feature.player_id)
            if feature.partner_player is not None:
                partner = by_match_rank.get((match_id, int(feature.partner_player)))
                if partner is not None and partner.player_id not in seen:
                    members.append(partner)
                    seen.add(partner.player_id)
            teams.append(members)

        teams.sort(key=lambda members: min(member.rank for member in members))
        for team_rank, members in enumerate(teams, start=1):
            team_best_rank = min(member.rank for member in members)
            for member in members:
                member.team_rank = team_rank
                member.team_best_rank = team_best_rank


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def feature_core_carries(feature: PlayerFeature) -> set[str]:
    return {hero.name for hero in feature.carry_candidates[:2]}


def feature_identity_anchor(feature: PlayerFeature) -> str:
    if feature.archetype in {"美食社收菜", "高费拼多多", "拼多多"}:
        return feature.archetype
    return feature.trait_investment.get("dominant_trait") or feature.main_bond or feature.archetype


def compatible_feature_identities(left: PlayerFeature, right: PlayerFeature) -> bool:
    left_anchor = feature_identity_anchor(left)
    right_anchor = feature_identity_anchor(right)
    if left_anchor == right_anchor:
        return True
    # A normal trait stage can transition through a low-investment form, but
    # never through a special gameplay archetype.
    special = {"美食社收菜", "高费拼多多"}
    if left_anchor in special or right_anchor in special:
        return False
    return bool(
        set(left.trait_investment.get("stable_traits", []))
        & set(right.trait_investment.get("stable_traits", []))
    )


def features_are_similar(left: PlayerFeature, right: PlayerFeature) -> bool:
    overlap = jaccard(left.hero_set, right.hero_set)
    shared_carries = feature_core_carries(left) & feature_core_carries(right)
    if not compatible_feature_identities(left, right):
        return False
    if left.archetype == right.archetype == "高费拼多多":
        # High-cost PDD has no trait anchor to absorb structural variance.
        # Keep only clearly continuous final-board shapes together: neither a
        # shared archetype nor a flexible temporary carry is sufficient.
        left_high_cost = {
            hero.name for hero in left.heroes if hero.tier is not None and hero.tier >= 4
        }
        right_high_cost = {
            hero.name for hero in right.heroes if hero.tier is not None and hero.tier >= 4
        }
        return (
            overlap >= 0.67
            and jaccard(left_high_cost, right_high_cost) >= 0.60
            and abs(len(left_high_cost) - len(right_high_cost)) <= 1
        )
    # Carry overlap alone was the former source of accidental merges.  It now
    # only strengthens a materially similar board, while a near-identical
    # board can remain together if its investment identity agrees.
    return overlap >= 0.62 or (overlap >= 0.50 and bool(shared_carries))


def cluster_reason(members: list[PlayerFeature]) -> dict[str, Any]:
    pair_overlaps = [
        jaccard(left.hero_set, right.hero_set)
        for left, right in itertools.combinations(members, 2)
    ]
    archetypes = Counter(member.archetype for member in members)
    return {
        "method": "deterministic_identity_representative",
        "hero_jaccard_threshold": 0.5,
        "strict_hero_jaccard_threshold": 0.62,
        "high_cost_pdd_hero_jaccard_threshold": 0.67,
        "high_cost_pdd_requires_high_cost_core_continuity": True,
        "requires_identity_compatibility": True,
        "avg_pair_hero_jaccard": round(avg_number(pair_overlaps) or 1.0, 3),
        "archetype_distribution": [
            {"archetype": name, "appearances": count}
            for name, count in sorted(archetypes.items())
        ],
    }


def parse_trait_tier(key: str) -> tuple[str, int]:
    trait, tier_raw = key.rsplit("-", 1)
    return trait, int(tier_raw)


def carry_trait_names(main_carries: list[tuple[str, int]], members: list[PlayerFeature]) -> set[str]:
    names = {name for name, _ in main_carries[:3]}
    traits: set[str] = set()
    for member in members:
        for hero in member.heroes:
            if hero.name in names:
                traits.update(hero.traits)
    return traits


def derive_family_label(
    members: list[PlayerFeature],
    active_bond_counter: Counter[str],
    main_carries: list[tuple[str, int]],
) -> dict[str, Any]:
    _, dict_bond = load_game_config()
    total = len(members) or 1
    carry_traits = carry_trait_names(main_carries, members)
    candidates = []
    high_tier_member_count = 0
    for member in members:
        if any(
            tier > min(dict_bond.get(trait, [tier]))
            for trait, tier in member.active_traits.items()
            if trait in dict_bond
        ):
            high_tier_member_count += 1

    mostly_first_tier = high_tier_member_count / total < 0.5
    for key, count in active_bond_counter.items():
        trait, tier = parse_trait_tier(key)
        thresholds = dict_bond.get(trait)
        if not thresholds:
            continue
        share = count * 100.0 / total
        if share < 50:
            continue
        first_threshold = min(thresholds)
        second_threshold = sorted(thresholds)[1] if len(thresholds) > 1 else first_threshold
        carry_bonus = 30 if trait in carry_traits else 0
        score = tier * 100 + share + carry_bonus
        candidates.append(
            {
                "key": key,
                "trait": trait,
                "tier": tier,
                "share": round(share, 1),
                "is_first_tier": tier <= first_threshold,
                "carry_aligned": trait in carry_traits,
                "score": score,
                "is_high_tier": tier >= second_threshold,
            }
        )

    if not candidates:
        return {
            "key": "拼多多",
            "label_trait": "拼多多",
            "label_confidence": "低",
            "label_reason": "没有稳定占比足够的主羁绊",
        }

    candidates.sort(key=lambda item: (item["score"], item["share"]), reverse=True)
    best = candidates[0]
    if mostly_first_tier and not (best["carry_aligned"] and best["share"] >= 65):
        return {
            "key": "拼多多",
            "label_trait": "拼多多",
            "label_confidence": "中",
            "label_reason": "激活羁绊主要停留在第一档，按拼多多处理",
        }

    return {
        "key": best["key"],
        "label_trait": best["trait"],
        "label_confidence": "高" if best["share"] >= 60 else "中",
        "label_reason": "主C羁绊主导" if best["carry_aligned"] else "家族稳定高占比羁绊",
        "label_share": best["share"],
    }


def score_composition(row: dict[str, Any]) -> tuple[float, dict[str, float]]:
    """Rank with shrunk performance; popularity remains descriptive only."""
    stats = row["stats"]
    evidence = row.get("confidence_evidence", {})
    shrunk = evidence.get("shrunk_metrics", {})
    avg_rank = float(shrunk.get("avg_rank", stats["avg_rank"]))
    top4_lower = float(
        shrunk.get("top4", {}).get("lower_bound", float(stats["top4_rate"]) / 100.0)
    )
    win_lower = float(
        shrunk.get("win", {}).get("lower_bound", float(stats["win_rate"]) / 100.0)
    )
    n_eff = float(evidence.get("n_eff", stats.get("n_eff", stats["appearances"])))
    uncertainty_penalty = max(0.0, 1.0 - n_eff / RECOMMENDATION_MIN_EFFECTIVE_N) * 0.35
    difficulty_penalty = float(row["difficulty"].get("score", 0.5)) * 0.15
    three_star_penalty = 0.45 if row.get("high_cost_three_star_dependency") else 0.0
    trend_adjustment = {
        "上升": -0.08,
        "下滑": 0.08,
    }.get(row.get("trend", {}).get("label"), 0.0)
    score = (
        avg_rank
        - top4_lower * 0.35
        - win_lower * 0.20
        + uncertainty_penalty
        + difficulty_penalty
        + three_star_penalty
        + trend_adjustment
    )
    return round(score, 4), {
        "shrunk_avg_rank": round(avg_rank, 4),
        "top4_lower_bound": round(top4_lower, 4),
        "win_lower_bound": round(win_lower, 4),
        "uncertainty_penalty": round(uncertainty_penalty, 4),
        "difficulty_penalty": round(difficulty_penalty, 4),
        "high_cost_three_star_penalty": three_star_penalty,
        "trend_adjustment": trend_adjustment,
    }


def composition_recommendation_score(row: dict[str, Any]) -> float:
    return score_composition(row)[0]


def overall_strength_score(row: dict[str, Any]) -> float:
    stats = row["stats"]
    difficulty_score = row["difficulty"].get("score", 0.5)
    n = stats["appearances"]
    score = stats["avg_rank"]
    score -= stats["top4_rate"] / 100.0 * 0.45
    score -= stats["win_rate"] / 100.0 * 0.2
    score += difficulty_score * 0.75
    score -= min(n, 80) / 80.0 * 0.25
    if n < 10:
        score += 1.2
    elif n < 20:
        score += 0.4
    return round(score, 4)


def build_composition_row(
    members: list[PlayerFeature],
    family_id: int,
    total_players: float,
    total_matches: float,
    *,
    is_subfamily: bool = False,
    subfamily_key: str | None = None,
    cluster_evidence: dict[str, Any] | None = None,
) -> dict[str, Any]:
    stats = RankStats()
    hero_counter: Counter[str] = Counter()
    carry_counter: Counter[str] = Counter()
    carry_score_sums: dict[str, float] = defaultdict(float)
    carry_score_counts: Counter[str] = Counter()
    active_bond_counter: Counter[str] = Counter()
    member_weight = sum(member.sample_weight for member in members) or 1.0
    for member in members:
        stats.add(member.rank, member.sample_weight)
        hero_counter.update(member.hero_set)
        seen_carries: set[str] = set()
        for hero in member.carry_candidates[:3]:
            if hero.name in seen_carries:
                continue
            seen_carries.add(hero.name)
            carry_counter[hero.name] += 1
            carry_score_sums[hero.name] += hero.carry_score
            carry_score_counts[hero.name] += 1
        for trait, tier in member.active_traits.items():
            active_bond_counter[f"{trait}-{tier}"] += 1

    match_counts = Counter(member.match_id for member in members)
    avg_contest = sum(match_counts.values()) / len(match_counts)
    unfinished = 0
    carry_complete = 0
    for member in members:
        carry = member.main_carry
        if carry and carry.equipment_count >= 3:
            carry_complete += 1
        if member.rank > 4 and carry and (carry.equipment_count < 3 or carry.stars < 2):
            unfinished += 1

    unfinished_rate = unfinished * 100.0 / len(members)
    carry_complete_rate = carry_complete * 100.0 / len(members)
    three_star_counts = [three_star_lineup_count(member) for member in members]
    top4_three_star_counts = [
        three_star_lineup_count(member) for member in members if member.rank <= 4
    ]
    difficulty_score = (
        (unfinished_rate / 100.0) * 0.5
        + min(avg_contest / 3.0, 1.0) * 0.3
        + (1.0 - carry_complete_rate / 100.0) * 0.2
    )
    difficulty = "高" if difficulty_score >= 0.58 else "中" if difficulty_score >= 0.34 else "低"
    pick_rate = member_weight * 100.0 / total_players
    match_weight_by_id = {member.match_id: member.sample_weight for member in members}
    match_weight = sum(match_weight_by_id[mid] for mid in match_counts)
    match_share = match_weight * 100.0 / total_matches
    popularity_score = pick_rate / 20.0 + avg_contest / 3.0 + match_share / 80.0
    popularity = "高" if popularity_score >= 1.5 else "中" if popularity_score >= 0.8 else "低"

    top_bonds = active_bond_counter.most_common(8)
    main_carries = carry_counter.most_common(3)
    label_info = derive_family_label(members, active_bond_counter, main_carries)
    main_bond = subfamily_key or label_info["key"]
    variants = build_level_variants(members, hero_counter, main_bond=main_bond)
    if subfamily_key:
        label_info = {
            **label_info,
            "key": subfamily_key,
            "label_trait": parse_trait_tier(subfamily_key)[0],
            "label_confidence": "高",
            "label_reason": "样本充足的高档羁绊子形态",
        }
    if main_bond != "拼多多" and main_bond not in {key for key, _ in top_bonds}:
        top_bonds.append((main_bond, sum(1 for member in members if main_bond in {
            f"{trait}-{tier}" for trait, tier in member.active_traits.items()
        })))
    common_bonds = [
        {"bond": bond, "share": round(count * 100.0 / len(members), 1)}
        for bond, count in top_bonds[:8]
        if count > 0
    ]
    play_style, play_style_breakdown = play_style_summary(members)
    archetypes = Counter(member.archetype for member in members)
    archetype, _ = sorted(
        archetypes.items(), key=lambda item: (-item[1], item[0])
    )[0]
    archetype_signals = [
        signal
        for member in members
        for signal in member.archetype_signals
        if member.archetype == archetype
    ]
    carry_requirements = summarize_carry_requirements(members, main_carries)
    # Mature/stage rows that still ask for a low-cost 3-star carry are reroll
    # comps even when a minority of boards look high-cost.
    if has_low_cost_three_star_carry_requirement({"carry_requirements": carry_requirements}):
        play_style = "赌狗"
    carry_equipment_notes = summarize_comp_carry_equipment(members, main_carries)
    jiujiu_requirements = analyze_comp_jiujiu_dependency(members, main_bond)
    high_cost_three_star_dependency = any(
        row.get("high_cost_three_star_dependency") for row in carry_requirements
    )
    carry_label = "+".join(name for name, _ in main_carries[:2]) or "无核心"
    row = {
        "family_id": family_id,
        # Gameplay identity takes priority in the readable label, while the
        # factual bond and recurring carries remain visible for comparison.
        "label": f"{archetype} / {main_bond} / {carry_label}",
        "main_bond": main_bond,
        "is_subfamily": is_subfamily,
        "subfamily_key": subfamily_key,
        "label_confidence": label_info.get("label_confidence", "中"),
        "label_reason": label_info.get("label_reason", ""),
        "main_carries": [
            {
                "hero_name": name,
                "share": round(count * 100.0 / len(members), 1),
                "carry_rank": rank,
                "avg_carry_score": round(
                    carry_score_sums[name] / max(carry_score_counts[name], 1),
                    1,
                ),
            }
            for rank, (name, count) in enumerate(main_carries, start=1)
        ],
        "core_heroes": [
            {"hero_name": name, "share": round(count * 100.0 / len(members), 1)}
            for name, count in hero_counter.most_common(10)
        ],
        "common_bonds": common_bonds,
        "play_style": play_style,
        "play_style_breakdown": play_style_breakdown,
        "archetype": archetype,
        "archetype_distribution": [
            {"archetype": name, "appearances": count, "share": round(count * 100.0 / len(members), 1)}
            for name, count in sorted(archetypes.items(), key=lambda item: (-item[1], item[0]))
        ],
        "archetype_signals": archetype_signals[:12],
        "trait_investment": {
            "dominant_traits": Counter(
                member.trait_investment.get("dominant_trait")
                for member in members
                if member.trait_investment.get("dominant_trait")
            ).most_common(3),
            "stable_trait_rate": round(
                sum(bool(member.trait_investment.get("stable_traits")) for member in members)
                * 100.0 / len(members),
                1,
            ),
        },
        "high_cost_structure": {
            "avg_four_five_cost_count": round(
                avg_number([
                    member.high_cost_structure.get("four_five_cost_count", 0)
                    for member in members
                ]) or 0.0,
                2,
            ),
            "avg_four_five_cost_share": round(
                avg_number([
                    member.high_cost_structure.get("four_five_cost_share", 0.0)
                    for member in members
                ]) or 0.0,
                3,
            ),
            "avg_low_cost_three_star_count": round(
                avg_number([
                    member.high_cost_structure.get("low_cost_three_star_count", 0)
                    for member in members
                ]) or 0.0,
                2,
            ),
            "high_cost_two_star_main_carry_rate": round(
                sum(
                    1
                    for member in members
                    if member.high_cost_structure.get("main_carry_is_high_cost_two_star")
                )
                * 100.0
                / len(members),
                1,
            ),
        },
        "cluster_reason": cluster_evidence or cluster_reason(members),
        "variants": variants,
        "carry_requirements": carry_requirements,
        "carry_equipment_notes": carry_equipment_notes,
        "jiujiu_requirements": jiujiu_requirements,
        "stats": stats.to_dict(),
        "difficulty": {
            "label": difficulty,
            "score": round(difficulty_score, 3),
            "unfinished_bottom_rate": round(unfinished_rate, 1),
            "carry_complete_rate": round(carry_complete_rate, 1),
            "avg_same_match_contest": round(avg_contest, 2),
            "avg_family_contest": round(avg_contest, 2),
            "avg_three_star_units": round(avg_number(three_star_counts) or 0.0, 2),
            "avg_top4_three_star_units": round(
                avg_number(top4_three_star_counts) or avg_number(three_star_counts) or 0.0,
                2,
            ),
        },
        "popularity": {
            "label": popularity,
            "score": round(popularity_score, 3),
            "pick_rate": round(pick_rate, 1),
            "match_share": round(match_share, 1),
            "avg_same_match_contest": round(avg_contest, 2),
            "avg_family_contest": round(avg_contest, 2),
        },
        "confidence": confidence_label(len(members)),
        "member_player_ids": [member.player_id for member in members],
        "high_cost_three_star_dependency": high_cost_three_star_dependency,
    }
    row["confidence_evidence"] = build_confidence_evidence(row, members)
    row["recommendation_score"] = composition_recommendation_score(row)
    row["score_breakdown"] = score_composition(row)[1]
    row["overall_strength_score"] = overall_strength_score(row)
    return row


def high_tier_subgroups(
    members: list[PlayerFeature],
    min_apps: int,
) -> list[tuple[str, list[PlayerFeature]]]:
    _, dict_bond = load_game_config()
    by_key: dict[str, list[PlayerFeature]] = defaultdict(list)
    for member in members:
        for trait, tier in member.active_traits.items():
            thresholds = sorted(dict_bond.get(trait, []))
            if len(thresholds) < 2:
                continue
            if tier >= thresholds[1]:
                by_key[f"{trait}-{tier}"].append(member)
    result = []
    for key, rows in by_key.items():
        if len(rows) >= max(15, min_apps * 2) and len(rows) < len(members) * 0.92:
            result.append((key, rows))
    result.sort(key=lambda item: (-len(item[1]), item[0]))
    return result[:4]


def cluster_compositions(features: list[PlayerFeature], min_apps: int) -> list[dict[str, Any]]:
    candidates = [
        feature
        for feature in sorted(
            features,
            key=lambda item: (
                feature_identity_anchor(item),
                tuple(sorted(item.hero_set)),
                tuple(sorted(feature_core_carries(item))),
                item.player_id,
            ),
        )
        if len(feature.hero_set) >= 5
    ]
    # Stable structural ordering removes both input-order dependence and rank
    # leakage.  Requiring similarity to one fixed representative avoids the
    # transitive chaining of connected components (A~B, B~C, A!~C).
    families_by_identity: dict[str, list[list[PlayerFeature]]] = defaultdict(list)
    for feature in candidates:
        identity = feature_identity_anchor(feature)
        identity_families = families_by_identity[identity]
        destination = next(
            (
                members
                for members in identity_families
                if features_are_similar(feature, members[0])
            ),
            None,
        )
        if destination is None:
            identity_families.append([feature])
        else:
            destination.append(feature)
    families = [
        sorted(members, key=lambda item: (item.player_id, item.match_id))
        for identity in sorted(families_by_identity)
        for members in families_by_identity[identity]
    ]

    output: list[dict[str, Any]] = []
    family_id = 1
    total_players, total_match_weight, _ = weighted_totals(features)
    total_matches = total_match_weight
    for members in families:
        if len(members) < min_apps:
            continue
        for member in members:
            member.family_id = family_id
        evidence = cluster_reason(members)
        base_row = build_composition_row(
            members, family_id, total_players, total_matches, cluster_evidence=evidence
        )
        output.append(base_row)
        family_id += 1
        for sub_key, sub_members in high_tier_subgroups(members, min_apps):
            if sub_key == base_row["main_bond"]:
                continue
            output.append(
                build_composition_row(
                    sub_members,
                    family_id,
                    total_players,
                    total_matches,
                    is_subfamily=True,
                    subfamily_key=sub_key,
                    cluster_evidence=cluster_reason(sub_members),
                )
            )
            family_id += 1

    output.sort(key=lambda row: (min(row["member_player_ids"]), row["label"]))
    return output


def trait_name_from_bond_key(key: str) -> str:
    if key == "拼多多":
        return key
    return parse_trait_tier(key)[0]


def strategy_carry_key(row: dict[str, Any]) -> str:
    # Use a stable gameplay anchor plus only carries that recur through the
    # family.  Top3 investment candidates are intentionally not a signature:
    # a shared temporary holder must not combine otherwise different comps.
    archetype = row.get("archetype", "未分类")
    if archetype.startswith("羁绊运营:"):
        bond_anchor = archetype.split(":", 1)[1]
    elif archetype in {"美食社收菜", "高费拼多多", "拼多多"}:
        bond_anchor = archetype
    else:
        bond_anchor = trait_name_from_bond_key(row.get("main_bond", "拼多多"))
    carries = [
        item["hero_name"]
        for item in row.get("main_carries", [])
        if item.get("share", 0) >= 40
    ][:2]
    carry_anchor = "+".join(sorted(carries)) or "无稳定主C"
    return f"{archetype}|{bond_anchor}|{carry_anchor}"


def strategy_rows_compatible(left: dict[str, Any], right: dict[str, Any]) -> bool:
    left_key = strategy_carry_key(left).split("|")
    right_key = strategy_carry_key(right).split("|")
    if left_key[:2] != right_key[:2]:
        return False
    if set(left.get("member_player_ids", [])) & set(right.get("member_player_ids", [])):
        return True
    left_carries = {
        item["hero_name"]
        for item in left.get("main_carries", [])
        if item.get("share", 0) >= 40
    }
    right_carries = {
        item["hero_name"]
        for item in right.get("main_carries", [])
        if item.get("share", 0) >= 40
    }
    left_core = {
        item["hero_name"]
        for item in left.get("core_heroes", [])
        if item.get("share", 0) >= 50
    }
    right_core = {
        item["hero_name"]
        for item in right.get("core_heroes", [])
        if item.get("share", 0) >= 50
    }
    core_overlap = jaccard(left_core, right_core)
    if left.get("archetype") == right.get("archetype") == "高费拼多多":
        # Preserve a relationship only when the recurring high-cost structure
        # is continuous.  This blocks large PDD strategy buckets built from a
        # shared archetype or a temporary carry.
        return core_overlap >= 0.65
    if core_overlap >= 0.55:
        return True
    # A shared carry is supporting evidence only.  Requiring lineup continuity
    # prevents unrelated strategies from merging just because their Top3
    # investment candidates happen to include the same flexible unit.
    return bool(left_carries & right_carries) and core_overlap >= 0.35


def strategy_core_overlap(left: dict[str, Any], right: dict[str, Any]) -> float:
    left_core = {
        item["hero_name"]
        for item in left.get("core_heroes", [])
        if item.get("share", 0) >= 50
    }
    right_core = {
        item["hero_name"]
        for item in right.get("core_heroes", [])
        if item.get("share", 0) >= 50
    }
    return jaccard(left_core, right_core)


def group_high_cost_pdd_rows(
    rows: list[dict[str, Any]],
) -> list[list[dict[str, Any]]]:
    """Group PDD stages against a fixed representative, without bridge chaining."""
    ordered = sorted(
        rows,
        key=lambda row: (
            bool(row.get("is_subfamily")),
            -int(row.get("stats", {}).get("appearances", 0)),
            strategy_carry_key(row),
            min(row.get("member_player_ids", [0])),
            row.get("label", ""),
        ),
    )
    groups: list[list[dict[str, Any]]] = []
    for row in ordered:
        row_ids = set(row.get("member_player_ids", []))
        candidates: list[tuple[float, int, list[dict[str, Any]]]] = []
        for index, members in enumerate(groups):
            representative = members[0]
            directly_compatible = strategy_rows_compatible(row, representative)
            overlapping_parent = bool(row.get("is_subfamily")) and any(
                row_ids & set(member.get("member_player_ids", []))
                for member in members
            )
            if directly_compatible or overlapping_parent:
                candidates.append(
                    (strategy_core_overlap(row, representative), -index, members)
                )
        if not candidates:
            groups.append([row])
            continue
        destination = max(candidates, key=lambda item: (item[0], item[1]))[2]
        destination.append(row)
    return groups


def group_strategy_rows(comp_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Build deterministic strategy components from compatible stage rows."""
    rows = sorted(
        comp_rows,
        key=lambda row: (
            strategy_carry_key(row),
            min(row.get("member_player_ids", [0])),
            row.get("label", ""),
        ),
    )
    high_cost_pdd_rows = [
        row for row in rows if row.get("archetype") == "高费拼多多"
    ]
    normal_rows = [
        row for row in rows if row.get("archetype") != "高费拼多多"
    ]
    parent = list(range(len(normal_rows)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    for left_index, right_index in itertools.combinations(range(len(normal_rows)), 2):
        if not strategy_rows_compatible(
            normal_rows[left_index], normal_rows[right_index]
        ):
            continue
        left_root, right_root = find(left_index), find(right_index)
        if left_root != right_root:
            parent[max(left_root, right_root)] = min(left_root, right_root)
    grouped_rows: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for index, row in enumerate(normal_rows):
        grouped_rows[find(index)].append(row)
    groups = [
        members for _, members in sorted(grouped_rows.items())
    ] + group_high_cost_pdd_rows(high_cost_pdd_rows)
    groups.sort(
        key=lambda members: (
            strategy_carry_key(members[0]),
            min(
                player_id
                for row in members
                for player_id in row.get("member_player_ids", [0])
            ),
            members[0].get("label", ""),
        )
    )
    return {
        f"strategy-group-{group_index:04d}": members
        for group_index, members in enumerate(groups, start=1)
    }


def bond_stage_score(row: dict[str, Any]) -> tuple[int, float, float, int]:
    """Describe structural completion; performance is handled by stage selection."""
    key = row.get("main_bond", "")
    tier = 0
    if key and key != "拼多多" and "-" in key:
        _, tier = parse_trait_tier(key)
    carry_complete = float(row.get("difficulty", {}).get("carry_complete_rate", 0.0))
    n_eff = float(row.get("stats", {}).get("n_eff", 0.0))
    return (tier, carry_complete, n_eff, int(row["stats"]["appearances"]))


def select_mature_stage(
    rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Choose a completed stage without allowing an obvious performance inversion."""
    high_cost_pdd = all(
        row.get("archetype") == "高费拼多多" for row in rows
    )
    reliable = [
        row
        for row in rows
        if float(row.get("stats", {}).get("n_eff", 0.0))
        >= MATURE_STAGE_MIN_RELIABLE_N_EFF
    ]
    performance_pool = reliable or rows
    performance_reference = min(
        performance_pool,
        key=lambda row: (
            float(row.get("recommendation_score", float("inf"))),
            float(row["stats"]["avg_rank"]),
            -float(row["stats"]["top4_rate"]),
        ),
    )
    reference_stats = performance_reference["stats"]

    def inversion_reasons(row: dict[str, Any]) -> list[str]:
        if row is performance_reference or row not in reliable:
            return []
        reasons = []
        avg_regression = float(row["stats"]["avg_rank"]) - float(reference_stats["avg_rank"])
        top4_regression = float(reference_stats["top4_rate"]) - float(row["stats"]["top4_rate"])
        if avg_regression >= MATURE_STAGE_MAX_AVG_RANK_REGRESSION:
            reasons.append("avg_rank_regression")
        if top4_regression >= MATURE_STAGE_MAX_TOP4_REGRESSION:
            reasons.append("top4_regression")
        return reasons

    candidates = [row for row in rows if not inversion_reasons(row)]
    if high_cost_pdd:
        # PDD has no factual deep-trait progression. Its incidental first-tier
        # bond and a tiny carry-completion delta must not let a small flex
        # branch replace the recurring final-board shape.
        high_cost_candidates = [
            row for row in candidates if row.get("play_style") == "高费"
        ]
        selected = max(
            high_cost_candidates or candidates,
            key=lambda row: (
                float(row.get("stats", {}).get("n_eff", 0.0)),
                int(row.get("stats", {}).get("appearances", 0)),
                float(row.get("difficulty", {}).get("carry_complete_rate", 0.0)),
                -float(row.get("recommendation_score", float("inf"))),
            ),
        )
    else:
        selected = max(
            candidates,
            key=lambda row: (
                bond_stage_score(row),
                -float(row.get("recommendation_score", float("inf"))),
            ),
        )
    candidate_audit = [
        {
            "label": row.get("label"),
            "bond": row.get("main_bond"),
            "structural_score": list(bond_stage_score(row)),
            "recommendation_score": row.get("recommendation_score"),
            "stats": row.get("stats"),
            "reliable_performance": row in reliable,
            "inversion_reasons": inversion_reasons(row),
            "selected": row is selected,
        }
        for row in sorted(rows, key=bond_stage_score, reverse=True)
    ]
    rejected_higher_tier = [
        item
        for item in candidate_audit
        if item["inversion_reasons"]
        and item["structural_score"][0] > bond_stage_score(selected)[0]
    ]
    return selected, {
        "method": (
            "reliable_performance_guard_then_representative_sample"
            if high_cost_pdd
            else "reliable_performance_guard_then_structural_completion"
        ),
        "reliable_n_eff_minimum": MATURE_STAGE_MIN_RELIABLE_N_EFF,
        "max_avg_rank_regression": MATURE_STAGE_MAX_AVG_RANK_REGRESSION,
        "max_top4_percentage_point_regression": MATURE_STAGE_MAX_TOP4_REGRESSION,
        "performance_reference_label": performance_reference.get("label"),
        "candidates": candidate_audit,
        "stage_inversion_detected": bool(rejected_higher_tier),
        "rejected_higher_tier_stages": rejected_higher_tier,
    }


def merge_comp_strategies(
    comp_rows: list[dict[str, Any]],
    features: list[PlayerFeature],
) -> list[dict[str, Any]]:
    player_by_id = {feature.player_id: feature for feature in features}
    grouped = group_strategy_rows(comp_rows)

    # A base family and its high-tier subfamily overlap by design.  Resolve
    # ownership before aggregation so every player board appears in exactly
    # one top-level strategy, even when future clustering emits overlapping
    # stage candidates.
    owner_candidates: dict[int, list[str]] = defaultdict(list)
    for key, rows in grouped.items():
        for player_id in {pid for row in rows for pid in row["member_player_ids"]}:
            owner_candidates[player_id].append(key)
    owned_by_strategy = {
        player_id: sorted(
            keys,
            key=lambda key: (
                max(bond_stage_score(row) for row in grouped[key]),
                key,
            ),
            reverse=True,
        )[0]
        for player_id, keys in owner_candidates.items()
    }

    strategies: list[dict[str, Any]] = []
    strategy_index = 1
    total_players, total_match_weight, _ = weighted_totals(features)
    for group_key, rows in sorted(grouped.items()):
        member_ids = sorted(
            pid
            for pid in {pid for row in rows for pid in row["member_player_ids"]}
            if owned_by_strategy.get(pid) == group_key
        )
        members = [player_by_id[pid] for pid in member_ids if pid in player_by_id]
        if not members:
            continue
        mature, mature_stage_selection = select_mature_stage(rows)
        carry_key = strategy_carry_key(mature)
        mature_ids = set(mature["member_player_ids"]) & set(member_ids)
        if not mature_ids:
            mature_ids = set(member_ids)
        mature_members = [player_by_id[pid] for pid in sorted(mature_ids) if pid in player_by_id]
        transition_ids = sorted(set(member_ids) - mature_ids)
        transition_members = [
            player_by_id[pid] for pid in transition_ids if pid in player_by_id
        ]
        mature_rank_stats = RankStats()
        transition_rank_stats = RankStats()
        for member in mature_members:
            mature_rank_stats.add(member.rank, member.sample_weight)
        for member in transition_members:
            transition_rank_stats.add(member.rank, member.sample_weight)
        transition_rows = [
            {
                "label": row["label"],
                "bond": row["main_bond"],
                "role": "大成" if row is mature else "过渡",
                "stats": row["stats"],
                "difficulty": row["difficulty"],
                "popularity": row["popularity"],
                "play_style": row.get("play_style", "高费"),
                "play_style_breakdown": row.get("play_style_breakdown", []),
                "member_player_ids": sorted(
                    pid
                    for pid in row["member_player_ids"]
                    if pid in member_ids and pid not in mature_ids
                ),
                "recommendation_score": row["recommendation_score"],
            }
            for row in sorted(rows, key=lambda row: bond_stage_score(row), reverse=True)
        ]
        aggregate = build_composition_row(
            members,
            strategy_index,
            total_players,
            total_match_weight,
        )
        pairwise_core_overlaps = [
            strategy_core_overlap(left, right)
            for left, right in itertools.combinations(rows, 2)
        ]
        strategy_id = f"{trait_name_from_bond_key(mature['main_bond'])}|{carry_key}"
        mature_carry_label = "+".join(
            item["hero_name"] for item in mature.get("main_carries", [])[:2]
        ) or "无核心"
        aggregate.update(
            {
                "strategy_id": strategy_id,
                "family_id": strategy_index,
                "label": (
                    f"{mature.get('archetype', aggregate.get('archetype', '未分类'))} / "
                    f"{mature['main_bond']} / {mature_carry_label}"
                ),
                "main_bond": mature["main_bond"],
                "archetype": mature.get("archetype", aggregate.get("archetype")),
                "archetype_signals": mature.get(
                    "archetype_signals", aggregate.get("archetype_signals", [])
                ),
                "mature_stage": {
                    "label": mature["label"],
                    "bond": mature["main_bond"],
                    "stats": mature_rank_stats.to_dict(),
                    "variants": mature["variants"],
                    "play_style": mature.get("play_style", aggregate.get("play_style", "高费")),
                    "play_style_breakdown": mature.get("play_style_breakdown", []),
                    "carry_requirements": mature.get("carry_requirements", []),
                    "carry_equipment_notes": mature.get("carry_equipment_notes", []),
                },
                "mature_stage_selection": mature_stage_selection,
                "stage_inversion_diagnostics": {
                    "detected": mature_stage_selection["stage_inversion_detected"],
                    "rejected_higher_tier_stages": mature_stage_selection[
                        "rejected_higher_tier_stages"
                    ],
                },
                "transition_stages": transition_rows,
                "mature_stats": mature_rank_stats.to_dict(),
                "aggregate_stats": aggregate["stats"],
                "mature_member_player_ids": sorted(mature_ids),
                "transition_stats": transition_rank_stats.to_dict()
                if transition_members
                else None,
                # Retain the legacy key, but make strategy strength explicitly
                # use completed boards.  Aggregate/transition outcomes remain
                # separately available for formation difficulty analysis.
                "stats": mature_rank_stats.to_dict(),
                "variants": mature["variants"],
                "carry_requirements": mature.get("carry_requirements", aggregate.get("carry_requirements", [])),
                "carry_equipment_notes": mature.get("carry_equipment_notes", aggregate.get("carry_equipment_notes", [])),
                "high_cost_three_star_dependency": mature.get(
                    "high_cost_three_star_dependency", False
                ),
                "member_player_ids": member_ids,
                "strategy_stage_count": len(rows),
                "merge_reason": {
                    "strategy_signature": carry_key,
                    "archetype": mature.get("archetype"),
                    "compatible_bond_anchor": trait_name_from_bond_key(mature["main_bond"]),
                    "stable_carries": carry_key.rsplit("|", 1)[-1],
                    "grouping_method": (
                        "fixed_representative"
                        if mature.get("archetype") == "高费拼多多"
                        else "compatible_connected_components"
                    ),
                    "representative_label": rows[0].get("label"),
                    "pairwise_core_overlap_min": round(
                        min(pairwise_core_overlaps), 3
                    )
                    if pairwise_core_overlaps
                    else 1.0,
                    "pairwise_core_overlap_avg": round(
                        avg_number(pairwise_core_overlaps) or 1.0, 3
                    ),
                    "ownership_rule": "one_player_one_top_level_strategy",
                    "mature_member_count": len(mature_members),
                    "transition_member_count": len(transition_members),
                    "play_style_source": "mature_stage_with_low_cost_three_star_gate",
                },
            }
        )
        # Recommendation bucket follows the mature stage, not aggregate majority
        # vote across transition boards. Low-cost 3-star carry advice still
        # forces 赌狗 even if the mature stage was labeled 高费.
        aggregate["play_style"] = resolve_strategy_play_style(aggregate)
        mature_stage = aggregate.get("mature_stage")
        if isinstance(mature_stage, dict):
            mature_stage["play_style"] = aggregate["play_style"]
        aggregate["confidence_evidence"] = build_confidence_evidence(
            aggregate, mature_members
        )
        aggregate["recommendation_score"] = composition_recommendation_score(aggregate)
        aggregate["score_breakdown"] = score_composition(aggregate)[1]
        aggregate["overall_strength_score"] = overall_strength_score(aggregate)
        strategies.append(aggregate)
        for feature in members:
            feature.family_id = strategy_index
        strategy_index += 1

    strategies.sort(
        key=lambda row: (
            row["overall_strength_score"],
            row["recommendation_score"],
            -row["aggregate_stats"]["appearances"],
            row["aggregate_stats"]["avg_rank"],
        )
    )
    for rank, strategy in enumerate(strategies, start=1):
        strategy["strength_rank"] = rank
    return strategies


def calibrate_composition_confidence(
    comp_rows: list[dict[str, Any]],
    features: list[PlayerFeature],
) -> None:
    """Apply play-style-specific priors after strategy maturity is resolved."""
    player_by_id = {feature.player_id: feature for feature in features}
    baselines = composition_baselines(features)
    for row in comp_rows:
        member_ids = row.get("mature_member_player_ids", row.get("member_player_ids", []))
        members = [player_by_id[player_id] for player_id in member_ids if player_id in player_by_id]
        baseline = baselines.get(row.get("play_style"), baselines["all"])
        row["confidence_evidence"] = build_confidence_evidence(row, members, baseline)
        row["recommendation_score"] = composition_recommendation_score(row)
        row["score_breakdown"] = score_composition(row)[1]


def trend_window_batches(
    features: list[PlayerFeature],
    *,
    boundary_batch: str | None = None,
    recent_batches: int = TREND_RECENT_BATCHES,
    prior_batches: int = TREND_PRIOR_BATCHES,
    reference_date: date | None = None,
) -> tuple[list[str], list[str], str]:
    batches = ordered_batches(features, reference_date)
    if boundary_batch and boundary_batch in batches:
        boundary_index = batches.index(boundary_batch)
        previous = batches[max(0, boundary_index - prior_batches) : boundary_index]
        current = batches[boundary_index : boundary_index + recent_batches]
        return current, previous, "balance_boundary"
    current = batches[-recent_batches:]
    previous_end = max(0, len(batches) - len(current))
    previous = batches[max(0, previous_end - prior_batches) : previous_end]
    return current, previous, "rolling"


def summarize_trend_window(
    members: list[PlayerFeature],
    population: list[PlayerFeature],
    baseline: dict[str, float],
) -> dict[str, Any]:
    stats = RankStats()
    for member in members:
        stats.add(member.rank)
    raw = stats.to_dict() if members else {
        "appearances": 0,
        "weighted_appearances": 0.0,
        "n_eff": 0.0,
        "avg_rank": None,
        "win_rate": None,
        "top4_rate": None,
    }
    n = len(members)
    shrunk_avg = None
    shrunk_top4 = None
    if n:
        shrunk_avg = (
            float(raw["avg_rank"]) * n
            + baseline["avg_rank"] * COMPOSITION_RANK_PRIOR_STRENGTH
        ) / (n + COMPOSITION_RANK_PRIOR_STRENGTH)
        shrunk_top4 = beta_posterior_summary(
            float(raw["top4_rate"]) / 100.0,
            float(n),
            baseline["top4_rate"],
        )["posterior_mean"] * 100.0
    return {
        "samples": n,
        "population_samples": len(population),
        "pick_rate": round(n * 100.0 / len(population), 2) if population else None,
        "raw_avg_rank": raw["avg_rank"],
        "raw_top4_rate": raw["top4_rate"],
        "shrunk_avg_rank": round(shrunk_avg, 3) if shrunk_avg is not None else None,
        "shrunk_top4_rate": round(shrunk_top4, 2) if shrunk_top4 is not None else None,
    }


def classify_trend(recent: dict[str, Any], prior: dict[str, Any]) -> tuple[str, list[str]]:
    if (
        recent["samples"] < TREND_MIN_SAMPLES_PER_WINDOW
        or prior["samples"] < TREND_MIN_SAMPLES_PER_WINDOW
        or not recent["population_samples"]
        or not prior["population_samples"]
    ):
        return "insufficient", ["任一窗口阵容样本少于判定门槛"]
    changes = {
        "pick_rate": recent["pick_rate"] - prior["pick_rate"],
        "shrunk_avg_rank": recent["shrunk_avg_rank"] - prior["shrunk_avg_rank"],
        "shrunk_top4_rate": recent["shrunk_top4_rate"] - prior["shrunk_top4_rate"],
    }
    up = sum(
        (
            changes["pick_rate"] >= TREND_PICK_RATE_THRESHOLD,
            changes["shrunk_avg_rank"] <= -TREND_AVG_RANK_THRESHOLD,
            changes["shrunk_top4_rate"] >= TREND_TOP4_THRESHOLD,
        )
    )
    down = sum(
        (
            changes["pick_rate"] <= -TREND_PICK_RATE_THRESHOLD,
            changes["shrunk_avg_rank"] >= TREND_AVG_RANK_THRESHOLD,
            changes["shrunk_top4_rate"] <= -TREND_TOP4_THRESHOLD,
        )
    )
    if up >= 2 and down == 0:
        return "上升", []
    if down >= 2 and up == 0:
        return "下滑", []
    return "稳定", []


def attach_composition_trends(
    comp_rows: list[dict[str, Any]],
    features: list[PlayerFeature],
    *,
    balance_boundary: dict[str, Any] | None = None,
    reference_date: date | None = None,
) -> dict[str, Any]:
    boundary_batch = (
        balance_boundary.get("batch")
        if balance_boundary and balance_boundary.get("supported")
        else None
    )
    recent_batches, prior_batches, mode = trend_window_batches(
        features,
        boundary_batch=boundary_batch,
        reference_date=reference_date,
    )
    recent_set, prior_set = set(recent_batches), set(prior_batches)
    recent_population = [feature for feature in features if feature.match_batch in recent_set]
    prior_population = [feature for feature in features if feature.match_batch in prior_set]
    player_by_id = {feature.player_id: feature for feature in features}
    baselines = composition_baselines(features)
    for row in comp_rows:
        member_ids = set(row.get("member_player_ids", []))
        recent_members = [
            player_by_id[player_id]
            for player_id in member_ids
            if player_id in player_by_id and player_by_id[player_id].match_batch in recent_set
        ]
        prior_members = [
            player_by_id[player_id]
            for player_id in member_ids
            if player_id in player_by_id and player_by_id[player_id].match_batch in prior_set
        ]
        baseline = baselines.get(row.get("play_style"), baselines["all"])
        recent = summarize_trend_window(recent_members, recent_population, baseline)
        prior = summarize_trend_window(prior_members, prior_population, baseline)
        label, reasons = classify_trend(recent, prior)
        changes = {
            key: (
                round(recent[key] - prior[key], 3)
                if recent[key] is not None and prior[key] is not None
                else None
            )
            for key in ("pick_rate", "shrunk_avg_rank", "shrunk_top4_rate")
        }
        row["trend"] = {
            "label": label,
            "status": "insufficient" if label == "insufficient" else "determined",
            "mode": mode,
            "recent_batches": recent_batches,
            "prior_batches": prior_batches,
            "recent": recent,
            "prior": prior,
            "changes": changes,
            "reasons": reasons,
        }
        row["recommendation_score"] = composition_recommendation_score(row)
        row["score_breakdown"] = score_composition(row)[1]
    return {
        "mode": mode,
        "recent_batches": recent_batches,
        "prior_batches": prior_batches,
        "min_samples_per_window": TREND_MIN_SAMPLES_PER_WINDOW,
        "thresholds": {
            "pick_rate_percentage_points": TREND_PICK_RATE_THRESHOLD,
            "shrunk_avg_rank": TREND_AVG_RANK_THRESHOLD,
            "shrunk_top4_percentage_points": TREND_TOP4_THRESHOLD,
            "direction_rule": "at least two metrics agree and none strongly opposes",
        },
    }


def three_star_required_carries(row: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "hero_name": item["hero_name"],
            "tier": item.get("tier"),
            "share": item.get("share"),
            "samples": item.get("samples"),
            "three_star_rate": item.get("three_star_rate"),
            "avg_stars_top4": item.get("avg_stars_top4"),
        }
        for item in row.get("carry_requirements", [])
        if item.get("recommended_min_stars", 0) >= 3
    ]


def relabel_difficulty(score: float) -> str:
    return "高" if score >= 0.58 else "中" if score >= 0.34 else "低"


def relabel_popularity(score: float) -> str:
    return "高" if score >= 1.5 else "中" if score >= 0.8 else "低"


def enrich_three_star_contest(
    comp_rows: list[dict[str, Any]],
    features: list[PlayerFeature],
) -> list[dict[str, Any]]:
    player_to_strategy: dict[int, dict[str, Any]] = {}
    for row in comp_rows:
        required = three_star_required_carries(row)
        row["three_star_required_carries"] = required
        row["low_cost_three_star_required_carries"] = [
            item for item in required if item.get("tier") is not None and item["tier"] <= 3
        ]
        for player_id in row.get("member_player_ids", []):
            player_to_strategy[player_id] = row

    features_by_match: dict[int, list[PlayerFeature]] = defaultdict(list)
    for feature in features:
        features_by_match[feature.match_id].append(feature)

    hero_match_counts: dict[str, Counter[int]] = defaultdict(Counter)
    hero_rank_stats: dict[str, RankStats] = defaultdict(RankStats)
    hero_strategy_labels: dict[str, Counter[str]] = defaultdict(Counter)

    for feature in features:
        strategy = player_to_strategy.get(feature.player_id)
        if not strategy:
            continue
        for required in strategy.get("three_star_required_carries", []):
            hero_name = required["hero_name"]
            hero_match_counts[hero_name][feature.match_id] += 1
            hero_rank_stats[hero_name].add(feature.rank, feature.sample_weight)
            hero_strategy_labels[hero_name][strategy["label"]] += 1

    for row in comp_rows:
        required_names = {
            item["hero_name"] for item in row.get("three_star_required_carries", [])
        }
        overlap_values: list[int] = []
        overlap_label_counter: Counter[str] = Counter()
        for player_id in row.get("member_player_ids", []):
            feature = next((item for item in features if item.player_id == player_id), None)
            if feature is None or not required_names:
                continue
            match_features = features_by_match.get(feature.match_id, [])
            max_overlap = 1
            for hero_name in required_names:
                same_need = [
                    other
                    for other in match_features
                    if (
                        other_strategy := player_to_strategy.get(other.player_id)
                    )
                    and any(
                        req["hero_name"] == hero_name
                        for req in other_strategy.get("three_star_required_carries", [])
                    )
                ]
                if len(same_need) > max_overlap:
                    max_overlap = len(same_need)
                for other in same_need:
                    other_strategy = player_to_strategy.get(other.player_id)
                    if other_strategy and other_strategy is not row:
                        overlap_label_counter[other_strategy["label"]] += 1
            overlap_values.append(max_overlap)

        avg_overlap = round(avg_number(overlap_values) or 0.0, 2)
        difficulty = row["difficulty"]
        popularity = row["popularity"]
        family_contest = difficulty.get("avg_family_contest", difficulty["avg_same_match_contest"])
        combined_contest = max(family_contest, avg_overlap)
        difficulty["avg_required_carry_contest"] = avg_overlap
        difficulty["avg_same_match_contest"] = round(combined_contest, 2)
        difficulty["contest_basis"] = "3星主C重叠" if avg_overlap > family_contest else "阵容相似"
        difficulty["overlap_strategies"] = [
            {"label": label, "samples": count}
            for label, count in overlap_label_counter.most_common(5)
        ]
        difficulty_score = (
            (difficulty["unfinished_bottom_rate"] / 100.0) * 0.5
            + min(combined_contest / 3.0, 1.0) * 0.3
            + (1.0 - difficulty["carry_complete_rate"] / 100.0) * 0.2
        )
        difficulty["score"] = round(difficulty_score, 3)
        difficulty["label"] = relabel_difficulty(difficulty_score)
        popularity["avg_required_carry_contest"] = avg_overlap
        popularity["avg_same_match_contest"] = round(combined_contest, 2)
        popularity["contest_basis"] = difficulty["contest_basis"]
        popularity_score = (
            popularity["pick_rate"] / 20.0
            + combined_contest / 3.0
            + popularity["match_share"] / 80.0
        )
        popularity["score"] = round(popularity_score, 3)
        popularity["label"] = relabel_popularity(popularity_score)
        row["overall_strength_score"] = overall_strength_score(row)
        row["recommendation_score"] = composition_recommendation_score(row)
        row["score_breakdown"] = score_composition(row)[1]

    rows: list[dict[str, Any]] = []
    for hero_name, match_counts in hero_match_counts.items():
        stat = hero_rank_stats[hero_name]
        if not match_counts:
            continue
        avg_contest = sum(match_counts.values()) / len(match_counts)
        multi_match_rate = (
            sum(1 for count in match_counts.values() if count >= 2)
            * 100.0
            / len(match_counts)
        )
        top_strategy_labels = [
            {"label": label, "samples": count}
            for label, count in hero_strategy_labels[hero_name].most_common(4)
        ]
        tier = None
        for row in comp_rows:
            for item in row.get("three_star_required_carries", []):
                if item["hero_name"] == hero_name:
                    tier = item.get("tier")
                    break
            if tier is not None:
                break
        rows.append(
            {
                "hero_name": hero_name,
                "tier": tier,
                **stat.to_dict(),
                "match_appearances": len(match_counts),
                "avg_same_match_needers": round(avg_contest, 2),
                "max_same_match_needers": max(match_counts.values()),
                "multi_needer_match_rate": round(multi_match_rate, 1),
                "top_strategies": top_strategy_labels,
                "is_low_cost": tier is not None and tier <= 3,
            }
        )
    rows.sort(
        key=lambda row: (
            not row["is_low_cost"],
            -row["avg_same_match_needers"],
            -row["appearances"],
            row["avg_rank"],
        )
    )
    for rank, row in enumerate(sorted(comp_rows, key=lambda item: item["overall_strength_score"]), start=1):
        row["strength_rank"] = rank
    comp_rows.sort(
        key=lambda row: (
            row["overall_strength_score"],
            row["recommendation_score"],
            -row["aggregate_stats"]["appearances"],
            row["aggregate_stats"]["avg_rank"],
        )
    )
    return rows


def high_cost_ceiling_completion(feature: PlayerFeature) -> dict[str, Any]:
    """Describe a completed high-investment *final board*, never a transition."""
    lineup = [hero for hero in feature.heroes if is_lineup_hero(hero.name)]
    high_cost = [hero for hero in lineup if hero.tier is not None and hero.tier >= 4]
    two_star_high_cost = [hero for hero in high_cost if hero.stars >= 2]
    three_star_high_cost = [hero for hero in high_cost if hero.stars >= 3]
    main_carry = feature.main_carry
    high_cost_share = len(high_cost) / max(len(lineup), 1)
    conditions = {
        "level_9_or_higher": feature.level >= 9,
        "four_five_cost_count": len(high_cost) >= 4,
        "four_five_cost_share": high_cost_share >= 0.50,
        "key_high_cost_two_star": bool(two_star_high_cost),
        "main_carry_equipment_complete": bool(
            main_carry and main_carry.equipment_count >= 3
        ),
    }
    return {
        "qualifies": classify_play_style(feature) == "高费" and all(conditions.values()),
        "conditions": conditions,
        "level": feature.level,
        "four_five_cost_count": len(high_cost),
        "four_five_cost_share": round(high_cost_share, 3),
        "two_star_high_cost_heroes": sorted({hero.name for hero in two_star_high_cost}),
        "three_star_high_cost_heroes": sorted({hero.name for hero in three_star_high_cost}),
        "main_carry": main_carry.name if main_carry else None,
        "main_carry_equipment_count": main_carry.equipment_count if main_carry else 0,
        "core_heroes": sorted(hero.name for hero in lineup),
        "active_traits": dict(feature.active_traits),
    }


def build_representative_final_boards(
    members: list[PlayerFeature],
    *,
    limit: int = CEILING_REPRESENTATIVE_BOARD_LIMIT,
) -> list[dict[str, Any]]:
    """Aggregate observed final boards by exact slot-ordered hero lineup."""
    if not members:
        return []
    groups: dict[tuple[str, ...], list[PlayerFeature]] = defaultdict(list)
    for member in members:
        heroes = tuple(unique_heroes_by_slot(member))
        if heroes:
            groups[heroes].append(member)
    boards: list[dict[str, Any]] = []
    total = len(members)
    for heroes, group in groups.items():
        stats = RankStats()
        carry_counter: Counter[str] = Counter()
        for member in group:
            stats.add(member.rank, member.sample_weight)
            if member.main_carry:
                carry_counter[member.main_carry.name] += 1
        main_carry = carry_counter.most_common(1)[0][0] if carry_counter else None
        boards.append(
            {
                "source": "sample",
                "final_board_only": True,
                "heroes": list(heroes),
                "level": len(heroes),
                "appearances": len(group),
                "weighted_appearances": round(stats.weighted_appearances, 2),
                "share": round(len(group) * 100.0 / total, 1),
                "avg_rank": stats.to_dict()["avg_rank"],
                "top4_rate": stats.to_dict()["top4_rate"],
                "win_rate": stats.to_dict()["win_rate"],
                "main_carry": main_carry,
                "confidence": confidence_label(len(group)),
            }
        )
    boards.sort(
        key=lambda row: (
            -row["appearances"],
            -row["weighted_appearances"],
            row["avg_rank"],
            row["heroes"],
        )
    )
    return boards[:limit]


def build_high_cost_ceiling_samples(
    comp_rows: list[dict[str, Any]],
    features: list[PlayerFeature],
) -> list[dict[str, Any]]:
    """Discover exceptional completed boards without treating snapshots as paths."""
    player_by_id = {feature.player_id: feature for feature in features}
    baselines = composition_baselines(features)
    samples: list[dict[str, Any]] = []
    for strategy in comp_rows:
        if strategy.get("play_style") != "高费":
            continue
        mature_ids = strategy.get(
            "mature_member_player_ids", strategy.get("member_player_ids", [])
        )
        members = [
            player_by_id[player_id]
            for player_id in mature_ids
            if player_id in player_by_id
            and high_cost_ceiling_completion(player_by_id[player_id])["qualifies"]
        ]
        if not members:
            continue
        completions = [high_cost_ceiling_completion(member) for member in members]
        stats = RankStats()
        for member in members:
            stats.add(member.rank, member.sample_weight)
        archetypes = Counter(member.archetype for member in members)
        three_star_names = sorted(
            {
                hero_name
                for completion in completions
                for hero_name in completion["three_star_high_cost_heroes"]
            }
        )
        representative_final_boards = build_representative_final_boards(members)
        ceiling_hero_counter: Counter[str] = Counter()
        for member in members:
            ceiling_hero_counter.update(unique_heroes_by_slot(member))
        ceiling_core_heroes = [
            {
                "hero_name": name,
                "share": round(count * 100.0 / len(members), 1),
            }
            for name, count in ceiling_hero_counter.most_common(12)
        ]
        ceiling_stage = {
            "kind": "高费大成上限样本",
            "final_board_only": True,
            "interpretation": "仅根据最终盘完成度归纳形态，不代表观测到真实过渡过程",
            "sample_count": len(members),
            "avg_level": round(avg_number([item["level"] for item in completions]) or 0.0, 2),
            "avg_four_five_cost_count": round(
                avg_number([item["four_five_cost_count"] for item in completions]) or 0.0, 2
            ),
            "avg_four_five_cost_share": round(
                avg_number([item["four_five_cost_share"] for item in completions]) or 0.0, 3
            ),
            "two_star_high_cost_heroes": sorted(
                {
                    hero_name
                    for completion in completions
                    for hero_name in completion["two_star_high_cost_heroes"]
                }
            ),
            "main_carry_equipment_complete_rate": round(
                sum(item["main_carry_equipment_count"] >= 3 for item in completions)
                * 100.0 / len(completions),
                1,
            ),
            "core_heroes": ceiling_core_heroes,
            "common_bonds": strategy.get("common_bonds", []),
            "representative_final_boards": representative_final_boards,
            "high_cost_three_star_nonstandard": bool(three_star_names),
            "high_cost_three_star_heroes": three_star_names,
            "high_investment_conditions": [
                "等级>=9",
                "4/5费至少4张且占阵容>=50%",
                "至少一张关键4/5费两星",
                "主C三件装备完整",
            ],
        }
        sample = {
            "label": f"高费大成上限 / {strategy['label']}",
            "strategy_id": strategy.get("strategy_id"),
            "source_strategy_label": strategy["label"],
            "play_style": "高费",
            "archetype": strategy.get("archetype", "未分类"),
            "archetype_distribution": [
                {
                    "archetype": name,
                    "appearances": count,
                    "share": round(count * 100.0 / len(members), 1),
                }
                for name, count in sorted(archetypes.items(), key=lambda item: (-item[1], item[0]))
            ],
            "stats": stats.to_dict(),
            "member_player_ids": [member.player_id for member in members],
            "core_heroes": ceiling_core_heroes,
            "representative_final_boards": representative_final_boards,
            "ceiling_stage": ceiling_stage,
            "high_cost_three_star_dependency": bool(three_star_names),
        }
        sample["confidence_evidence"] = build_confidence_evidence(
            sample, members, baselines["高费"], allow_high_cost_ceiling=True
        )
        sample["recommendation_status"] = (
            "正式高费大成上限"
            if sample["confidence_evidence"]["recommendation_eligible"]
            else "观察/低置信"
        )
        sample["ceiling_score"] = round(
            float(sample["stats"]["avg_rank"])
            - float(sample["stats"]["top2_rate"]) / 100.0 * 0.55
            - float(sample["stats"]["win_rate"]) / 100.0 * 0.25,
            4,
        )
        strategy["ceiling_stage"] = ceiling_stage
        samples.append(sample)
    samples.sort(
        key=lambda row: (
            -row["stats"]["top2_rate"],
            -row["stats"]["win_rate"],
            row["stats"]["avg_rank"],
            row["label"],
        )
    )
    return samples


def build_composition_recommendations(
    comp_rows: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Split discovered strategies into 赌狗 / 高费 only.

    Every strategy that already passed the discovery threshold
    (``min_comp_apps`` during clustering/merge) enters its play-style bucket.
    Formal ``recommendation_eligible`` remains audit/display evidence and does
    not gate inclusion. There is no per-style count limit.
    """
    return {
        style: sorted(
            (row for row in comp_rows if row.get("play_style") == style),
            key=lambda row: (row["recommendation_score"], row["label"]),
        )
        for style in PLAY_STYLES
    }


def avg_number(values: list[int | float]) -> float | None:
    if not values:
        return None
    return sum(values) / len(values)


def carry_for_name(feature: PlayerFeature, hero_name: str) -> Hero | None:
    for hero in feature.carry_candidates[:3]:
        if hero.name == hero_name:
            return hero
    if feature.main_carry and feature.main_carry.name == hero_name:
        return feature.main_carry
    return None


def compute_variant_bond_status(
    feature: PlayerFeature,
    main_bond: str | None,
) -> dict[str, Any] | None:
    if not main_bond or main_bond == "拼多多" or "-" not in main_bond:
        return None
    _, dict_bond = load_game_config()
    trait, target_tier = parse_trait_tier(main_bond)
    thresholds = dict_bond.get(trait, [])
    if not thresholds:
        return None
    hero_only_tier = active_tier(feature.trait_counts.get(trait, 0), thresholds)
    actual_tier = feature.active_traits.get(trait, 0)
    jiujiu_wearers: list[dict[str, str]] = []
    for hero in feature.heroes:
        for equipment in hero.equipments:
            if jiujiu_trait(equipment.raw_name) == trait:
                jiujiu_wearers.append(
                    {"hero_name": hero.name, "equipment_name": equipment.name}
                )
    return {
        "trait": trait,
        "target_tier": target_tier,
        "hero_only_tier": hero_only_tier,
        "actual_tier": actual_tier,
        "needs_jiujiu": hero_only_tier < target_tier and actual_tier >= target_tier,
        "meets_title_bond": actual_tier >= target_tier,
        "jiujiu_wearers": jiujiu_wearers,
        "active_traits": dict(feature.active_traits),
        "jiujiu_bonus": dict(feature.jiujiu_bonus),
    }


def analyze_comp_jiujiu_dependency(
    members: list[PlayerFeature],
    main_bond: str,
) -> list[dict[str, Any]]:
    if not main_bond or main_bond == "拼多多" or "-" not in main_bond:
        return []
    _, dict_bond = load_game_config()
    trait, target_tier = parse_trait_tier(main_bond)
    thresholds = dict_bond.get(trait, [])
    if not thresholds:
        return []

    dependency_samples = 0
    wearer_counter: Counter[str] = Counter()
    jiujiu_item_counter: Counter[str] = Counter()
    for member in members:
        hero_only_tier = active_tier(member.trait_counts.get(trait, 0), thresholds)
        actual_tier = member.active_traits.get(trait, 0)
        if hero_only_tier < target_tier and actual_tier >= target_tier:
            dependency_samples += 1
            for hero in member.heroes:
                for equipment in hero.equipments:
                    if jiujiu_trait(equipment.raw_name) == trait:
                        wearer_counter[hero.name] += 1
                        jiujiu_item_counter[equipment.name] += 1

    if dependency_samples == 0:
        return []

    dependency_rate = dependency_samples * 100.0 / len(members)
    recommended_jiujiu = (
        jiujiu_item_counter.most_common(1)[0][0]
        if jiujiu_item_counter
        else f"{trait}啾啾"
    )
    return [
        {
            "trait": trait,
            "target_tier": target_tier,
            "dependency_rate": round(dependency_rate, 1),
            "dependency_samples": dependency_samples,
            "recommended_jiujiu": recommended_jiujiu,
            "recommended_wearers": [
                {
                    "hero_name": name,
                    "share": round(count * 100.0 / dependency_samples, 1),
                }
                for name, count in wearer_counter.most_common(4)
            ],
        }
    ]


def variant_bond_note(bond_status: dict[str, Any] | None) -> str:
    if not bond_status:
        return "—"
    trait = bond_status["trait"]
    target = bond_status["target_tier"]
    if bond_status["meets_title_bond"]:
        if bond_status["needs_jiujiu"]:
            wearers = "、".join(
                item["hero_name"] for item in bond_status.get("jiujiu_wearers", [])[:2]
            )
            return f"需{trait}啾啾({wearers or '待观察'})"
        return f"已达成{trait}-{target}"
    return f"未达{trait}-{target}(纯{trait}{bond_status['hero_only_tier']})"


def build_level_variants(
    members: list[PlayerFeature],
    family_hero_counter: Counter[str],
    *,
    main_bond: str | None = None,
) -> dict[str, dict[str, Any]]:
    variants: dict[str, dict[str, Any]] = {}
    hero_order = [name for name, _ in family_hero_counter.most_common()]
    for target in (7, 8, 9):
        exact = [
            member
            for member in members
            if len(unique_heroes_by_slot(member)) == target
        ]
        if exact:
            candidates = [
                (member, compute_variant_bond_status(member, main_bond))
                for member in exact
            ]
            best, bond_status = sorted(
                candidates,
                key=lambda item: (
                    not (
                        item[1] is None
                        or item[1].get("meets_title_bond", False)
                    ),
                    item[0].rank,
                    -len(item[0].hero_set),
                ),
            )[0]
            variants[str(target)] = {
                "source": "sample",
                "confidence": confidence_label(len(exact)),
                "rank": best.rank,
                "heroes": unique_heroes_by_slot(best),
                "main_carry": best.main_carry.name if best.main_carry else None,
                "sample_count": len(exact),
                "bond_status": bond_status,
                "bond_note": variant_bond_note(bond_status),
                "jiujiu_wearers": bond_status.get("jiujiu_wearers", []) if bond_status else [],
            }
        else:
            variants[str(target)] = {
                "source": "derived",
                "confidence": "低",
                "rank": None,
                "heroes": hero_order[:target],
                "main_carry": members[0].main_carry.name if members and members[0].main_carry else None,
                "sample_count": 0,
                "bond_status": None,
                "bond_note": "推导阵容，未绑定样本羁绊",
                "jiujiu_wearers": [],
            }
    return variants


def unique_heroes_by_slot(feature: PlayerFeature) -> list[str]:
    seen: set[str] = set()
    heroes: list[str] = []
    for hero in sorted(feature.heroes, key=lambda item: item.slot_index):
        if not is_lineup_hero(hero.name):
            continue
        if hero.name in seen:
            continue
        seen.add(hero.name)
        heroes.append(hero.name)
    return heroes


def median_number(values: list[int]) -> float | None:
    if not values:
        return None
    values = sorted(values)
    mid = len(values) // 2
    if len(values) % 2 == 1:
        return float(values[mid])
    return (values[mid - 1] + values[mid]) / 2.0


def summarize_carry_requirements(
    members: list[PlayerFeature],
    main_carries: list[tuple[str, int]],
) -> list[dict[str, Any]]:
    requirements = []
    for hero_name, count in main_carries[:3]:
        carry_samples = [
            carry
            for member in members
            if (carry := carry_for_name(member, hero_name)) is not None
        ]
        top4_samples = [
            carry
            for member in members
            if member.rank <= 4
            and (carry := carry_for_name(member, hero_name)) is not None
        ]
        if not carry_samples:
            continue

        stars = [carry.stars for carry in top4_samples] or [carry.stars for carry in carry_samples]
        equipment_counts = [
            carry.equipment_count for carry in top4_samples
        ] or [carry.equipment_count for carry in carry_samples]
        two_star_rate = sum(1 for value in stars if value >= 2) * 100.0 / len(stars)
        three_star_rate = sum(1 for value in stars if value >= 3) * 100.0 / len(stars)
        three_item_rate = sum(1 for value in equipment_counts if value >= 3) * 100.0 / len(equipment_counts)
        bottom_underbuilt = 0
        bottom_samples = 0
        for member in members:
            carry = carry_for_name(member, hero_name)
            if carry is None or member.rank <= 4:
                continue
            bottom_samples += 1
            if carry.stars < 2 or carry.equipment_count < 3:
                bottom_underbuilt += 1
        bottom_underbuilt_rate = (
            bottom_underbuilt * 100.0 / bottom_samples
            if bottom_samples
            else 0.0
        )
        tier = carry_samples[0].tier or 0
        high_cost_three_star_dependency = tier >= 4 and three_star_rate >= 45
        if tier >= 4:
            recommended_star = 2 if two_star_rate >= 45 else max(1, min(stars))
        else:
            recommended_star = 3 if three_star_rate >= 45 else 2 if two_star_rate >= 60 else max(1, min(stars))
        requirements.append(
            {
                "hero_name": hero_name,
                "tier": tier,
                "share": round(count * 100.0 / len(members), 1),
                "samples": len(carry_samples),
                "top4_samples": len(top4_samples),
                "min_stars_top4": min(stars),
                "avg_stars_top4": round(avg_number(stars) or 0, 2),
                "recommended_min_stars": recommended_star,
                "high_cost_three_star_dependency": high_cost_three_star_dependency,
                "two_star_rate": round(two_star_rate, 1),
                "three_star_rate": round(three_star_rate, 1),
                "three_item_rate": round(three_item_rate, 1),
                "bottom_underbuilt_rate": round(bottom_underbuilt_rate, 1),
            }
        )
    return requirements


def summarize_comp_carry_equipment(
    members: list[PlayerFeature],
    main_carries: list[tuple[str, int]],
) -> list[dict[str, Any]]:
    notes = []
    for hero_name, _ in main_carries[:3]:
        samples: list[tuple[PlayerFeature, Hero]] = []
        item_counter: Counter[str] = Counter()
        for member in members:
            carry = carry_for_name(member, hero_name)
            if carry is None:
                continue
            samples.append((member, carry))
            item_counter.update(equipment.name for equipment in carry.equipments)
        if len(samples) < 4:
            continue

        item_rows = []
        for item_name, appearances in item_counter.most_common():
            if appearances < 3:
                continue
            with_ranks = [
                member.rank
                for member, carry in samples
                if item_name in {equipment.name for equipment in carry.equipments}
            ]
            without_ranks = [
                member.rank
                for member, carry in samples
                if item_name not in {equipment.name for equipment in carry.equipments}
            ]
            if not with_ranks:
                continue
            with_avg = sum(with_ranks) / len(with_ranks)
            without_avg = sum(without_ranks) / len(without_ranks) if without_ranks else None
            with_top4 = sum(1 for rank in with_ranks if rank <= 4) * 100.0 / len(with_ranks)
            without_top4 = (
                sum(1 for rank in without_ranks if rank <= 4) * 100.0 / len(without_ranks)
                if without_ranks
                else None
            )
            penalty = (
                round(without_avg - with_avg, 2)
                if without_avg is not None
                else None
            )
            use_rate = appearances * 100.0 / len(samples)
            selected_rate = (
                sum(
                    1
                    for _, carry in samples
                    for equipment in carry.equipments
                    if equipment.name == item_name and equipment.selected
                )
                * 100.0
                / appearances
            )
            if appearances >= 8 and penalty is not None and penalty >= 0.45 and with_top4 >= 60:
                label = "疑似刚需"
            elif appearances >= 8 and penalty is not None and penalty >= 0.25:
                label = "高价值"
            else:
                label = "观察"
            item_rows.append(
                {
                    "equipment_name": item_name,
                    "label": label,
                    "appearances": appearances,
                    "use_rate": round(use_rate, 1),
                    "with_avg_rank": round(with_avg, 2),
                    "without_avg_rank": round(without_avg, 2) if without_avg is not None else None,
                    "without_item_penalty": penalty,
                    "with_top4_rate": round(with_top4, 1),
                    "without_top4_rate": round(without_top4, 1) if without_top4 is not None else None,
                    "selected_rate": round(selected_rate, 1),
                }
            )
        item_rows.sort(
            key=lambda row: (
                {"疑似刚需": 0, "高价值": 1, "观察": 2}[row["label"]],
                -row["appearances"],
                -(row["without_item_penalty"] or 0),
                -row["use_rate"],
            )
        )
        notes.append(
            {
                "hero_name": hero_name,
                "sample_count": len(samples),
                "items": item_rows[:5],
            }
        )
    return notes


def analyze_cards(
    features: list[PlayerFeature],
    comp_rows: list[dict[str, Any]],
    min_apps: int,
    baseline: float,
    team_baseline: float,
) -> dict[str, Any]:
    single_items: list[StatItem] = []
    first_card_items: list[StatItem] = []
    blue_team_rank_items: list[StatItem] = []
    blue_team_top2: Counter[str] = Counter()
    pair_items: list[StatItem] = []
    triple_items: list[StatItem] = []
    teammate_pair_items: list[StatItem] = []
    first_card_duo_items: list[StatItem] = []
    first_with_partner_any_items: list[StatItem] = []
    comp_card_items: dict[int, list[StatItem]] = defaultdict(list)
    family_labels = {row["family_id"]: row["label"] for row in comp_rows}
    by_match_rank = {
        (feature.match_id, feature.rank): feature
        for feature in features
    }
    seen_partner_pairs: set[tuple[int, int]] = set()
    duo_contribution: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "appearances": 0,
            "holder_rank_sum": 0.0,
            "team_rank_sum": 0.0,
            "team_top2": 0,
            "rank_gap_sum": 0.0,
        }
    )
    total_matches = len({feature.match_id for feature in features}) or 1

    for feature in features:
        cards = sorted(set(feature.cards))
        weight = feature.sample_weight
        for card in cards:
            single_items.append((card, feature.rank, weight))
            if card_prefix_type(card) == "蓝":
                blue_team_rank_items.append((card, team_rank_value(feature), weight))
                if team_rank_value(feature) <= 2:
                    blue_team_top2[card] += weight
            if feature.family_id:
                comp_card_items[feature.family_id].append((card, feature.rank, weight))
        if (card := first_card(feature)) is not None:
            first_card_items.append((card, feature.rank, weight))
        for pair in itertools.combinations(cards, 2):
            pair_items.append((" + ".join(pair), feature.rank, weight))
        for triple in itertools.combinations(cards, 3):
            triple_items.append((" + ".join(triple), feature.rank, weight))
        if feature.partner_player is not None:
            partner = by_match_rank.get((feature.match_id, int(feature.partner_player)))
            if partner is not None:
                pair_key = tuple(sorted((feature.player_id, partner.player_id)))
                if pair_key not in seen_partner_pairs:
                    seen_partner_pairs.add(pair_key)
                    team_rank = min(team_rank_value(feature), team_rank_value(partner))
                    team_weight = feature.sample_weight
                    for left in set(feature.cards):
                        for right in set(partner.cards):
                            teammate_pair_items.append(
                                (" + ".join(sorted((left, right))), team_rank, team_weight)
                            )
                    left_first = first_card(feature)
                    right_first = first_card(partner)
                    if left_first and right_first:
                        first_key = " + ".join(sorted((left_first, right_first)))
                        first_card_duo_items.append((first_key, team_rank, team_weight))
                        contribution = duo_contribution[first_key]
                        contribution["appearances"] += 1
                        contribution["holder_rank_sum"] += (feature.rank + partner.rank) / 2.0
                        contribution["team_rank_sum"] += team_rank
                        if team_rank <= 2:
                            contribution["team_top2"] += 1
                        contribution["rank_gap_sum"] += abs(feature.rank - partner.rank)
                    if left_first:
                        for card in set(partner.cards):
                            first_with_partner_any_items.append(
                                (f"{left_first} + 队友{card}", team_rank, team_weight)
                            )
                    if right_first:
                        for card in set(feature.cards):
                            first_with_partner_any_items.append(
                                (f"{right_first} + 队友{card}", team_rank, team_weight)
                            )

    by_comp = []
    for family_id, items in comp_card_items.items():
        rows = aggregate_key_stats(items, max(4, min_apps // 3), baseline)[:8]
        if rows:
            add_avg_appearances_per_match(rows, total_matches)
            by_comp.append(
                {
                    "family_id": family_id,
                    "family_label": family_labels.get(family_id, str(family_id)),
                    "cards": rows,
                }
            )

    contribution_rows = []
    for key, row in duo_contribution.items():
        n = row["appearances"]
        if n < max(5, min_apps // 2):
            continue
        holder_avg = row["holder_rank_sum"] / n
        team_avg = row["team_rank_sum"] / n
        contribution_rows.append(
            {
                "key": key,
                "appearances": n,
                "holder_avg_rank": round(holder_avg, 2),
                "team_avg_rank": round(team_avg, 2),
                "team_top2_rate": round(row["team_top2"] * 100.0 / n, 1),
                "team_lift_vs_baseline": round(team_baseline - team_avg, 2),
                "team_lift_vs_holder": round((holder_avg / 2.0) - team_avg, 2),
                "partner_delta": round(row["rank_gap_sum"] / n, 2),
            }
        )
    contribution_rows.sort(
        key=lambda item: (
            -item["team_lift_vs_baseline"],
            item["team_avg_rank"],
            -item["appearances"],
        )
    )

    single_cards, single_cards_by_prefix = aggregate_single_cards_by_catalog(
        single_items,
        baseline,
        load_report_card_catalog(),
        sample_first=True,
    )
    first_card_rankings, first_card_rankings_by_prefix = aggregate_key_stats_by_prefix(
        first_card_items, max(6, min_apps // 2), baseline, sample_first=True
    )
    blue_cards_team_rank, blue_cards_team_rank_by_prefix = aggregate_key_stats_by_prefix(
        blue_team_rank_items,
        max(6, min_apps // 2),
        team_baseline,
        sample_first=True,
    )
    add_avg_appearances_per_match(single_cards, total_matches)
    add_avg_appearances_to_prefix_groups(single_cards_by_prefix, total_matches)
    add_avg_appearances_per_match(first_card_rankings, total_matches)
    add_avg_appearances_to_prefix_groups(first_card_rankings_by_prefix, total_matches)
    add_avg_appearances_per_match(blue_cards_team_rank, total_matches)
    add_avg_appearances_to_prefix_groups(blue_cards_team_rank_by_prefix, total_matches)
    for row in blue_cards_team_rank:
        row["team_top2_rate"] = round(
            blue_team_top2[row["key"]] * 100.0 / max(row.get("weighted_appearances", row["appearances"]), 1e-9),
            1,
        )
    for rows in blue_cards_team_rank_by_prefix.values():
        for row in rows:
            row["team_top2_rate"] = round(
                blue_team_top2[row["key"]] * 100.0 / max(row.get("weighted_appearances", row["appearances"]), 1e-9),
                1,
            )

    return {
        "single_cards": single_cards,
        "single_cards_by_prefix": single_cards_by_prefix,
        "first_card_rankings": first_card_rankings,
        "first_card_rankings_by_prefix": first_card_rankings_by_prefix,
        "blue_cards_team_rank": blue_cards_team_rank,
        "blue_cards_team_rank_by_prefix": blue_cards_team_rank_by_prefix,
        "card_pairs_observation": aggregate_key_stats(pair_items, max(6, min_apps // 2), baseline)[:20],
        "card_triples_observation": aggregate_key_stats(triple_items, max(5, min_apps // 2), baseline)[:20],
        "teammate_card_pairs_observation": aggregate_key_stats(
            teammate_pair_items,
            max(6, min_apps // 2),
            team_baseline,
        )[:20],
        "first_card_duo_synergy": aggregate_key_stats(
            first_card_duo_items,
            max(5, min_apps // 2),
            team_baseline,
        )[:20],
        "first_with_partner_any_observation": aggregate_key_stats(
            first_with_partner_any_items,
            max(6, min_apps // 2),
            team_baseline,
        )[:20],
        "duo_card_contribution": contribution_rows[:20],
        "composition_cards": by_comp,
    }


def analyze_heroes_and_equipment(
    features: list[PlayerFeature],
    min_apps: int,
    baseline: float,
) -> dict[str, Any]:
    hero_stats: dict[str, RankStats] = defaultdict(RankStats)
    carry_stats: dict[str, RankStats] = defaultdict(RankStats)
    hero_item_stats: dict[tuple[str, str], RankStats] = defaultdict(RankStats)
    hero_item_selected: Counter[tuple[str, str]] = Counter()
    item_stats: dict[str, RankStats] = defaultdict(RankStats)
    set_stats: dict[tuple[str, str], RankStats] = defaultdict(RankStats)
    trait_items: list[StatItem] = []
    jiujiu_items: list[StatItem] = []
    hero_tiers: dict[str, int | None] = {}

    for feature in features:
        weight = feature.sample_weight
        for trait, tier in feature.active_traits.items():
            trait_items.append((f"{trait}-{tier}", feature.rank, weight))
        for trait, count in feature.jiujiu_bonus.items():
            if count > 0:
                jiujiu_items.append((trait, feature.rank, weight))
        for hero in feature.heroes:
            hero_stats[hero.name].add(feature.rank, feature.sample_weight)
            hero_tiers.setdefault(hero.name, hero.tier)
            if any(hero.id == candidate.id for candidate in feature.carry_candidates[:3]):
                carry_stats[hero.name].add(feature.rank, feature.sample_weight)
            equipment_names = []
            for equipment in hero.equipments:
                item_stats[equipment.name].add(feature.rank, feature.sample_weight)
                hero_item_stats[(hero.name, equipment.name)].add(feature.rank, feature.sample_weight)
                equipment_names.append(equipment.name)
                if equipment.selected:
                    hero_item_selected[(hero.name, equipment.name)] += 1
            if len(equipment_names) == 3:
                set_key = " + ".join(sorted(equipment_names))
                set_stats[(hero.name, set_key)].add(feature.rank, feature.sample_weight)

    heroes = []
    for hero_name, stat in hero_stats.items():
        if stat.appearances < min_apps:
            continue
        carry_stat = carry_stats.get(hero_name, RankStats())
        heroes.append(
            {
                "hero_name": hero_name,
                "tier": hero_tiers.get(hero_name),
                **stat.to_dict(baseline_rank=baseline, prior=8),
                "carry_appearances": carry_stat.appearances,
                "carry_rate": round(carry_stat.appearances * 100.0 / stat.appearances, 1),
            }
        )
    heroes.sort(
        key=lambda row: (
            -row["carry_appearances"],
            row["adjusted_avg_rank"],
            -row["top4_rate"],
        )
    )

    dict_character, _ = load_game_config()
    recommendations = []
    for hero in heroes:
        hero_name = hero["hero_name"]
        config_entry = dict_character.get(hero_name)
        hero_traits = [str(trait) for trait in config_entry[1:]] if config_entry else []
        items = []
        super_items = []
        food_items = []
        low_sample_items = []
        reliable_item_min = max(8, int(hero["appearances"] * 0.05))
        for (item_hero, item_name), stat in hero_item_stats.items():
            if item_hero != hero_name or stat.appearances < max(4, min_apps // 3):
                continue
            selected_rate = hero_item_selected[(item_hero, item_name)] * 100.0 / stat.appearances
            kind = equipment_kind(item_name)
            row = {
                "equipment_name": item_name,
                "equipment_kind": kind,
                **stat.to_dict(baseline_rank=baseline, prior=8),
                "selected_rate": round(selected_rate, 1),
                "sample_quality": "高样本" if stat.appearances >= reliable_item_min else "低样本观察",
                "selected_priority": selected_priority_label(selected_rate, stat.to_dict()["avg_rank"], baseline)
                if stat.appearances >= reliable_item_min
                else "低",
            }
            if kind == "super":
                super_items.append(row)
                continue
            if kind == "food":
                food_items.append(row)
                continue
            if stat.appearances >= reliable_item_min:
                items.append(row)
            else:
                low_sample_items.append(row)
        sets = []
        for (set_hero, set_name), stat in set_stats.items():
            if set_hero == hero_name and stat.appearances >= max(3, min_apps // 4):
                sets.append({"equipment_set": set_name, **stat.to_dict(baseline_rank=baseline, prior=8)})
        items.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        super_items.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        food_items.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        low_sample_items.sort(key=lambda row: (row["adjusted_avg_rank"], -row["top4_rate"], -row["appearances"]))
        sets.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        detail_items: list[dict[str, Any]] = []
        for (item_hero, item_name), stat in hero_item_stats.items():
            if item_hero != hero_name:
                continue
            if stat.appearances < HERO_EQUIPMENT_DETAIL_MIN_APPEARANCES:
                continue
            selected_rate = hero_item_selected[(item_hero, item_name)] * 100.0 / stat.appearances
            kind = equipment_kind(item_name)
            detail_items.append(
                {
                    "equipment_name": item_name,
                    "equipment_kind": kind,
                    **stat.to_dict(baseline_rank=baseline, prior=8),
                    "selected_rate": round(selected_rate, 1),
                    "sample_quality": (
                        "高样本" if stat.appearances >= reliable_item_min else "低样本观察"
                    ),
                    "selected_priority": selected_priority_label(
                        selected_rate, stat.to_dict()["avg_rank"], baseline
                    )
                    if stat.appearances >= reliable_item_min
                    else "低",
                }
            )
        detail_items.sort(
            key=lambda row: (
                -row["appearances"],
                row["adjusted_avg_rank"],
                -row["top4_rate"],
                row["equipment_name"],
            )
        )
        recommendations.append(
            {
                "hero_name": hero_name,
                "detail_slug": hero_equipment_detail_slug(hero_name),
                "hero_traits": hero_traits,
                "hero_stats": hero,
                "recommended_items": items[:6],
                "recommended_super_items": super_items[:4],
                "recommended_food_items": food_items[:4],
                "low_sample_observations": low_sample_items[:4],
                "recommended_sets": sets[:4],
                "detail_items": detail_items,
                "has_equipment_data": bool(items or super_items or food_items or low_sample_items),
            }
        )

    recommendations.sort(
        key=lambda row: (
            row["hero_stats"].get("tier") or 99,
            -row["hero_stats"].get("carry_appearances", 0),
            row["hero_name"],
        )
    )

    equipment_rows = []
    for item_name, stat in item_stats.items():
        if stat.appearances >= min_apps:
            equipment_rows.append({"equipment_name": item_name, **stat.to_dict(baseline_rank=baseline, prior=8)})
    equipment_rows.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], row["avg_rank"], -row["top4_rate"]))

    return {
        "heroes": heroes,
        "carry_equipment_recommendations": recommendations,
        "equipment": equipment_rows,
        "bonds": aggregate_key_stats(trait_items, min_apps, baseline),
        "jiujiu_bonds": aggregate_key_stats(jiujiu_items, max(5, min_apps // 3), baseline),
    }


def special_equipment_confidence(appearances: int, n_eff: float) -> str:
    if appearances >= SPECIAL_EQUIPMENT_RELIABLE_MIN and n_eff >= 5.0:
        return "高"
    if appearances >= 4 and n_eff >= 2.5:
        return "中"
    return "低"


def analyze_special_equipment(
    features: list[PlayerFeature],
    baseline: float,
    *,
    kind: str,
    always_include: frozenset[str] | None = None,
    min_apps: int = 1,
) -> dict[str, Any]:
    """Rank special equipment and recommend wearers with low-sample confidence tags."""
    item_stats: dict[str, RankStats] = defaultdict(RankStats)
    hero_item_stats: dict[tuple[str, str], RankStats] = defaultdict(RankStats)
    for feature in features:
        seen_items: set[str] = set()
        for hero in feature.heroes:
            for equipment in hero.equipments:
                item_name = equipment.name
                if equipment_kind(item_name) != kind:
                    continue
                hero_item_stats[(item_name, hero.name)].add(feature.rank, feature.sample_weight)
                if item_name in seen_items:
                    continue
                seen_items.add(item_name)
                item_stats[item_name].add(feature.rank, feature.sample_weight)

    force_include = always_include or frozenset()
    candidate_names = set(item_stats) | set(force_include)
    rankings: list[dict[str, Any]] = []
    for item_name in candidate_names:
        if kind == "super" and item_name not in SUPER_EQUIPMENT_NAMES:
            continue
        if kind == "food" and not is_food_equipment(item_name):
            continue
        stat = item_stats.get(item_name, RankStats())
        if stat.appearances < min_apps and item_name not in force_include:
            continue
        stats = (
            stat.to_dict(baseline_rank=baseline, prior=8)
            if stat.appearances
            else {
                "appearances": 0,
                "weighted_appearances": 0.0,
                "n_eff": 0.0,
                "avg_rank": None,
                "win_rate": None,
                "top2_rate": None,
                "top4_rate": None,
                "adjusted_avg_rank": None,
            }
        )
        wearer_min = (
            1
            if stats["appearances"] < SPECIAL_EQUIPMENT_RELIABLE_MIN
            else SPECIAL_EQUIPMENT_WEARER_MIN
        )
        wearers: list[dict[str, Any]] = []
        for (wearer_item, hero_name), wearer_stat in hero_item_stats.items():
            if wearer_item != item_name or wearer_stat.appearances < wearer_min:
                continue
            wearers.append(
                {
                    "hero_name": hero_name,
                    **wearer_stat.to_dict(baseline_rank=baseline, prior=6),
                    "share": round(
                        wearer_stat.appearances * 100.0 / max(stat.appearances, 1),
                        1,
                    ),
                }
            )
        wearers.sort(
            key=lambda row: (
                -row["appearances"],
                row.get("adjusted_avg_rank", 99),
                -row.get("top4_rate", 0),
            )
        )
        confidence = special_equipment_confidence(
            stats["appearances"],
            float(stats.get("n_eff") or 0.0),
        )
        sample_quality = (
            "高样本"
            if stats["appearances"] >= SPECIAL_EQUIPMENT_RELIABLE_MIN
            else "低样本观察"
        )
        note = ""
        if item_name == "岛好锅" and stats["appearances"] < SPECIAL_EQUIPMENT_RELIABLE_MIN:
            note = "名称样本极少，低置信观察，勿作高置信推荐"
        elif confidence == "低":
            note = "样本不足，仅供观察"
        rankings.append(
            {
                "equipment_name": item_name,
                "equipment_kind": kind,
                **stats,
                "confidence": confidence,
                "sample_quality": sample_quality,
                "note": note,
                "recommended_wearers": wearers[:5],
            }
        )

    rankings.sort(
        key=lambda row: (
            0 if row["appearances"] >= SPECIAL_EQUIPMENT_RELIABLE_MIN else 1,
            row.get("adjusted_avg_rank") if row.get("adjusted_avg_rank") is not None else 99,
            -(row.get("top4_rate") or 0),
            -row["appearances"],
            row["equipment_name"],
        )
    )
    for index, row in enumerate(rankings, start=1):
        row["strength_rank"] = index
    catalog = (
        sorted(SUPER_EQUIPMENT_NAMES)
        if kind == "super"
        else sorted(FOOD_SPECIAL_EQUIPMENT_NAMES)
    )
    return {
        "definition": (
            "超级装备按固定白名单统计；排序优先修正名次，高样本优先于低样本观察。"
            if kind == "super"
            else "美食社装备含美味/绝味/暗黑前缀及杏仁豆腐/椒盐酥糖/岛好锅；低样本仅观察。"
        ),
        "catalog": catalog,
        "rankings": rankings,
    }


def format_active_count_distribution(counter: Counter[int | str]) -> str:
    parts = [
        f"{count}({appearances})"
        for count, appearances in sorted(
            counter.items(),
            key=lambda item: (-item[1], -int(item[0]) if str(item[0]).isdigit() else 0),
        )
    ]
    return " / ".join(parts)


def format_primary_bond_source_distribution(
    distribution: dict[str, int] | None,
) -> str:
    if not distribution:
        return "—"
    return " / ".join(
        f"{source}({appearances})"
        for source, appearances in sorted(
            distribution.items(), key=lambda item: (-item[1], item[0])
        )
    )


def analyze_primary_bond_strength(
    features: list[PlayerFeature],
    min_apps: int,
    baseline: float,
) -> dict[str, Any]:
    stats: dict[str, RankStats] = defaultdict(RankStats)
    count_distribution: dict[str, Counter[int]] = defaultdict(Counter)
    tier_distribution: dict[str, Counter[int]] = defaultdict(Counter)
    source_distribution: dict[str, Counter[str]] = defaultdict(Counter)
    category_distribution: Counter[str] = Counter()
    overall_source_distribution: Counter[str] = Counter()
    _, bond_thresholds = load_game_config()

    for feature in features:
        for selection in primary_bond_business_selections(feature, bond_thresholds):
            category = selection["category"]
            activation_count = selection["activation_count"]
            selected_tier = selection["active_tier"]
            source = selection["source"]
            stats[category].add(feature.rank, feature.sample_weight)
            count_distribution[category][activation_count] += 1
            tier_distribution[category][selected_tier] += 1
            source_distribution[category][source] += 1
            category_distribution[category] += 1
            overall_source_distribution[source] += 1

    rows: list[dict[str, Any]] = []
    for bond, stat in stats.items():
        if stat.appearances < min_apps:
            continue
        row = {
            "bond": bond,
            "category": bond,
            **stat.to_dict(baseline_rank=baseline, prior=8),
            "bottom4_rate": round(
                (stat.weighted_appearances - stat.top4) * 100.0 / max(stat.weighted_appearances, 1e-9),
                1,
            ),
            "active_count_distribution": {
                str(count): appearances
                for count, appearances in sorted(count_distribution[bond].items())
            },
            "active_tier_distribution": {
                str(tier): appearances
                for tier, appearances in sorted(tier_distribution[bond].items())
            },
            "source_distribution": dict(sorted(source_distribution[bond].items())),
            "common_activation_summary": format_active_count_distribution(count_distribution[bond]),
            "common_tier_summary": format_active_count_distribution(tier_distribution[bond]),
        }
        rows.append(row)

    rows.sort(
        key=lambda row: (
            row["avg_rank"],
            -row["top4_rate"],
            -row["appearances"],
            row["bond"],
        )
    )
    for index, row in enumerate(rows, start=1):
        row["strength_rank"] = index

    return {
        "definition": (
            "主羁绊榜采用业务归类：学习社达到 config_s2.py 阈值列表的第三档（4学习）时"
            "独占归类为学习社，覆盖美食社收菜与其他全部业务归类；否则有收菜装备或原型为"
            "美食社收菜的棋盘计入美食社；普通事实羁绊须达到对应阈值列表的第二档，合格羁绊"
            "按最终激活人数（含啾啾加成）取最大值，人数并列则同时计入；无合格羁绊且原型为"
            "高费拼多多时归入高费拼多多，普通一级散搭不入榜。PlayerFeature 与阵容的"
            "factual main_bond 不变。"
        ),
        "source_definition": {
            "study_override": "学习社达到配置第三档（4学习）时独占归类",
            "food_harvest": "收菜装备或美食社收菜原型",
            "qualified_bond": "达到配置第二档的事实羁绊",
            "high_cost_pdd": "无合格事实羁绊的高费拼多多兜底",
        },
        "source_distribution": dict(sorted(overall_source_distribution.items())),
        "category_distribution": dict(sorted(category_distribution.items())),
        "rows": rows,
    }


def analyze_jiujiu(
    features: list[PlayerFeature],
    comp_rows: list[dict[str, Any]],
    baseline: float,
    min_apps: int = 5,
) -> dict[str, Any]:
    total_stats: dict[str, RankStats] = defaultdict(RankStats)
    effective_stats: dict[str, RankStats] = defaultdict(RankStats)
    incidental_stats: dict[str, RankStats] = defaultdict(RankStats)
    reason_counts: dict[str, Counter[str]] = defaultdict(Counter)
    player_to_comps: dict[int, dict[str, dict[str, Any]]] = defaultdict(dict)
    comp_item_stats: dict[tuple[str, str], RankStats] = defaultdict(RankStats)
    comp_item_wearers: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    hero_item_stats: dict[tuple[str, str], RankStats] = defaultdict(RankStats)
    item_strategy_seen: dict[str, set[str]] = defaultdict(set)

    for comp in comp_rows:
        for player_id in comp["member_player_ids"]:
            player_to_comps[player_id].setdefault(comp["label"], comp)

    pending_generalist: dict[str, list[tuple[PlayerFeature, str]]] = defaultdict(list)
    for feature in features:
        comp_values = list(player_to_comps.get(feature.player_id, {}).values())
        seen_items: set[str] = set()
        for hero in feature.heroes:
            for equipment in hero.equipments:
                trait = jiujiu_trait(equipment.name)
                if not trait:
                    continue
                item_name = equipment.name
                if item_name in seen_items:
                    continue
                seen_items.add(item_name)
                total_stats[item_name].add(feature.rank, feature.sample_weight)
                reasons = classify_jiujiu_sample(feature, hero, item_name, trait, comp_values)
                if reasons:
                    for reason in reasons:
                        reason_counts[item_name][reason] += 1
                    effective_stats[item_name].add(feature.rank, feature.sample_weight)
                    if "final_bond" in reasons:
                        for comp in comp_values:
                            if jiujiu_matches_strategy_bond(trait, comp):
                                comp_item_stats[(item_name, comp["label"])].add(feature.rank, feature.sample_weight)
                                comp_item_wearers[(item_name, comp["label"])][hero.name] += 1
                    if "hero_boost" in reasons:
                        hero_item_stats[(item_name, hero.name)].add(feature.rank, feature.sample_weight)
                else:
                    incidental_stats[item_name].add(feature.rank, feature.sample_weight)
                    pending_generalist[item_name].append((feature, item_name))
                for comp in comp_values:
                    item_strategy_seen[item_name].add(comp["label"])

    baseline_rank = baseline
    for item_name, samples in pending_generalist.items():
        if len(item_strategy_seen[item_name]) < 4:
            continue
        total = total_stats[item_name]
        if total.appearances < 12 or total.to_dict()["avg_rank"] > baseline_rank - 0.25:
            continue
        for feature, _ in samples:
            reason_counts[item_name]["generalist"] += 1
            effective_stats[item_name].add(feature.rank, feature.sample_weight)

    rankings = []
    recommended: dict[str, list[dict[str, Any]]] = {}
    hero_recommendations: dict[str, list[dict[str, Any]]] = {}
    for item_name, stat in total_stats.items():
        if stat.appearances < min_apps:
            continue
        comps = []
        for (comp_item, comp_label), comp_stat in comp_item_stats.items():
            if comp_item != item_name or comp_stat.appearances < 3:
                continue
            comps.append(
                {
                    "family_label": comp_label,
                    **comp_stat.to_dict(baseline_rank=baseline, prior=6),
                    "share": round(comp_stat.appearances * 100.0 / stat.appearances, 1),
                    "recommended_wearers": [
                        {
                            "hero_name": hero_name,
                            "appearances": count,
                            "share": round(count * 100.0 / comp_stat.appearances, 1),
                        }
                        for hero_name, count in comp_item_wearers[
                            (comp_item, comp_label)
                        ].most_common(3)
                    ],
                }
            )
        comps.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        recommended[item_name] = comps[:4]
        hero_rows = []
        for (hero_item, hero_name), hero_stat in hero_item_stats.items():
            if hero_item != item_name or hero_stat.appearances < 3:
                continue
            hero_rows.append(
                {
                    "hero_name": hero_name,
                    **hero_stat.to_dict(baseline_rank=baseline, prior=6),
                }
            )
        hero_rows.sort(key=lambda row: (-row["appearances"], row["adjusted_avg_rank"], -row["top4_rate"]))
        hero_recommendations[item_name] = hero_rows[:4]
        effective = effective_stats.get(item_name, RankStats())
        incidental = incidental_stats.get(item_name, RankStats())
        rankings.append(
            {
                "equipment_name": item_name,
                **stat.to_dict(baseline_rank=baseline, prior=8),
                "effective_appearances": effective.appearances,
                "effective_rate": round(effective.appearances * 100.0 / stat.appearances, 1),
                "effective_stats": effective.to_dict(baseline_rank=baseline, prior=8) if effective.appearances else None,
                "incidental_stats": incidental.to_dict(baseline_rank=baseline, prior=8) if incidental.appearances else None,
                "reason_counts": dict(reason_counts[item_name]),
                "recommended_comps": recommended[item_name],
                "recommended_heroes": hero_recommendations[item_name],
            }
        )
    rankings.sort(key=lambda row: (-row["effective_appearances"], row["effective_stats"]["adjusted_avg_rank"] if row["effective_stats"] else 99, -row["top4_rate"]))
    return {
        "jiujiu_rankings": rankings,
        "jiujiu_recommended_comps": recommended,
        "jiujiu_recommended_heroes": hero_recommendations,
    }


def jiujiu_matches_strategy_bond(trait: str, comp: dict[str, Any]) -> bool:
    main_bond = comp.get("main_bond", "")
    if main_bond != "拼多多" and "-" in main_bond and parse_trait_tier(main_bond)[0] == trait:
        return True
    for bond in comp.get("common_bonds", [])[:4]:
        if "-" in bond["bond"]:
            bond_trait, _ = parse_trait_tier(bond["bond"])
            if bond_trait == trait and bond.get("share", 0) >= 50:
                return True
    return False


def classify_jiujiu_sample(
    feature: PlayerFeature,
    hero: Hero,
    item_name: str,
    trait: str,
    comps: list[dict[str, Any]],
) -> list[str]:
    reasons: list[str] = []
    if any(jiujiu_matches_strategy_bond(trait, comp) for comp in comps):
        reasons.append("final_bond")
    is_key_hero = any(hero.id == candidate.id for candidate in feature.carry_candidates[:3])
    if is_key_hero and feature.rank <= 4 and hero.equipment_count >= 2:
        reasons.append("hero_boost")
    return reasons


def analyze_duo_composition_synergy(
    features: list[PlayerFeature],
    comp_rows: list[dict[str, Any]],
    team_baseline: float,
    min_apps: int = 5,
) -> list[dict[str, Any]]:
    player_to_comp: dict[int, dict[str, Any]] = {}
    for comp in comp_rows:
        for player_id in comp.get("member_player_ids", []):
            player_to_comp[player_id] = comp

    by_match_rank = {
        (feature.match_id, feature.rank): feature
        for feature in features
    }
    pair_stats: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "appearances": 0,
            "team_rank_sum": 0.0,
            "team_top2": 0,
            "team_wins": 0,
            "individual_rank_sum": 0.0,
            "strategy_labels": None,
        }
    )
    seen_pairs: set[tuple[int, int]] = set()
    for feature in features:
        if feature.partner_player is None:
            continue
        partner = by_match_rank.get((feature.match_id, int(feature.partner_player)))
        if partner is None:
            continue
        pair_key = tuple(sorted((feature.player_id, partner.player_id)))
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)
        left = player_to_comp.get(feature.player_id)
        right = player_to_comp.get(partner.player_id)
        if not left or not right:
            continue
        labels = tuple(sorted((left["label"], right["label"])))
        team_rank = min(team_rank_value(feature), team_rank_value(partner))
        row = pair_stats[labels]
        row["appearances"] += 1
        row["team_rank_sum"] += team_rank
        row["individual_rank_sum"] += (feature.rank + partner.rank) / 2.0
        row["strategy_labels"] = labels
        if team_rank <= 2:
            row["team_top2"] += 1
        if team_rank == 1:
            row["team_wins"] += 1

    rows: list[dict[str, Any]] = []
    for labels, row in pair_stats.items():
        n = row["appearances"]
        if n < min_apps:
            continue
        team_avg = row["team_rank_sum"] / n
        holder_avg = row["individual_rank_sum"] / n
        rows.append(
            {
                "strategy_a": labels[0],
                "strategy_b": labels[1],
                "key": " + ".join(labels),
                "appearances": n,
                "team_avg_rank": round(team_avg, 2),
                "team_top2_rate": round(row["team_top2"] * 100.0 / n, 1),
                "team_win_rate": round(row["team_wins"] * 100.0 / n, 1),
                "team_lift_vs_baseline": round(team_baseline - team_avg, 2),
                "holder_avg_rank": round(holder_avg, 2),
                "confidence": confidence_label(n),
            }
        )
    rows.sort(
        key=lambda row: (
            -row["appearances"],
            row["team_avg_rank"],
            -row["team_top2_rate"],
        )
    )
    return rows[:20]


def selected_priority_label(selected_rate: float, avg_rank: float, baseline: float) -> str:
    if selected_rate >= 30 and avg_rank < baseline:
        return "高"
    if selected_rate >= 12 and avg_rank <= baseline + 0.2:
        return "中"
    return "低"


def find_card_traps(
    cards_by_prefix: dict[str, list[dict[str, Any]]],
    baseline: float,
) -> list[dict[str, Any]]:
    def weak(row: dict[str, Any]) -> bool:
        return row.get("adjusted_avg_rank", row.get("avg_rank", 0)) >= baseline + 0.45 or row.get(
            "top4_rate", 100
        ) <= 42

    traps: list[dict[str, Any]] = []
    for prefix_type in CARD_PREFIX_TYPES:
        for row in cards_by_prefix.get(prefix_type, []):
            if row["appearances"] >= 12 and weak(row):
                traps.append(
                    {
                        **row,
                        "trap_reason": f"{prefix_type}类内样本充足但表现偏弱",
                    }
                )
    traps.sort(key=lambda row: (-row["appearances"], -row["adjusted_avg_rank"]))
    return traps[:10]


def find_traps(
    comp_rows: list[dict[str, Any]],
    hero_rows: list[dict[str, Any]],
    card_rows: list[dict[str, Any]],
    cards_by_prefix: dict[str, list[dict[str, Any]]],
    bond_rows: list[dict[str, Any]],
    equipment_rows: list[dict[str, Any]],
    baseline: float,
) -> dict[str, list[dict[str, Any]]]:
    def weak(row: dict[str, Any]) -> bool:
        return row.get("adjusted_avg_rank", row.get("avg_rank", 0)) >= baseline + 0.45 or row.get("top4_rate", 100) <= 42

    def comp_trait(row: dict[str, Any]) -> str | None:
        key = row.get("main_bond", "")
        if not key or key == "拼多多" or "-" not in key:
            return None
        return parse_trait_tier(key)[0]

    strong_trait_tiers: dict[str, int] = {}
    for row in comp_rows:
        trait = comp_trait(row)
        if (
            trait
            and row["stats"]["appearances"] >= 20
            and row["stats"]["top4_rate"] >= 60
            and row["stats"]["avg_rank"] <= baseline
        ):
            _, tier = parse_trait_tier(row["main_bond"])
            strong_trait_tiers[trait] = max(strong_trait_tiers.get(trait, 0), tier)
    strong_traits = set(strong_trait_tiers)

    comp_traps = [
        row
        for row in comp_rows
        if row["stats"]["appearances"] >= 5
        and (row["stats"]["avg_rank"] >= baseline + 0.45 or row["stats"]["top4_rate"] <= 42)
        and comp_trait(row) not in strong_traits
    ]
    for row in comp_traps:
        row["trap_reason"] = "策略整体表现偏弱，且没有同羁绊强势大成形态覆盖"
    comp_traps.sort(key=lambda row: (-row["popularity"]["pick_rate"], -row["stats"]["avg_rank"]))

    bond_traps: list[dict[str, Any]] = []
    covered_bond_pressure: list[dict[str, Any]] = []
    for row in bond_rows:
        if row["appearances"] < 10 or not weak(row):
            continue
        key = row.get("key", "")
        if "-" in key:
            trait, tier = parse_trait_tier(key)
            mature_tier = strong_trait_tiers.get(trait)
            if mature_tier and tier < mature_tier:
                covered_bond_pressure.append(
                    {
                        **row,
                        "covered_by": f"{trait}-{mature_tier}",
                        "trap_reason": f"更像{trait}-{mature_tier}的未成型阶段，计入成型压力而非独立陷阱",
                    }
                )
                continue
        bond_traps.append(row)

    return {
        "compositions": comp_traps[:10],
        "heroes": [row for row in hero_rows if row["appearances"] >= 10 and weak(row)][:10],
        "cards": find_card_traps(cards_by_prefix, baseline),
        "bonds": bond_traps[:10],
        "formation_pressure_bonds": covered_bond_pressure[:10],
        "equipment": [row for row in equipment_rows if row["appearances"] >= 10 and weak(row)][:10],
    }


def extract_balance_boundary(
    notes_path: Path | None,
    features: list[PlayerFeature],
) -> dict[str, Any]:
    """Read only explicit boundary metadata; never infer dates from prose."""
    base = {
        "requested": notes_path is not None,
        "supported": False,
        "batch": None,
        "effective_date": None,
        "source": None,
        "reason": "未提供平衡公告，使用保守滚动窗口",
        "parser_policy": "only explicit JSON/front-matter batch or effective_date fields",
    }
    if notes_path is None:
        return base
    path = notes_path if notes_path.is_absolute() else ROOT / notes_path
    text = path.read_text(encoding="utf-8")
    batch: str | None = None
    effective_date: str | None = None
    source: str | None = None
    if path.suffix.lower() == ".json":
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            payload = None
        if isinstance(payload, dict):
            raw_batch = payload.get("batch")
            raw_date = payload.get("effective_date")
            if isinstance(raw_batch, str) and re.fullmatch(r"\d{4}", raw_batch):
                batch, source = raw_batch, "json.batch"
            elif isinstance(raw_date, str):
                try:
                    parsed = date.fromisoformat(raw_date)
                except ValueError:
                    parsed = None
                if parsed:
                    effective_date = parsed.isoformat()
                    batch, source = parsed.strftime("%m%d"), "json.effective_date"
    else:
        header = text[:2000]
        batch_match = re.search(
            r"(?mi)^(?:batch|effective_batch|公告批次)\s*:\s*(\d{4})\s*$",
            header,
        )
        date_match = re.search(
            r"(?mi)^(?:effective_date|公告日期)\s*:\s*(\d{4}-\d{2}-\d{2})\s*$",
            header,
        )
        if batch_match:
            batch, source = batch_match.group(1), "explicit_text.batch"
        elif date_match:
            try:
                parsed = date.fromisoformat(date_match.group(1))
            except ValueError:
                parsed = None
            if parsed:
                effective_date = parsed.isoformat()
                batch, source = parsed.strftime("%m%d"), "explicit_text.effective_date"
    observed = ordered_batches(features)
    if not batch:
        return {
            **base,
            "reason": "公告中没有显式 batch/effective_date 元数据；未从正文猜测日期",
        }
    if batch not in observed:
        return {
            **base,
            "batch": batch,
            "effective_date": effective_date,
            "source": source,
            "reason": "已读取显式公告边界，但该批次不在当前数据中，回退滚动窗口",
        }
    return {
        **base,
        "supported": True,
        "batch": batch,
        "effective_date": effective_date,
        "source": source,
        "reason": "使用显式公告边界做前后窗口对比",
    }


def extract_balance_targets(
    notes_path: Path | None,
    dict_character: dict[str, list[Any]],
    dict_bond: dict[str, list[int]],
    observed_names: set[str],
) -> dict[str, list[str]]:
    if notes_path is None:
        return {"heroes": [], "bonds": [], "equipment_or_cards": []}
    path = notes_path if notes_path.is_absolute() else ROOT / notes_path
    text = path.read_text(encoding="utf-8")
    heroes = sorted(name for name in dict_character if name in text)
    bonds = sorted(name for name in dict_bond if name in text)
    others = sorted(name for name in observed_names if name in text and name not in heroes and name not in bonds)
    return {"heroes": heroes, "bonds": bonds, "equipment_or_cards": others}


def build_analysis(args: argparse.Namespace) -> dict[str, Any]:
    db_path = find_latest_db(args.db)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        dict_character, dict_bond = load_game_config()
        bot_ids = find_bot_player_ids(conn)
        quality = data_quality(conn, bot_ids)
        validation = validate_config(conn, dict_character, dict_bond)
        features = load_player_features(conn, bot_ids, dict_character, dict_bond)
        if not features:
            raise SystemExit("No usable player records after filtering.")
        half_life = getattr(args, "recency_half_life_days", DEFAULT_RECENCY_HALF_LIFE_DAYS)
        compute_sample_weights(features, half_life_days=half_life)
        total_weight = sum(feature.sample_weight for feature in features) or 1.0
        baseline = (
            sum(feature.rank * feature.sample_weight for feature in features) / total_weight
        )
        team_baseline = (
            sum(team_rank_value(feature) * feature.sample_weight for feature in features)
            / total_weight
        )
        recency = recency_overview(features, half_life)
        balance_boundary = extract_balance_boundary(args.balance_notes, features)
        stage_rows = cluster_compositions(features, args.min_comp_apps)
        attach_composition_trends(
            stage_rows,
            features,
            balance_boundary=balance_boundary,
        )
        comp_rows = merge_comp_strategies(stage_rows, features)
        calibrate_composition_confidence(comp_rows, features)
        trend_methodology = attach_composition_trends(
            comp_rows,
            features,
            balance_boundary=balance_boundary,
        )
        low_cost_carry_difficulty = enrich_three_star_contest(comp_rows, features)
        hero_equipment = analyze_heroes_and_equipment(features, args.min_entity_apps, baseline)
        super_equipment = analyze_special_equipment(
            features,
            baseline,
            kind="super",
            always_include=SUPER_EQUIPMENT_NAMES,
        )
        food_equipment = analyze_special_equipment(
            features,
            baseline,
            kind="food",
            always_include=FOOD_SPECIAL_EQUIPMENT_NAMES,
        )
        primary_bond_strength = analyze_primary_bond_strength(
            features,
            args.min_entity_apps,
            baseline,
        )
        cards = analyze_cards(features, comp_rows, args.min_card_apps, baseline, team_baseline)
        jiujiu_analysis = analyze_jiujiu(features, comp_rows, baseline)
        duo_compositions = analyze_duo_composition_synergy(
            features,
            comp_rows,
            team_baseline,
            min_apps=max(4, args.min_comp_apps // 2),
        )
        observed_names = {
            hero.name for feature in features for hero in feature.heroes
        } | {
            card for feature in features for card in feature.cards
        } | {
            equipment.name for feature in features for hero in feature.heroes for equipment in hero.equipments
        }
        balance_targets = extract_balance_targets(
            args.balance_notes,
            dict_character,
            dict_bond,
            observed_names,
        )
        traps = find_traps(
            comp_rows,
            hero_equipment["heroes"],
            cards["single_cards"],
            cards["single_cards_by_prefix"],
            hero_equipment["bonds"],
            hero_equipment["equipment"],
            baseline,
        )
        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "data_source": rel(db_path),
            "methodology": {
                "implementation": "rebuilt skill analyzer",
                "bot_filter": "rank 7/8 paired players excluded from all rankings",
                "unknown_filter": "unknown heroes/cards/equipment excluded from reference statistics",
                "equipment_normalization": "核选 prefix removed for equipment identity; selected rate retained",
                "super_equipment": sorted(SUPER_EQUIPMENT_NAMES),
                "food_equipment": (
                    "prefix 美味/绝味/暗黑 plus exact names "
                    + "、".join(sorted(FOOD_SPECIAL_EQUIPMENT_NAMES))
                ),
                "jiujiu_rule": "X啾啾 adds +1 to bond X when X exists in dict_bond",
                "primary_bond_rule": (
                    "food harvest -> 美食社; otherwise factual bonds must reach the configured "
                    "second threshold and ties by activation count are retained; high-cost PDD "
                    "is used only when no factual bond qualifies"
                ),
                "carry_score": "equipment_count*30 + selected_count*12 + stars*10 + tier*2 + max(0, 8-slot_index)*1.5",
                "play_style_rule": (
                    "any lineup 1/2/3-cost 3-star is 赌狗; level<=6 is 赌狗; "
                    "level 7 with low-cost main carry is 赌狗; otherwise 高费. "
                    "Strategy recommendation buckets follow mature-stage play_style "
                    "and force 赌狗 when mature carry advice requires a low-cost 3-star. "
                    "高费拼多多 requires 高费 play_style, no low-cost 3-star, enough 4/5-cost "
                    "units, no stable deep trait, and a 2-star+ 4/5-cost main carry"
                ),
                "archetype_rule": (
                    "美食社收菜 (food harvest) > 高费拼多多 (high-cost scattered, no low-cost "
                    "3-star, 2-star+ 4/5-cost main carry) > 羁绊运营:X > 拼多多"
                ),
                "card_order": "cards are ordered by slot_index; cards[0] is treated as the first/duo card",
                "card_prefix_ranking": "single/first-card rankings are grouped by card template prefix (彩/黄/蓝/白/其他) and ranked within each group",
                "team_rank": "per match, teams are ranked 1..N by each team's best individual rank",
                "card_granted_heroes": sorted(CARD_GRANTED_HEROES),
                "recency_weighting": {
                    "enabled": True,
                    "batch_source": recency["source"],
                    "half_life_days": half_life,
                    "min_weight": MIN_RECENCY_WEIGHT,
                    "latest_batch": recency["latest_batch"],
                    "batch_range": recency["batch_range"],
                    "note": "批次日期来自 path 中的 screenshots.MMDD；均分/胜率/前四率按加权样本计算，样本阈值仍看原始 n。",
                },
                "version_tracking": {
                    **trend_methodology,
                    "balance_boundary": balance_boundary,
                    "note": "公告边界仅接受显式结构化元数据；无法可靠取得时保守回退滚动窗口。",
                },
                "min_samples": {
                    "composition": args.min_comp_apps,
                    "entity": args.min_entity_apps,
                    "card": args.min_card_apps,
                },
            },
            "overview": {
                "quality": quality,
                "filtered_players": len(features),
                "effective_sample_weight": recency["effective_sample_weight"],
                "recency": recency,
                "version_tracking": {
                    **trend_methodology,
                    "balance_boundary": balance_boundary,
                },
                "baseline_rank": round(baseline, 3),
                "team_baseline_rank": round(team_baseline, 3),
                "validation": validation,
            },
            "rankings": {
                "compositions": comp_rows,
                "composition_recommendations": build_composition_recommendations(comp_rows),
                "composition_stages": stage_rows,
                "low_cost_carry_three_star_difficulty": low_cost_carry_difficulty,
                "duo_composition_synergy": duo_compositions,
                "cards": cards,
                "heroes_and_equipment": hero_equipment,
                "super_equipment": super_equipment,
                "food_equipment": food_equipment,
                "primary_bond_strength": primary_bond_strength,
                "jiujiu": jiujiu_analysis,
                "traps": traps,
                "balance_targets": balance_targets,
            },
        }
    finally:
        conn.close()


def append_hero_equipment_block(lines: list[str], row: dict[str, Any]) -> None:
    hero = row["hero_stats"]
    tier = hero.get("tier")
    tier_label = f"{tier}费，" if tier is not None else ""
    lines.append(
        f"### {row['hero_name']}（{tier_label}主C率 {render_pct(hero['carry_rate'])}，"
        f"avg {hero['avg_rank']:.2f}，n={hero['appearances']}）"
    )
    if row.get("recommended_items"):
        lines.append("")
        lines.append("| 装备 | 修正名次 | 核选占比 | 核选优先级 | 样本 |")
        lines.append("| --- | ---: | ---: | --- | ---: |")
        for item in row["recommended_items"][:6]:
            lines.append(
                f"| {item['equipment_name']} | {item['adjusted_avg_rank']:.2f} | "
                f"{render_pct(item['selected_rate'])} | {item['selected_priority']} | {item['appearances']} |"
            )
    elif row.get("low_sample_observations"):
        lines.append("")
        lines.append("| 装备 | 修正名次 | 核选占比 | 核选优先级 | 样本 |")
        lines.append("| --- | ---: | ---: | --- | ---: |")
        for item in row["low_sample_observations"][:6]:
            lines.append(
                f"| {item['equipment_name']} | {item['adjusted_avg_rank']:.2f} | "
                f"{render_pct(item['selected_rate'])} | {item['selected_priority']} | {item['appearances']} |"
            )
    else:
        lines.append("- 出装样本不足：该棋子在过滤后样本中缺少足够单件出装记录，多为副C/前排或出装不完整。")
    if row.get("recommended_items") and row.get("low_sample_observations"):
        obs = " / ".join(
            f"{item['equipment_name']}(修正{item['adjusted_avg_rank']:.2f}, n={item['appearances']})"
            for item in row["low_sample_observations"][:3]
        )
        lines.append(f"- 低样本观察：{obs}")
    if row.get("recommended_super_items"):
        super_items = " / ".join(
            f"{item['equipment_name']}(修正{item['adjusted_avg_rank']:.2f}, n={item['appearances']})"
            for item in row["recommended_super_items"][:3]
        )
        lines.append(f"- 超级装备：{super_items}")
    if row.get("recommended_food_items"):
        food_items = " / ".join(
            f"{item['equipment_name']}(修正{item['adjusted_avg_rank']:.2f}, n={item['appearances']})"
            for item in row["recommended_food_items"][:3]
        )
        lines.append(f"- 美食社装备：{food_items}")
    if row["recommended_sets"]:
        sets = "；".join(
            f"{item['equipment_set']}({item['adjusted_avg_rank']:.2f}, n={item['appearances']})"
            for item in row["recommended_sets"][:3]
        )
        lines.append(f"- 常见三件套：{sets}")
    lines.append("")


def render_pct(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.1f}%"


def render_card_metric(value: float | None, *, digits: int = 2) -> str:
    if value is None:
        return "—"
    return f"{value:.{digits}f}"


def format_trend_label(trend: dict[str, Any] | None) -> str:
    if not trend:
        return "—"
    label = trend.get("label", "—")
    if label == "insufficient":
        return "样本不足"
    return str(label)


def format_archetype_signals_text(comp: dict[str, Any], *, limit: int = 6) -> str:
    signals = comp.get("archetype_signals") or []
    if not signals:
        return "无明确归类信号"
    parts: list[str] = []
    for signal in signals[:limit]:
        signal_type = signal.get("type", "")
        if signal_type == "美食装备":
            equipment = signal.get("equipment") or {}
            parts.append(
                f"{equipment.get('hero_name', '—')}·{equipment.get('equipment_name', '—')}"
            )
        elif signal_type == "高费散羁绊":
            parts.append(
                f"高费散羁绊（{signal.get('four_five_cost_count', '—')}"
                f"/{signal.get('threshold', '—')}；"
                f"主C{signal.get('main_carry_tier', '—')}费"
                f"{signal.get('main_carry_stars', '—')}星；"
                f"低费三星{signal.get('low_cost_three_star_count', 0)}）"
            )
        elif signal_type == "羁绊投入":
            parts.append(f"羁绊投入·{signal.get('trait', '—')}")
        else:
            parts.append(signal.get("detail") or signal.get("reason") or signal_type or str(signal))
    return "；".join(parts)


def format_confidence_criteria_text(evidence: dict[str, Any] | None) -> str:
    if not evidence:
        return "样本信息不足"
    criteria = evidence.get("recommendation_criteria", {})
    parts: list[str] = []
    for key, label in RECOMMENDATION_CRITERION_LABELS.items():
        item = criteria.get(key, {})
        if not item:
            continue
        status = "达标" if item.get("met") else "未达标"
        required = item.get("minimum", item.get("required", "—"))
        parts.append(
            f"{label} {item.get('value', '—')}/{required}（{status}）"
        )
    eligible = "可正式推荐" if evidence.get("recommendation_eligible") else "仅观察"
    return f"{eligible}；{'；'.join(parts)}"


def format_recommendation_failure_reasons_text(evidence: dict[str, Any] | None) -> str:
    if not evidence:
        return "样本信息不足"
    failures = evidence.get("recommendation_failure_reasons") or []
    if not failures:
        return "全部门槛达标"
    parts: list[str] = []
    for item in failures:
        criterion = item.get("criterion", "")
        label = RECOMMENDATION_CRITERION_LABELS.get(criterion, criterion or "未知门槛")
        required = item.get("required", "—")
        parts.append(f"{label} 实际 {item.get('value', '—')} / 要求 {required}")
    return "；".join(parts)


def format_stage_inversion_diagnostics_text(comp: dict[str, Any]) -> str:
    diagnostics = comp.get("stage_inversion_diagnostics") or {}
    if not diagnostics:
        return "—"
    if not diagnostics.get("detected"):
        return "未检测到成熟/过渡倒挂"
    rejected = diagnostics.get("rejected_higher_tier_stages") or []
    if not rejected:
        return "检测到倒挂，但缺少被否决阶段明细"
    parts: list[str] = []
    for item in rejected[:3]:
        label = item.get("label") or item.get("bond") or "—"
        reasons = [
            INVERSION_REASON_LABELS.get(reason, reason)
            for reason in item.get("inversion_reasons", [])
        ]
        parts.append(f"{label}（{'、'.join(reasons) or '结构更高但表现更差'}）")
    return "；".join(parts)


def format_ceiling_conditions_text(ceiling_stage: dict[str, Any] | None) -> str:
    if not ceiling_stage:
        return "—"
    conditions = ceiling_stage.get("high_investment_conditions") or []
    if conditions:
        return "；".join(conditions)
    return "；".join(
        CEILING_CONDITION_LABELS.get(key, key)
        for key, met in (ceiling_stage.get("conditions") or {}).items()
        if met
    )


def format_primary_bond_audit_text(primary_bond: dict[str, Any]) -> str:
    source_labels = primary_bond.get("source_definition") or {}
    if source_labels:
        source_text = "；".join(
            f"{PRIMARY_BOND_SOURCE_LABELS.get(key, key)}={label}"
            for key, label in source_labels.items()
        )
    else:
        source_text = "4学习独占；收菜归美食社；普通羁绊第二档门；高费拼多多兜底"
    source_distribution = format_primary_bond_source_distribution(
        primary_bond.get("source_distribution")
    )
    category_distribution = format_primary_bond_source_distribution(
        primary_bond.get("category_distribution")
    )
    parts = [source_text]
    if source_distribution != "—":
        parts.append(f"来源审计 {source_distribution}")
    if category_distribution != "—":
        parts.append(f"归类审计 {category_distribution}")
    return "；".join(parts)


def format_mature_transition_lines(comp: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    mature_stats = comp.get("mature_stats") or comp.get("stats", {})
    if mature_stats.get("appearances"):
        lines.append(
            f"- 成熟表现：avg {mature_stats['avg_rank']:.2f}，"
            f"top4 {render_pct(mature_stats['top4_rate'])}，"
            f"n={mature_stats['appearances']}。"
        )
    transition_stats = comp.get("transition_stats")
    if transition_stats and transition_stats.get("appearances"):
        lines.append(
            f"- 过渡表现：avg {transition_stats['avg_rank']:.2f}，"
            f"top4 {render_pct(transition_stats['top4_rate'])}，"
            f"n={transition_stats['appearances']}。"
        )
    elif comp.get("transition_stages"):
        lines.append("- 过渡表现：存在过渡阶段样本，计入成型难度但不单独推荐。")
    return lines


def format_score_breakdown_text(breakdown: dict[str, Any] | None) -> str:
    if not breakdown:
        return "—"
    return (
        f"收缩均分 {breakdown.get('shrunk_avg_rank', '—')}；"
        f"前四下界 {breakdown.get('top4_lower_bound', '—')}；"
        f"吃鸡下界 {breakdown.get('win_lower_bound', '—')}；"
        f"不确定惩罚 {breakdown.get('uncertainty_penalty', '—')}；"
        f"成型惩罚 {breakdown.get('difficulty_penalty', '—')}；"
        f"趋势调整 {breakdown.get('trend_adjustment', '—')}"
    )


def format_cluster_merge_reason_lines(comp: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    cluster = comp.get("cluster_reason") or {}
    if cluster:
        dist = cluster.get("archetype_distribution") or []
        if dist:
            dist_text = "、".join(
                f"{item.get('archetype', '—')}({item.get('appearances', 0)})"
                for item in dist[:4]
            )
            lines.append(f"- 聚类证据：原型分布 {dist_text}。")
        if cluster.get("avg_pair_hero_jaccard") is not None:
            lines.append(
                f"- 聚类证据：平均英雄 Jaccard {cluster['avg_pair_hero_jaccard']}。"
            )
    merge = comp.get("merge_reason") or {}
    if merge:
        lines.append(
            f"- 合并证据：{merge.get('archetype', '—')} / "
            f"成熟 {merge.get('mature_member_count', 0)} / "
            f"过渡 {merge.get('transition_member_count', 0)}；"
            f"{merge.get('ownership_rule', 'one_player_one_top_level_strategy')}。"
        )
    return lines


def format_trend_detail_text(trend: dict[str, Any] | None) -> str:
    if not trend:
        return "—"
    label = format_trend_label(trend)
    if label == "样本不足":
        reasons = trend.get("reasons") or []
        return f"{label}（{'；'.join(reasons) or '窗口样本不足'}）"
    changes = trend.get("changes") or {}
    return (
        f"{label}（选秀 {changes.get('pick_rate', '—')}pp / "
        f"均分 {changes.get('shrunk_avg_rank', '—')} / "
        f"前四 {changes.get('shrunk_top4_rate', '—')}pp；"
        f"窗口 {trend.get('mode', '—')}）"
    )


def known_archetypes_from_data(data: dict[str, Any]) -> list[str]:
    archetypes: set[str] = set()
    for comp in data["rankings"].get("compositions", []):
        archetype = comp.get("archetype")
        if archetype and archetype != "未分类":
            archetypes.add(archetype)
    recommendations = data["rankings"].get("composition_recommendations", {})
    for style in PLAY_STYLES:
        for comp in recommendations.get(style, []):
            archetype = comp.get("archetype")
            if archetype and archetype != "未分类":
                archetypes.add(archetype)
    return sorted(archetypes)


def append_comp_markdown(
    lines: list[str],
    comp: dict[str, Any],
) -> None:
    stats = comp["stats"]
    evidence = comp.get("confidence_evidence") or {}
    lines.append(
        f"### {comp['label']}（{comp.get('play_style', '高费')}，"
        f"{comp.get('archetype', '未分类')}，"
        f"{comp['confidence']}置信，n={stats['appearances']}）"
    )
    lines.append("")
    lines.append(
        f"- 玩法原型：{comp.get('archetype', '未分类')}；归类证据：{format_archetype_signals_text(comp)}。"
    )
    lines.extend(format_cluster_merge_reason_lines(comp))
    lines.extend(format_mature_transition_lines(comp))
    trend = comp.get("trend")
    if trend:
        lines.append(f"- 版本趋势：{format_trend_detail_text(trend)}。")
    lines.append(
        f"- 样本：raw {evidence.get('raw_n', stats.get('appearances', '—'))} / "
        f"weighted {evidence.get('weighted_n', stats.get('weighted_appearances', '—'))} / "
        f"n_eff {evidence.get('n_eff', stats.get('n_eff', '—'))}。"
    )
    lines.append(f"- 置信解释：{format_confidence_criteria_text(evidence)}。")
    if not evidence.get("recommendation_eligible", True):
        lines.append(
            f"- 低置信提示：{format_recommendation_failure_reasons_text(evidence)}。"
        )
    inversion_text = format_stage_inversion_diagnostics_text(comp)
    if inversion_text != "—":
        lines.append(f"- 成熟/过渡倒挂：{inversion_text}。")
    if comp.get("score_breakdown"):
        lines.append(f"- 评分分解：{format_score_breakdown_text(comp['score_breakdown'])}。")
    lines.append(
        f"- 表现：avg {stats['avg_rank']:.2f}，top4 {render_pct(stats['top4_rate'])}，吃鸡 {render_pct(stats['win_rate'])}。"
    )
    difficulty = comp.get("difficulty", {})
    if difficulty:
        lines.append(
            f"- 三星压力：阵容平均{difficulty.get('avg_three_star_units', 0):.2f}个三星棋子，"
            f"前四样本平均{difficulty.get('avg_top4_three_star_units', 0):.2f}个；"
            f"同行数{difficulty.get('avg_same_match_contest', 0):.2f}"
            f"（{difficulty.get('contest_basis', '阵容相似')}）。"
        )
    carries = "、".join(
        f"P{item.get('carry_rank', idx)} {item['hero_name']}({render_pct(item['share'])})"
        for idx, item in enumerate(comp["main_carries"], start=1)
    )
    lines.append(f"- 主C判断：{carries or '样本不足'}。")
    breakdown = "、".join(
        f"{row['play_style']}{render_pct(row['share'])}"
        for row in comp.get("play_style_breakdown", [])
    )
    if breakdown:
        lines.append(f"- 类型样本：{breakdown}。")
    if comp.get("carry_requirements"):
        req_text = "；".join(
            f"{row['hero_name']}建议至少{row['recommended_min_stars']}星"
            f"（前四平均{row['avg_stars_top4']:.1f}星，三件套{render_pct(row['three_item_rate'])}）"
            for row in comp["carry_requirements"][:3]
        )
        lines.append(f"- 主C成型门槛：{req_text}。")
        expensive_note = [
            row["hero_name"]
            for row in comp["carry_requirements"][:3]
            if row.get("high_cost_three_star_dependency")
        ]
        if expensive_note:
            lines.append(
                f"- 成型成本提醒：{'、'.join(expensive_note)} 的三星高费样本会拉高上限，常规推荐按 2 星门槛评估。"
            )
    if comp.get("carry_equipment_notes"):
        note_parts = []
        for note in comp["carry_equipment_notes"][:3]:
            important = [
                item
                for item in note["items"]
                if item["label"] in ("疑似刚需", "高价值")
            ][:3]
            if important:
                note_parts.append(
                    f"{note['hero_name']}："
                    + "、".join(
                        f"{item['equipment_name']}({item['label']}, 不带惩罚{item['without_item_penalty']})"
                        for item in important
                    )
                )
        if note_parts:
            lines.append(f"- 主C关键装备：{'；'.join(note_parts)}。")
    if comp.get("jiujiu_requirements"):
        jiujiu_parts = []
        for req in comp["jiujiu_requirements"]:
            wearers = "、".join(
                f"{item['hero_name']}({render_pct(item['share'])})"
                for item in req.get("recommended_wearers", [])[:3]
            )
            jiujiu_parts.append(
                f"{req['recommended_jiujiu']}（{render_pct(req['dependency_rate'])}样本需啾啾开"
                f"{req['trait']}-{req['target_tier']}，推荐穿戴：{wearers or '待观察'}）"
            )
        if jiujiu_parts:
            lines.append(f"- 啾啾成型：{'；'.join(jiujiu_parts)}。")
    bonds = "、".join(
        f"{item['bond']}({render_pct(item['share'])})" for item in comp["common_bonds"][:5]
    )
    lines.append(f"- 常见羁绊：{bonds or '无稳定羁绊'}。")
    lines.append("")
    lines.append("| 等级 | 来源 | 置信度 | 羁绊达成 | 棋子 |")
    lines.append("| ---: | --- | --- | --- | --- |")
    for level in ("7", "8", "9"):
        variant = comp["variants"][level]
        lines.append(
            f"| {level} | {variant['source']} | {variant['confidence']} | "
            f"{variant.get('bond_note', '—')} | "
            f"{'、'.join(variant['heroes'])} |"
        )
    lines.append("")


def append_ceiling_markdown(
    lines: list[str],
    sample: dict[str, Any],
    *,
    zone: str = "ceiling",
) -> None:
    stats = sample["stats"]
    evidence = sample.get("confidence_evidence") or {}
    ceiling_stage = sample.get("ceiling_stage") or {}
    zone_label = "观察" if zone == "ceiling-observation" else "正式"
    lines.append(
        f"### {sample['label']}（{zone_label}，"
        f"{sample.get('recommendation_status', '高费大成上限')}，"
        f"n={stats['appearances']}）"
    )
    lines.append("")
    lines.append(
        f"- 说明：{ceiling_stage.get('interpretation', '仅根据最终盘完成度归纳形态，不代表观测到真实过渡过程')}。"
    )
    lines.append(
        f"- 完成度条件：{format_ceiling_conditions_text(ceiling_stage)}。"
    )
    lines.append(
        f"- 表现：top2 {render_pct(stats.get('top2_rate', 0))}，"
        f"吃鸡 {render_pct(stats['win_rate'])}，"
        f"avg {stats['avg_rank']:.2f}。"
    )
    lines.append(
        f"- 样本：raw {evidence.get('raw_n', stats.get('appearances', '—'))} / "
        f"weighted {evidence.get('weighted_n', stats.get('weighted_appearances', '—'))} / "
        f"n_eff {evidence.get('n_eff', stats.get('n_eff', '—'))}。"
    )
    lines.append(f"- 置信证据：{format_confidence_criteria_text(evidence)}。")
    if zone == "ceiling-observation" or not evidence.get("recommendation_eligible"):
        lines.append(
            f"- 推荐资格失败：{format_recommendation_failure_reasons_text(evidence)}。"
        )
    if ceiling_stage.get("two_star_high_cost_heroes"):
        lines.append(
            f"- 关键高费两星：{'、'.join(ceiling_stage['two_star_high_cost_heroes'])}。"
        )
    if ceiling_stage.get("high_cost_three_star_nonstandard"):
        heroes = ceiling_stage.get("high_cost_three_star_heroes") or []
        lines.append(
            f"- 三星高费提醒：{'、'.join(heroes) or '存在'} 属于上限样本，不代表常规过渡要求。"
        )
    boards = (
        sample.get("representative_final_boards")
        or ceiling_stage.get("representative_final_boards")
        or []
    )
    if boards:
        lines.append("")
        lines.append("观测到的代表性最终阵容（按完整棋子组合聚合，仅最终盘样本）：")
        lines.append("")
        lines.append("| 排名 | 样本 | 占比 | 平均名次 | 主C | 棋子 |")
        lines.append("| ---: | ---: | ---: | ---: | --- | --- |")
        for index, board in enumerate(boards, start=1):
            heroes = "、".join(board.get("heroes") or []) or "—"
            lines.append(
                f"| {index} | {board.get('appearances', 0)} | "
                f"{render_pct(board.get('share', 0))} | "
                f"{board.get('avg_rank', 0):.2f} | "
                f"{board.get('main_carry') or '—'} | {heroes} |"
            )
    lines.append("")


def html_ceiling_final_board_cards(sample: dict[str, Any]) -> str:
    ceiling_stage = sample.get("ceiling_stage") or {}
    boards = (
        sample.get("representative_final_boards")
        or ceiling_stage.get("representative_final_boards")
        or []
    )
    if not boards:
        return '<p class="muted">当前上限样本尚未归纳出可展示的完整最终阵容。</p>'
    cards: list[str] = []
    for index, board in enumerate(boards, start=1):
        heroes = board.get("heroes") or []
        hero_chips = "".join(f'<span class="hero-chip">{esc(hero)}</span>' for hero in heroes)
        cards.append(
            f"""
            <article class="board-card">
              <div class="board-head">
                <span class="level-badge">Lv{esc(str(board.get('level', '—')))}</span>
                <span class="source-badge">观测最终盘</span>
                <span class="conf-badge">{esc(board.get('confidence', '—'))}</span>
                <span class="sample-badge">样本 {esc(str(board.get('appearances', 0)))} · {render_pct(board.get('share', 0))}</span>
              </div>
              <p class="bond-note">主C {esc(board.get('main_carry') or '—')} · avg {board.get('avg_rank', 0):.2f} · 前四 {render_pct(board.get('top4_rate', 0))}</p>
              <div class="hero-chips">{hero_chips or '<span class="muted">样本不足</span>'}</div>
            </article>
            """
        )
    return f'<div class="board-grid">{"".join(cards)}</div>'


def render_md(data: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append("# 蛋仔派对当前环境分析报告")
    lines.append("")
    lines.append(f"- 生成时间: `{data['generated_at']}`")
    lines.append(f"- 数据源: `{data['data_source']}`")
    lines.append(f"- 分析器: `{data['methodology']['implementation']}`")
    lines.append("")

    quality = data["overview"]["quality"]
    lines.append("## 数据概览与过滤摘要")
    lines.append("")
    lines.append("| 指标 | 数值 |")
    lines.append("| --- | ---: |")
    for key in (
        "matches",
        "players",
        "heroes",
        "hero_equipments",
        "cards",
        "unknown_heroes",
        "unknown_cards",
        "unknown_equipment",
        "card_granted_heroes",
        "seven_eight_bot_matches",
        "bot_player_records_excluded",
    ):
        lines.append(f"| {key} | {quality[key]} |")
    lines.append(f"| filtered_players | {data['overview']['filtered_players']} |")
    recency = data["overview"].get("recency", {})
    if recency:
        lines.append(f"| effective_sample_weight | {data['overview'].get('effective_sample_weight', '—')} |")
        lines.append(f"| latest_batch | {recency.get('latest_batch', '—')} |")
        batch_range = recency.get("batch_range") or []
        if batch_range:
            lines.append(f"| batch_range | {' → '.join(batch_range)} |")
    lines.append("")
    if recency.get("batch_counts"):
        lines.append("### 批次样本与近期加权")
        lines.append("")
        lines.append(
            f"- 批次来源：`{recency.get('source', 'screenshots.MMDD')}`；"
            f"半衰期 {recency.get('half_life_days', DEFAULT_RECENCY_HALF_LIFE_DAYS)} 天，"
            f"最低权重 {recency.get('min_weight', MIN_RECENCY_WEIGHT)}。"
        )
        lines.append(
            "- 均分/胜率/前四率按加权样本计算；样本阈值与置信度仍看原始 n。"
        )
        lines.append("")
        lines.append("| 批次 | 原始样本 | 加权样本 |")
        lines.append("| --- | ---: | ---: |")
        for batch, count in recency["batch_counts"].items():
            weighted = recency.get("batch_weighted_counts", {}).get(batch, count)
            lines.append(f"| {batch} | {count} | {weighted} |")
        lines.append("")

    comps = data["rankings"]["compositions"]
    heroes = data["rankings"]["heroes_and_equipment"]["heroes"]
    cards = data["rankings"]["cards"]["single_cards"]
    lines.append("## 当前环境结论摘要")
    lines.append("")
    if comps:
        top_comp = comps[0]
        lines.append(
            f"- 当前最优阵容族群：**{top_comp['label']}**，avg {top_comp['stats']['avg_rank']:.2f}，"
            f"top4 {render_pct(top_comp['stats']['top4_rate'])}，n={top_comp['stats']['appearances']}。"
        )
        recommendations = data["rankings"].get("composition_recommendations", {})
        for style in PLAY_STYLES:
            rows = recommendations.get(style, [])
            if not rows:
                continue
            top_style_comp = rows[0]
            lines.append(
                f"- {style}推荐首选：**{top_style_comp['label']}**，avg {top_style_comp['stats']['avg_rank']:.2f}，"
                f"top4 {render_pct(top_style_comp['stats']['top4_rate'])}，n={top_style_comp['stats']['appearances']}。"
            )
    if heroes:
        top_heroes = "、".join(
            f"{row['hero_name']}（carry {render_pct(row['carry_rate'])}, avg {row['avg_rank']:.2f}）"
            for row in heroes[:5]
        )
        lines.append(f"- 高投入核心棋子：{top_heroes}。")
    if cards:
        cards_by_prefix = data["rankings"]["cards"].get("single_cards_by_prefix", {})
        if cards_by_prefix:
            top_cards = "、".join(
                f"{prefix_type}类 {rows[0]['key']}（修正 {rows[0]['adjusted_avg_rank']:.2f}）"
                for prefix_type in CARD_PREFIX_TYPES
                if (rows := cards_by_prefix.get(prefix_type)) and rows[0]["appearances"] > 0
            )
        else:
            top_cards = "、".join(
                f"{row['key']}（修正 {row['adjusted_avg_rank']:.2f}）"
                for row in cards[:5]
            )
        lines.append(f"- 强势卡牌（分类型）：{top_cards}。")
    lines.append("")

    lines.append("## 赌狗阵容推荐")
    lines.append("")
    recommendations = data["rankings"].get("composition_recommendations", {})
    reroll_comps = recommendations.get("赌狗", [])
    if not reroll_comps:
        lines.append("当前样本下没有达到发现门槛的赌狗阵容。")
        lines.append("")
    for comp in reroll_comps:
        append_comp_markdown(lines, comp)

    lines.append("## 高费阵容推荐")
    lines.append("")
    high_cost_comps = recommendations.get("高费", [])
    if not high_cost_comps:
        lines.append("当前样本下没有达到发现门槛的高费阵容。")
        lines.append("")
    for comp in high_cost_comps:
        append_comp_markdown(lines, comp)

    lines.append("## 阵容成型难度与热门程度")
    lines.append("")
    lines.append(
        "强度排名综合成型后表现（平均名次、前四率、吃鸡率）与成型难度（未成型后四率、同行压力、装备完整率）。"
    )
    lines.append("")
    strength_sorted = sorted(
        comps,
        key=lambda row: (
            row.get("strength_rank", 999),
            row.get("overall_strength_score", 99),
        ),
    )
    lines.append("| 强度排名 | 阵容 | 类型 | 难度 | 热门 | 平均三星 | 后四未成型率 | 同行数 | 同行口径 | 出场率 |")
    lines.append("| ---: | --- | --- | --- | --- | ---: | ---: | ---: | --- | ---: |")
    for comp in strength_sorted[:12]:
        difficulty = comp["difficulty"]
        popularity = comp["popularity"]
        lines.append(
            f"| {comp.get('strength_rank', '—')} | {comp['label']} | {comp.get('play_style', '高费')} | "
            f"{difficulty['label']} | {popularity['label']} | "
            f"{difficulty.get('avg_three_star_units', 0):.2f} | "
            f"{render_pct(difficulty['unfinished_bottom_rate'])} | "
            f"{difficulty['avg_same_match_contest']:.2f} | "
            f"{difficulty.get('contest_basis', '阵容相似')} | {render_pct(popularity['pick_rate'])} |"
        )
    lines.append("")

    low_cost_difficulty = data["rankings"].get("low_cost_carry_three_star_difficulty", [])
    if low_cost_difficulty:
        lines.append("### 低费主C三星难度")
        lines.append("")
        lines.append("同场多家阵容需要同一个低费3星主C时，即使阵容路线不同，也计入同行压力。")
        lines.append("")
        lines.append("| 棋子 | 费用 | 平均同场需求 | 最高同场需求 | 多家需求对局率 | 平均名次 | 样本 | 主要阵容 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: | --- |")
        for row in [item for item in low_cost_difficulty if item.get("is_low_cost")][:12]:
            strategies = "；".join(
                f"{item['label']}({item['samples']})" for item in row.get("top_strategies", [])[:3]
            )
            lines.append(
                f"| {row['hero_name']} | {row.get('tier') or '—'} | "
                f"{row['avg_same_match_needers']:.2f} | {row['max_same_match_needers']} | "
                f"{render_pct(row['multi_needer_match_rate'])} | {row['avg_rank']:.2f} | "
                f"{row['appearances']} | {strategies or '样本不足'} |"
            )
        lines.append("")
        lines.append("### 主C三星需求热门程度")
        lines.append("")
        lines.append("| 棋子 | 费用 | 平均同场需求 | 最高同场需求 | 多家需求对局率 | 需要它三星的阵容 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | --- |")
        for row in low_cost_difficulty[:12]:
            strategies = "；".join(
                f"{item['label']}({item['samples']})" for item in row.get("top_strategies", [])[:3]
            )
            lines.append(
                f"| {row['hero_name']} | {row.get('tier') or '—'} | "
                f"{row['avg_same_match_needers']:.2f} | {row['max_same_match_needers']} | "
                f"{render_pct(row['multi_needer_match_rate'])} | {strategies or '样本不足'} |"
            )
        lines.append("")

    lines.append("## 卡牌强度分析")
    lines.append("")
    lines.append(
        "卡牌顺序按 `slot_index` 统计，第一张卡牌视为双人配合重点；队伍排名按每局队伍最高个人名次重新排序为 1-4。"
    )
    lines.append("单卡与第一卡强度按模板前缀类型（彩/黄/蓝/白/其他）分组，并在各组内优先按样本数排序。")
    lines.append("")
    cards_by_prefix = data["rankings"]["cards"].get("single_cards_by_prefix", {})
    if cards_by_prefix:
        for prefix_type in CARD_PREFIX_TYPES:
            prefix_rows = cards_by_prefix.get(prefix_type, [])
            if not prefix_rows:
                continue
            lines.append(f"### {prefix_type}类单卡")
            lines.append("")
            lines.append("| 组内排名 | 卡牌 | 样本 | 每局平均 | 修正名次 | 平均名次 | 前四率 | 吃鸡率 |")
            lines.append("| ---: | --- | ---: | ---: | ---: | ---: | ---: | ---: |")
            for row in prefix_rows:
                rank_label = "—" if row.get("prefix_rank") is None else str(row["prefix_rank"])
                lines.append(
                    f"| {rank_label} | {row['key']} | {row['appearances']} | "
                    f"{row.get('avg_appearances_per_match', 0):.2f} | "
                    f"{render_card_metric(row.get('adjusted_avg_rank'))} | "
                    f"{render_card_metric(row.get('avg_rank'))} | "
                    f"{render_pct(row.get('top4_rate'))} | "
                    f"{render_pct(row.get('win_rate'))} |"
                )
            lines.append("")
    else:
        lines.append("| 卡牌 | 样本 | 每局平均 | 修正名次 | 平均名次 | 前四率 | 吃鸡率 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
        for row in cards[:20]:
            lines.append(
                f"| {row['key']} | {row['appearances']} | {row.get('avg_appearances_per_match', 0):.2f} | "
                f"{row['adjusted_avg_rank']:.2f} | {row['avg_rank']:.2f} | "
                f"{render_pct(row['top4_rate'])} | {render_pct(row['win_rate'])} |"
            )
        lines.append("")
    first_cards = data["rankings"]["cards"]["first_card_rankings"]
    first_cards_by_prefix = data["rankings"]["cards"].get("first_card_rankings_by_prefix", {})
    if first_cards_by_prefix:
        lines.append("### 第一张卡牌强度（分类型）")
        lines.append("")
        for prefix_type in CARD_PREFIX_TYPES:
            prefix_rows = first_cards_by_prefix.get(prefix_type, [])
            if not prefix_rows:
                continue
            lines.append(f"#### {prefix_type}类第一卡")
            lines.append("")
            lines.append("| 组内排名 | 第一卡 | 样本 | 每局平均 | 修正名次 | 平均名次 | 前四率 |")
            lines.append("| ---: | --- | ---: | ---: | ---: | ---: | ---: |")
            for row in prefix_rows[:8]:
                lines.append(
                    f"| {row['prefix_rank']} | {row['key']} | {row['appearances']} | "
                    f"{row.get('avg_appearances_per_match', 0):.2f} | {row['adjusted_avg_rank']:.2f} | "
                    f"{row['avg_rank']:.2f} | {render_pct(row['top4_rate'])} |"
                )
            lines.append("")
    elif first_cards:
        lines.append("### 第一张卡牌强度")
        lines.append("")
        lines.append("| 第一卡 | 样本 | 每局平均 | 修正名次 | 平均名次 | 前四率 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: |")
        for row in first_cards[:12]:
            lines.append(
                f"| {row['key']} | {row['appearances']} | {row.get('avg_appearances_per_match', 0):.2f} | "
                f"{row['adjusted_avg_rank']:.2f} | {row['avg_rank']:.2f} | {render_pct(row['top4_rate'])} |"
            )
        lines.append("")
    blue_team_cards = data["rankings"]["cards"].get("blue_cards_team_rank", [])
    if blue_team_cards:
        lines.append("### 蓝卡队伍排名视角")
        lines.append("")
        lines.append("蓝卡按双人卡牌处理，额外使用队伍排名评估；队伍排名按每局队伍最高个人名次重新排序。")
        lines.append("")
        lines.append("| 蓝卡 | 样本 | 每局平均 | 修正队伍名次 | 队伍名次 | 队伍前二率 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: |")
        for row in blue_team_cards[:12]:
            lines.append(
                f"| {row['key']} | {row['appearances']} | {row.get('avg_appearances_per_match', 0):.2f} | "
                f"{row['adjusted_avg_rank']:.2f} | {row['avg_rank']:.2f} | {render_pct(row.get('team_top2_rate', 0))} |"
            )
        lines.append("")
    first_duos = data["rankings"]["cards"]["first_card_duo_synergy"]
    if first_duos:
        lines.append("### 双人第一卡配合")
        lines.append("")
        lines.append("| 第一卡组合 | 修正队伍名次 | 队伍名次 | 样本 |")
        lines.append("| --- | ---: | ---: | ---: |")
        for row in first_duos[:12]:
            lines.append(
                f"| {row['key']} | {row['adjusted_avg_rank']:.2f} | {row['avg_rank']:.2f} | {row['appearances']} |"
            )
        lines.append("")
    contribution = data["rankings"]["cards"]["duo_card_contribution"]
    if contribution:
        lines.append("### 第一卡贡献增量观察")
        lines.append("")
        lines.append("| 第一卡组合 | 队伍名次 | 相对基线提升 | 持有者折算提升 | 队伍前二率 | 样本 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: |")
        for row in contribution[:12]:
            lines.append(
                f"| {row['key']} | {row['team_avg_rank']:.2f} | {row['team_lift_vs_baseline']:.2f} | "
                f"{row['team_lift_vs_holder']:.2f} | {render_pct(row['team_top2_rate'])} | {row['appearances']} |"
            )
        lines.append("")
    comp_cards = data["rankings"]["cards"]["composition_cards"]
    if comp_cards:
        lines.append("### 阵容内卡牌观察")
        lines.append("")
        for row in comp_cards[:5]:
            picks = "、".join(
                f"{card['key']}({card['adjusted_avg_rank']:.2f}, n={card['appearances']}, 每局{card.get('avg_appearances_per_match', 0):.2f})"
                for card in row["cards"][:5]
            )
            lines.append(f"- {row['family_label']}：{picks}")
        lines.append("")
    teammate_cards = data["rankings"]["cards"]["teammate_card_pairs_observation"]
    if teammate_cards:
        lines.append("### 队友卡牌配合观察")
        lines.append("")
        lines.append("以下组合为低置信观察，需要结合样本量判断：")
        lines.append("")
        for row in teammate_cards[:10]:
            lines.append(
                f"- {row['key']}：修正 {row['adjusted_avg_rank']:.2f}，top4 {render_pct(row['top4_rate'])}，n={row['appearances']}"
            )
        lines.append("")

    duo_comps = data["rankings"].get("duo_composition_synergy", [])
    if duo_comps:
        lines.append("## 双人阵容配合推荐")
        lines.append("")
        lines.append("基于同队两家的最终策略组合与重算队伍排名，仅作双排分工参考。")
        lines.append("")
        lines.append("| 阵容组合 | 队伍名次 | 相对基线提升 | 队伍前二率 | 队伍吃鸡率 | 样本 | 置信度 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | ---: | --- |")
        for row in duo_comps[:12]:
            lines.append(
                f"| {row['strategy_a']} + {row['strategy_b']} | {row['team_avg_rank']:.2f} | "
                f"{row['team_lift_vs_baseline']:.2f} | {render_pct(row['team_top2_rate'])} | "
                f"{render_pct(row['team_win_rate'])} | {row['appearances']} | {row['confidence']} |"
            )
        lines.append("")

    lines.append("## 强势棋子与装备推荐")
    lines.append("")
    recommendations = data["rankings"]["heroes_and_equipment"]["carry_equipment_recommendations"]
    equipment_xlsx = data.get("outputs", {}).get("equipment_xlsx", "data/latest_meta_analysis_equipment.xlsx")
    interactive_html = data.get("outputs", {}).get(
        "interactive_html", "data/环境分析详情.html"
    )
    hero_equipment_dir = data.get("outputs", {}).get(
        "hero_equipment_dir", "data/hero-equipment"
    )
    with_items = sum(1 for row in recommendations if row.get("has_equipment_data"))
    lines.append(
        f"每位英雄的详细出装推荐已导出至 **`{equipment_xlsx}`**（Excel）与 **`{interactive_html}#equipment`**（可筛选 HTML），"
        f"覆盖过滤后样本中出现的全部 **{len(recommendations)}** 个棋子；"
        f"其中 **{with_items}** 个有可靠或低样本出装记录。"
        f"棋子详情为独立页面（目录 **`{hero_equipment_dir}/`**），仅展示单装 raw 样本 >10 的装备，"
        f"从装备概览点击棋子名会在新标签页打开。"
        f"超级装备见 **`{interactive_html}#super-equipment`**，美食社装备见 **`{interactive_html}#food-equipment`**。"
    )
    lines.append("")
    lines.append("排序：费用从低到高，同费按主C投入与名称。")
    lines.append("")
    top_carries = sorted(
        recommendations,
        key=lambda row: (
            -row["hero_stats"].get("carry_appearances", 0),
            row["hero_stats"].get("adjusted_avg_rank", 99),
        ),
    )[:8]
    if top_carries:
        lines.append("### 高投入主C速览")
        lines.append("")
        for row in top_carries:
            hero = row["hero_stats"]
            top_items = row.get("recommended_items") or row.get("low_sample_observations") or []
            item_hint = "、".join(item["equipment_name"] for item in top_items[:3]) or "出装样本不足"
            super_hint = "、".join(
                item["equipment_name"] for item in (row.get("recommended_super_items") or [])[:2]
            )
            food_hint = "、".join(
                item["equipment_name"] for item in (row.get("recommended_food_items") or [])[:2]
            )
            extra_parts = []
            if super_hint:
                extra_parts.append(f"超级 {super_hint}")
            if food_hint:
                extra_parts.append(f"美食社 {food_hint}")
            extra = f"；{'；'.join(extra_parts)}" if extra_parts else ""
            detail_path = hero_equipment_detail_relpath(
                row["hero_name"],
                slug=row.get("detail_slug"),
                directory=hero_equipment_dir,
            )
            lines.append(
                f"- **[{row['hero_name']}]({detail_path})**：主C率 {render_pct(hero['carry_rate'])}，"
                f"avg {hero['avg_rank']:.2f}，优先 {item_hint}{extra}"
            )
        lines.append("")

    def append_special_equipment_md(title: str, section_key: str, anchor: str) -> None:
        special = data["rankings"].get(section_key, {})
        rankings = special.get("rankings", [])
        lines.append(f"### {title}")
        lines.append("")
        lines.append(special.get("definition", ""))
        lines.append("")
        lines.append(f"完整可排序表格见 **`{interactive_html}#{anchor}`**。")
        lines.append("")
        if not rankings:
            lines.append("- 当前样本不足以形成该分类装备排名。")
            lines.append("")
            return
        lines.append("| 强度排名 | 装备 | 样本 | 修正名次 | 前四率 | 吃鸡率 | 置信度 | 推荐佩戴 | 备注 |")
        lines.append("| ---: | --- | ---: | ---: | ---: | ---: | --- | --- | --- |")
        for row in rankings[:20]:
            wearers = "、".join(
                f"{item['hero_name']}(n={item['appearances']})"
                for item in row.get("recommended_wearers", [])[:3]
            ) or "待观察"
            avg_rank = (
                f"{row['adjusted_avg_rank']:.2f}"
                if row.get("adjusted_avg_rank") is not None
                else "—"
            )
            top4 = render_pct(row["top4_rate"]) if row.get("top4_rate") is not None else "—"
            win = render_pct(row["win_rate"]) if row.get("win_rate") is not None else "—"
            lines.append(
                f"| {row.get('strength_rank', '—')} | {row['equipment_name']} | "
                f"{row.get('appearances', 0)} | {avg_rank} | {top4} | {win} | "
                f"{row.get('confidence', '低')} | {wearers} | {row.get('note') or '—'} |"
            )
        lines.append("")

    append_special_equipment_md("超级装备强度", "super_equipment", "super-equipment")
    append_special_equipment_md("美食社装备强度", "food_equipment", "food-equipment")

    primary_bond_html = data.get("outputs", {}).get(
        "interactive_html",
        "data/环境分析详情.html",
    )
    primary_bond = data["rankings"].get("primary_bond_strength", {})
    primary_rows = primary_bond.get("rows", [])
    lines.append("## 主羁绊强度")
    lines.append("")
    lines.append(primary_bond.get("definition", "按最终阵容激活数量最大的羁绊统计。"))
    lines.append("")
    lines.append(f"- 归类规则：{format_primary_bond_audit_text(primary_bond)}。")
    lines.append("")
    if primary_rows:
        lines.append(
            f"完整可排序表格见 **`{primary_bond_html}#primary-bond`**。"
        )
        lines.append("")
        lines.append("| 强度排名 | 主羁绊 | 样本 | 平均名次 | 前四率 | 后四率 | 吃鸡率 | 常见激活数量 | 归类 | 归类来源 |")
        lines.append("| ---: | --- | ---: | ---: | ---: | ---: | ---: | --- | --- | --- |")
        for row in primary_rows[:20]:
            lines.append(
                f"| {row['strength_rank']} | {row['bond']} | {row['appearances']} | "
                f"{row['avg_rank']:.2f} | {render_pct(row['top4_rate'])} | "
                f"{render_pct(row['bottom4_rate'])} | {render_pct(row['win_rate'])} | "
                f"{row.get('common_activation_summary', '—')} | "
                f"{row.get('category', row['bond'])} | "
                f"{format_primary_bond_source_distribution(row.get('source_distribution'))} |"
            )
    else:
        lines.append("- 当前样本不足以形成主羁绊强度排名。")
    lines.append("")

    lines.append("## 羁绊表现与啾啾影响")
    lines.append("")
    bonds = data["rankings"]["heroes_and_equipment"]["bonds"]
    lines.append("| 羁绊档位 | 修正名次 | 平均名次 | 前四率 | 样本 |")
    lines.append("| --- | ---: | ---: | ---: | ---: |")
    for row in bonds[:20]:
        lines.append(
            f"| {row['key']} | {row['adjusted_avg_rank']:.2f} | {row['avg_rank']:.2f} | "
            f"{render_pct(row['top4_rate'])} | {row['appearances']} |"
        )
    jiujiu = data["rankings"]["heroes_and_equipment"]["jiujiu_bonds"]
    if jiujiu:
        lines.append("")
        lines.append("### 啾啾辅助羁绊")
        lines.append("")
        for row in jiujiu[:12]:
            lines.append(
                f"- {row['key']}：修正 {row['adjusted_avg_rank']:.2f}，top4 {render_pct(row['top4_rate'])}，n={row['appearances']}"
            )
    jiujiu_analysis = data["rankings"].get("jiujiu", {})
    rankings = jiujiu_analysis.get("jiujiu_rankings", [])
    if rankings:
        lines.append("")
        lines.append("### 啾啾强度排名")
        lines.append("")
        lines.append("| 啾啾 | 有效样本 | 有效率 | 有效修正 | 前四率 | 推荐阵容/穿戴棋子 |")
        lines.append("| --- | ---: | ---: | ---: | ---: | --- |")
        for row in rankings[:16]:
            comps = "；".join(
                f"{comp['family_label']}→"
                f"{'、'.join(wearer['hero_name'] for wearer in comp.get('recommended_wearers', [])[:2]) or '待观察'}"
                f"({comp['appearances']})"
                for comp in row["recommended_comps"][:2]
            )
            heroes = "；".join(
                f"{hero['hero_name']}({hero['appearances']})"
                for hero in row.get("recommended_heroes", [])[:2]
            )
            effective_stats = row.get("effective_stats") or row
            targets = comps or heroes or "有效样本不足"
            lines.append(
                f"| {row['equipment_name']} | {row['effective_appearances']} | {render_pct(row['effective_rate'])} | "
                f"{effective_stats['adjusted_avg_rank']:.2f} | {render_pct(effective_stats['top4_rate'])} | {targets} |"
            )
        lines.append("")
    lines.append("")

    lines.append("## 版本陷阱分析")
    lines.append("")
    traps = data["rankings"]["traps"]
    for label, key in (
        ("阵容", "compositions"),
        ("棋子", "heroes"),
        ("卡牌", "cards"),
        ("羁绊", "bonds"),
        ("装备", "equipment"),
    ):
        rows = traps[key]
        if not rows:
            continue
        lines.append(f"### {label}")
        lines.append("")
        for row in rows[:8]:
            name = row.get("label") or row.get("hero_name") or row.get("key") or row.get("equipment_name")
            stats = row.get("stats", row)
            prefix_note = ""
            if key == "cards" and row.get("prefix_type"):
                prefix_note = f"[{row['prefix_type']}类内] "
            trap_reason = row.get("trap_reason")
            reason_note = f"，{trap_reason}" if trap_reason else ""
            lines.append(
                f"- {prefix_note}{name}：avg {stats['avg_rank']:.2f}，top4 {render_pct(stats['top4_rate'])}，"
                f"n={stats['appearances']}{reason_note}"
            )
        lines.append("")
    pressure_bonds = traps.get("formation_pressure_bonds", [])
    if pressure_bonds:
        lines.append("### 未成型压力（并入成熟阵容）")
        lines.append("")
        for row in pressure_bonds[:8]:
            lines.append(
                f"- {row['key']}：avg {row['avg_rank']:.2f}，top4 {render_pct(row['top4_rate'])}，"
                f"n={row['appearances']}，{row.get('trap_reason', '计入成型难度')}"
            )
        lines.append("")

    balance = data["rankings"]["balance_targets"]
    lines.append("## 平衡性调整追踪")
    lines.append("")
    if any(balance.values()):
        for key, values in balance.items():
            lines.append(f"- {key}: {', '.join(values) if values else '无'}")
    else:
        lines.append("本次未提供平衡性调整文本，跳过定向追踪。")
    lines.append("")

    lines.append("## 数据质量与可信度说明")
    lines.append("")
    validation = data["overview"]["validation"]
    lines.append(f"- 配置未映射棋子：{', '.join(validation['missing_config_heroes']) or '无'}。")
    lines.append(f"- 啾啾未映射装备：{', '.join(validation['jiujiu_unmapped']) or '无'}。")
    lines.append(
        f"- 正式推荐门槛：原始 n>={RECOMMENDATION_MIN_RAW_N}、"
        f"加权 n>={RECOMMENDATION_MIN_WEIGHTED_N}、"
        f"有效 n_eff>={RECOMMENDATION_MIN_EFFECTIVE_N}、"
        f"批次>={RECOMMENDATION_MIN_BATCHES}、"
        f"聚类纯度>={RECOMMENDATION_MIN_CLUSTER_PURITY}；"
        "发现门槛仍为 min_comp_apps=5。"
    )
    lines.append(
        "- 蓝卡 `一起刷刷刷` 与 `天降啾啾pro` 共用图标，但按最终阵容啾啾装备数（>=2 为 pro）分别统计，"
        "不再合并为同一排行项。"
    )
    lines.append(
        "- 黄卡 `巨神兵` 与 `迅迅迅捷双剑` 共用图标，按最终阵容 `巨神兵之斧`/`迅捷双剑` 数量分别统计："
        "仅斧 -> 巨神兵，仅剑 -> 迅迅迅捷双剑，都有则数量占优；"
        "数量相同则按本次数据库明确样本比例并以固定种子可复现分配。"
    )
    lines.append("- 低样本阵容、卡牌组合和队友配合只作为观察，不应单独作为上分结论。")
    return "\n".join(lines) + "\n"


def esc(value: Any) -> str:
    return html.escape(str(value), quote=True)


def hero_equipment_detail_slug(hero_name: str) -> str:
    """Stable URL-safe slug for hero equipment page links."""
    return quote(str(hero_name), safe="")


def hero_equipment_detail_filename(hero_name: str) -> str:
    """Unicode filename for a standalone hero equipment page."""
    return f"{hero_name}.html"


def hero_equipment_detail_relpath(
    hero_name: str,
    *,
    slug: str | None = None,
    directory: str = "data/hero-equipment",
) -> str:
    """Repository-relative path used in Markdown and JSON outputs."""
    del slug  # Slug is URL-only; file names keep the readable hero name.
    return f"{directory.rstrip('/')}/{hero_equipment_detail_filename(hero_name)}"


def hero_equipment_detail_href_from_dashboard(hero_name: str, *, slug: str | None = None) -> str:
    """Relative href from ``环境分析详情.html`` to a standalone hero page."""
    encoded = slug or hero_equipment_detail_slug(hero_name)
    return f"hero-equipment/{encoded}.html"


def hero_equipment_detail_hash(hero_name: str, *, slug: str | None = None) -> str:
    """Deprecated hash form kept for callers; prefer standalone page paths."""
    return f"equipment/{slug or hero_equipment_detail_slug(hero_name)}"


def hero_equipment_detail_element_id(slug: str) -> str:
    return f"equipment-hero-{slug}"


def equipment_kind_label(kind: str) -> str:
    return {"normal": "普通", "super": "超级", "food": "美食社"}.get(kind, kind or "普通")


def unique_route_bonds(comp: dict[str, Any]) -> str:
    route: list[str] = []
    mature_bond = comp.get("mature_stage", {}).get("bond") or comp.get("main_bond")
    for stage in comp.get("transition_stages", []):
        bond = stage.get("bond")
        if not bond or bond == mature_bond or bond in route:
            continue
        route.append(bond)
    if mature_bond and mature_bond not in route:
        route.append(mature_bond)
    return " → ".join(route[:4]) or str(comp.get("main_bond", "样本不足"))


def render_sortable_table_panel(
    *,
    panel_id: str,
    title: str,
    subtitle: str,
    note: str,
    headers: list[tuple[str, str]],
    rows: list[list[dict[str, Any]]],
) -> str:
    header_parts: list[str] = []
    for index, (label, sort_type) in enumerate(headers):
        th_class = "sortable sort-asc" if index == 0 else "sortable"
        data_dir = ' data-dir="asc"' if index == 0 else ""
        header_parts.append(
            f'<th class="{th_class}" data-sort="{esc(sort_type)}"{data_dir}>{esc(label)}</th>'
        )
    header_html = "\n".join(header_parts)
    body_rows: list[str] = []
    for row in rows:
        cells = "\n".join(
            (
                f'<td data-sort="{esc(cell["sort"])}">{cell["html"]}</td>'
                if cell.get("html")
                else f'<td data-sort="{esc(cell["sort"])}">{esc(cell["text"])}</td>'
            )
            for cell in row
        )
        body_rows.append(f"<tr>{cells}</tr>")
    body_html = "\n".join(body_rows) or '<tr><td colspan="{0}">样本不足</td></tr>'.format(
        len(headers)
    )
    initial_sort_label = esc(headers[0][0]) if headers else ""
    note_html = f'<div class="note">{esc(note)}</div>' if note.strip() else ""

    return f"""
  <div class="panel-section" id="{esc(panel_id)}">
    <header class="panel-header">
      <div class="title-row">
        <h2>{esc(title)}</h2>
        <span class="sort-status" data-role="sort-status">当前按 {initial_sort_label} 升序</span>
      </div>
      <div class="sub">{esc(subtitle)}</div>
      {note_html}
    </header>
    <div class="table-wrap">
      <table class="sortable-table">
        <thead><tr>{header_html}</tr></thead>
        <tbody>{body_html}</tbody>
      </table>
    </div>
  </div>
  <script>
(function() {{
  const panel = document.getElementById({json.dumps(panel_id)});
  if (!panel) return;
  const sortStatusEl = panel.querySelector('[data-role="sort-status"]');
  function updateSortStatus(th, dir) {{
    if (!sortStatusEl || !th) return;
    const label = th.textContent.replace(/\\s*[▲▼]\\s*$/, "").trim();
    sortStatusEl.textContent = `当前按 ${{label}} ${{dir === "desc" ? "降序" : "升序"}}`;
  }}
  panel.querySelectorAll("th.sortable").forEach((th, colIndex) => {{
    th.addEventListener("click", () => {{
      const table = th.closest("table");
      const tbody = table.querySelector("tbody");
      const tableRows = Array.from(tbody.querySelectorAll("tr"));
      const sortType = th.dataset.sort || "text";
      const isActive = th.classList.contains("sort-asc") || th.classList.contains("sort-desc");
      const newDir = isActive && th.dataset.dir === "asc" ? "desc" : "asc";
      table.querySelectorAll("th.sortable").forEach((header) => {{
        header.dataset.dir = "";
        header.classList.remove("sort-asc", "sort-desc");
      }});
      th.dataset.dir = newDir;
      th.classList.add(newDir === "asc" ? "sort-asc" : "sort-desc");
      updateSortStatus(th, newDir);
      tableRows.sort((left, right) => {{
        const leftVal = left.cells[colIndex]?.dataset.sort ?? "";
        const rightVal = right.cells[colIndex]?.dataset.sort ?? "";
        if (sortType === "numeric") {{
          const leftNum = parseFloat(leftVal);
          const rightNum = parseFloat(rightVal);
          const safeLeft = Number.isFinite(leftNum) ? leftNum : Number.MAX_VALUE;
          const safeRight = Number.isFinite(rightNum) ? rightNum : Number.MAX_VALUE;
          return newDir === "asc" ? safeLeft - safeRight : safeRight - safeLeft;
        }}
        const cmp = String(leftVal).localeCompare(String(rightVal), "zh-CN");
        return newDir === "asc" ? cmp : -cmp;
      }});
      tableRows.forEach((row) => tbody.appendChild(row));
    }});
  }});
  const initialSortHeader = panel.querySelector("th.sort-asc, th.sort-desc");
  if (initialSortHeader) {{
    updateSortStatus(initialSortHeader, initialSortHeader.dataset.dir || "asc");
  }}
}})();
  </script>
"""


def split_strategy_label(label: str) -> tuple[str, str]:
    if " / " in label:
        bond, carries = label.split(" / ", 1)
        return bond.strip(), carries.strip()
    return label.strip(), ""


def html_strategy_cell(label: str) -> str:
    bond, carries = split_strategy_label(label)
    carries_html = (
        f'<div class="carries">{esc(carries)}</div>' if carries else ""
    )
    return (
        f'<div class="strategy-cell">'
        f'<div class="bond">{esc(bond)}</div>'
        f"{carries_html}"
        f"</div>"
    )


def html_top_strategies_cell(strategies: list[dict[str, Any]], limit: int = 3) -> str:
    if not strategies:
        return '<span class="muted">样本不足</span>'
    items: list[str] = []
    for item in strategies[:limit]:
        bond, _ = split_strategy_label(item["label"])
        title = esc(f"{item['label']}({item['samples']})")
        items.append(
            f'<span class="strategy-brief" title="{title}">'
            f"{esc(bond)}({item['samples']})</span>"
        )
    return '<div class="strategy-list">' + "".join(items) + "</div>"


def _html_table_cell(
    text: str,
    *,
    sort_value: str | float | int | None = None,
    html: str | None = None,
) -> dict[str, Any]:
    return {
        "text": text,
        "sort": str(sort_value if sort_value is not None else text),
        "html": html,
    }


def render_card_prefix_table_panel(
    data: dict[str, Any], prefix_type: str, *, panel_id: str
) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    cards = data["rankings"]["cards"]
    prefix_rows = cards.get("single_cards_by_prefix", {}).get(prefix_type, [])
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = CARD_MERGE_NOTES.get(prefix_type, "")

    headers: list[tuple[str, str]] = [
        ("组内排名", "numeric"),
        ("卡牌", "text"),
        ("样本", "numeric"),
        ("每局平均", "numeric"),
        ("修正名次", "numeric"),
        ("平均名次", "numeric"),
        ("前四率", "numeric"),
        ("吃鸡率", "numeric"),
    ]
    if prefix_type == "蓝":
        headers.extend(
            [
                ("修正队伍名次", "numeric"),
                ("队伍名次", "numeric"),
                ("队伍前二率", "numeric"),
            ]
        )

    team_rank_map = {
        row["key"]: row
        for row in cards.get("blue_cards_team_rank_by_prefix", {}).get("蓝", [])
    }
    table_rows: list[list[dict[str, Any]]] = []
    for row in prefix_rows:
        rank_label = "—" if row.get("prefix_rank") is None else str(row["prefix_rank"])
        cells = [
            _html_table_cell(rank_label, sort_value=row.get("prefix_rank", 9999)),
            _html_table_cell(row["key"], sort_value=row["key"]),
            _html_table_cell(str(row["appearances"]), sort_value=row["appearances"]),
            _html_table_cell(
                f"{row.get('avg_appearances_per_match', 0):.2f}",
                sort_value=row.get("avg_appearances_per_match", 0),
            ),
            _html_table_cell(
                render_card_metric(row.get("adjusted_avg_rank")),
                sort_value=row.get("adjusted_avg_rank", 999),
            ),
            _html_table_cell(
                render_card_metric(row.get("avg_rank")),
                sort_value=row.get("avg_rank", 999),
            ),
            _html_table_cell(
                render_pct(row.get("top4_rate")),
                sort_value=row.get("top4_rate", -1),
            ),
            _html_table_cell(
                render_pct(row.get("win_rate")),
                sort_value=row.get("win_rate", -1),
            ),
        ]
        if prefix_type == "蓝":
            team = team_rank_map.get(row["key"], {})
            cells.extend(
                [
                    _html_table_cell(
                        render_card_metric(team.get("adjusted_avg_rank")) if team else "—",
                        sort_value=team.get("adjusted_avg_rank", 999),
                    ),
                    _html_table_cell(
                        render_card_metric(team.get("avg_rank")) if team else "—",
                        sort_value=team.get("avg_rank", 999),
                    ),
                    _html_table_cell(
                        render_pct(team.get("team_top2_rate")) if team else "—",
                        sort_value=team.get("team_top2_rate", -1),
                    ),
                ]
            )
        table_rows.append(cells)

    return render_sortable_table_panel(
        panel_id=panel_id,
        title=f"{prefix_type}类单卡排名",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=table_rows,
    )


def render_duo_composition_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    duo_rows = data["rankings"].get("duo_composition_synergy", [])
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = "基于同队两家的最终策略组合与重算队伍排名，仅作双排分工参考；阵容分列展示以避免拥挤。"

    headers = [
        ("阵容A", "text"),
        ("阵容B", "text"),
        ("队伍名次", "numeric"),
        ("相对基线提升", "numeric"),
        ("队伍前二率", "numeric"),
        ("队伍吃鸡率", "numeric"),
        ("样本", "numeric"),
        ("置信度", "text"),
    ]
    table_rows: list[list[dict[str, Any]]] = []
    for row in duo_rows:
        table_rows.append(
            [
                {
                    "text": row["strategy_a"],
                    "sort": row["strategy_a"],
                    "html": html_strategy_cell(row["strategy_a"]),
                },
                {
                    "text": row["strategy_b"],
                    "sort": row["strategy_b"],
                    "html": html_strategy_cell(row["strategy_b"]),
                },
                _html_table_cell(
                    f"{row['team_avg_rank']:.2f}",
                    sort_value=row["team_avg_rank"],
                ),
                _html_table_cell(
                    f"{row['team_lift_vs_baseline']:.2f}",
                    sort_value=row["team_lift_vs_baseline"],
                ),
                _html_table_cell(
                    render_pct(row["team_top2_rate"]),
                    sort_value=row["team_top2_rate"],
                ),
                _html_table_cell(
                    render_pct(row["team_win_rate"]),
                    sort_value=row["team_win_rate"],
                ),
                _html_table_cell(str(row["appearances"]), sort_value=row["appearances"]),
                _html_table_cell(row["confidence"], sort_value=row["confidence"]),
            ]
        )

    return render_sortable_table_panel(
        panel_id=panel_id,
        title="双人阵容配合推荐",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=table_rows,
    )


def render_low_cost_carry_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    rows = [
        row
        for row in data["rankings"].get("low_cost_carry_three_star_difficulty", [])
        if row.get("is_low_cost")
    ]
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = (
        "同场多家阵容需要同一个低费3星主C时，即使阵容路线不同，也计入同行压力；"
        "主要阵容仅展示羁绊标题与样本数，完整阵容可悬停查看。"
    )

    headers = [
        ("棋子", "text"),
        ("费用", "numeric"),
        ("平均同场需求", "numeric"),
        ("最高同场需求", "numeric"),
        ("多家需求对局率", "numeric"),
        ("平均名次", "numeric"),
        ("样本", "numeric"),
        ("主要阵容", "text"),
    ]
    table_rows: list[list[dict[str, Any]]] = []
    for row in rows:
        strategies = row.get("top_strategies", [])
        table_rows.append(
            [
                _html_table_cell(row["hero_name"], sort_value=row["hero_name"]),
                _html_table_cell(
                    str(row.get("tier") or "—"),
                    sort_value=row.get("tier") or 99,
                ),
                _html_table_cell(
                    f"{row['avg_same_match_needers']:.2f}",
                    sort_value=row["avg_same_match_needers"],
                ),
                _html_table_cell(
                    str(row["max_same_match_needers"]),
                    sort_value=row["max_same_match_needers"],
                ),
                _html_table_cell(
                    render_pct(row["multi_needer_match_rate"]),
                    sort_value=row["multi_needer_match_rate"],
                ),
                _html_table_cell(f"{row['avg_rank']:.2f}", sort_value=row["avg_rank"]),
                _html_table_cell(str(row["appearances"]), sort_value=row["appearances"]),
                {
                    "text": "；".join(
                        f"{item['label']}({item['samples']})" for item in strategies[:3]
                    )
                    or "样本不足",
                    "sort": "；".join(item["label"] for item in strategies[:3]) or "",
                    "html": html_top_strategies_cell(strategies),
                },
            ]
        )

    return render_sortable_table_panel(
        panel_id=panel_id,
        title="低费主C热门程度",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=table_rows,
    )


def render_primary_bond_strength_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    primary_bond = data["rankings"].get("primary_bond_strength", {})
    rows = primary_bond.get("rows", [])
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = primary_bond.get(
        "definition",
        "按最终阵容激活数量最大的羁绊统计；数量相同则同时计入；按羁绊名聚合。",
    )
    audit_note = format_primary_bond_audit_text(primary_bond)
    if audit_note:
        note = f"{note} {audit_note}"

    headers = [
        ("强度排名", "numeric"),
        ("主羁绊", "text"),
        ("归类", "text"),
        ("样本", "numeric"),
        ("修正名次", "numeric"),
        ("平均名次", "numeric"),
        ("前四率", "numeric"),
        ("后四率", "numeric"),
        ("吃鸡率", "numeric"),
        ("常见激活数量", "text"),
        ("常见激活档位", "text"),
        ("归类来源", "text"),
    ]
    table_rows: list[list[dict[str, Any]]] = []
    for row in rows:
        table_rows.append(
            [
                _html_table_cell(str(row["strength_rank"]), sort_value=row["strength_rank"]),
                _html_table_cell(row["bond"], sort_value=row["bond"]),
                _html_table_cell(row.get("category", row["bond"]), sort_value=row.get("category", row["bond"])),
                _html_table_cell(str(row["appearances"]), sort_value=row["appearances"]),
                _html_table_cell(
                    f"{row['adjusted_avg_rank']:.2f}",
                    sort_value=row["adjusted_avg_rank"],
                ),
                _html_table_cell(f"{row['avg_rank']:.2f}", sort_value=row["avg_rank"]),
                _html_table_cell(render_pct(row["top4_rate"]), sort_value=row["top4_rate"]),
                _html_table_cell(render_pct(row["bottom4_rate"]), sort_value=row["bottom4_rate"]),
                _html_table_cell(render_pct(row["win_rate"]), sort_value=row["win_rate"]),
                _html_table_cell(
                    row.get("common_activation_summary", "—"),
                    sort_value=row.get("common_activation_summary", ""),
                ),
                _html_table_cell(
                    row.get("common_tier_summary", "—"),
                    sort_value=row.get("common_tier_summary", ""),
                ),
                _html_table_cell(
                    format_primary_bond_source_distribution(row.get("source_distribution")),
                    sort_value=format_primary_bond_source_distribution(
                        row.get("source_distribution")
                    ),
                ),
            ]
        )

    return render_sortable_table_panel(
        panel_id=panel_id,
        title="主羁绊强度排行",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=table_rows,
    )


def html_variant_board_cards(comp: dict[str, Any]) -> str:
    variants = comp.get("variants", {})
    cards: list[str] = []
    for level in ("7", "8", "9"):
        variant = variants.get(level, {})
        heroes = variant.get("heroes", [])
        hero_chips = "".join(f'<span class="hero-chip">{esc(hero)}</span>' for hero in heroes)
        sample_note = (
            f"样本 {variant.get('sample_count', 0)}"
            if variant.get("sample_count")
            else "推导阵容"
            if variant.get("source") == "derived"
            else ""
        )
        cards.append(
            f"""
            <article class="board-card">
              <div class="board-head">
                <span class="level-badge">Lv{level}</span>
                <span class="source-badge">{esc(variant.get("source", "—"))}</span>
                <span class="conf-badge">{esc(variant.get("confidence", "—"))}</span>
                {f'<span class="sample-badge">{esc(sample_note)}</span>' if sample_note else ""}
              </div>
              <p class="bond-note">{esc(variant.get("bond_note", "—"))}</p>
              <div class="hero-chips">{hero_chips or '<span class="muted">样本不足</span>'}</div>
            </article>
            """
        )
    return f'<div class="board-grid">{"".join(cards)}</div>'


def html_comp_detail_page(
    comp: dict[str, Any],
    *,
    style: str,
    page_index: int,
) -> str:
    stats = comp["stats"]
    evidence = comp.get("confidence_evidence") or {}
    difficulty = comp.get("difficulty", {})
    archetype = comp.get("archetype", "未分类") or "未分类"
    trend_label = format_trend_label(comp.get("trend"))
    carry_text = " > ".join(
        f"P{item.get('carry_rank', idx)} {item['hero_name']}({render_pct(item['share'])})"
        for idx, item in enumerate(comp.get("main_carries", [])[:3], start=1)
    )
    req_text = "；".join(
        f"{row['hero_name']} {row['recommended_min_stars']}星起(前四均{row['avg_stars_top4']:.1f}，三件套{render_pct(row['three_item_rate'])})"
        for row in comp.get("carry_requirements", [])[:3]
    )
    equip_parts: list[str] = []
    for note in comp.get("carry_equipment_notes", [])[:3]:
        important = [
            item for item in note["items"] if item["label"] in ("疑似刚需", "高价值")
        ][:3]
        if important:
            equip_parts.append(
                f"{note['hero_name']}："
                + "、".join(
                    f"{item['equipment_name']}({item['label']})"
                    for item in important
                )
            )
    jiujiu_parts: list[str] = []
    for req in comp.get("jiujiu_requirements", []):
        wearers = "、".join(
            f"{item['hero_name']}({render_pct(item['share'])})"
            for item in req.get("recommended_wearers", [])[:3]
        )
        jiujiu_parts.append(
            f"{req['recommended_jiujiu']}（{render_pct(req['dependency_rate'])}需啾啾开"
            f"{req['trait']}-{req['target_tier']}，推荐：{wearers or '待观察'}）"
        )
    bonds = "、".join(
        f"{item['bond']}({render_pct(item['share'])})" for item in comp.get("common_bonds", [])[:5]
    )
    breakdown = " · ".join(
        f"{row['play_style']}{render_pct(row['share'])}"
        for row in comp.get("play_style_breakdown", [])[:3]
    )
    star_pressure = (
        f"三星均{difficulty.get('avg_three_star_units', 0):.2f} / "
        f"前四均{difficulty.get('avg_top4_three_star_units', 0):.2f} / "
        f"同行{difficulty.get('avg_same_match_contest', 0):.2f}"
        f"({difficulty.get('contest_basis', '阵容相似')})"
        if difficulty
        else "样本不足"
    )
    popularity = comp.get("popularity", {})
    pop_text = (
        f"玩家占比 {render_pct(popularity.get('pick_rate', 0))} / "
        f"对局出现 {render_pct(popularity.get('match_rate', 0))}"
        if popularity
        else "—"
    )
    mature_stats = comp.get("mature_stats") or stats
    transition_stats = comp.get("transition_stats")
    mature_text = (
        f"Avg {mature_stats.get('avg_rank', 0):.2f} / "
        f"Top4 {render_pct(mature_stats.get('top4_rate'))} / n={mature_stats.get('appearances', 0)}"
        if mature_stats.get("appearances")
        else "—"
    )
    transition_text = (
        f"Avg {transition_stats.get('avg_rank', 0):.2f} / "
        f"Top4 {render_pct(transition_stats.get('top4_rate'))} / n={transition_stats.get('appearances', 0)}"
        if transition_stats and transition_stats.get("appearances")
        else ("存在过渡样本" if comp.get("transition_stages") else "—")
    )
    cluster_lines = format_cluster_merge_reason_lines(comp)
    cluster_html = "".join(f"<p>{esc(line.lstrip('- '))}</p>" for line in cluster_lines)
    failure_text = format_recommendation_failure_reasons_text(evidence)
    inversion_text = format_stage_inversion_diagnostics_text(comp)
    expensive_note = [
        row["hero_name"]
        for row in comp.get("carry_requirements", [])[:3]
        if row.get("high_cost_three_star_dependency")
    ]
    expensive_html = (
        f"<p><strong>成型成本提醒：</strong>{esc('、'.join(expensive_note))} 的三星高费样本会拉高上限，常规推荐按 2 星门槛评估。</p>"
        if expensive_note
        else ""
    )
    return f"""
    <section class="comp-page" data-style="{esc(style)}" data-index="{page_index}">
      <div class="comp-page-inner">
        <header class="comp-page-head">
          <div class="comp-head">
            <span class="badge">{esc(style)}</span>
            <span class="trend-badge">{esc(trend_label)}</span>
            <span class="conf-pill">{esc(comp.get('confidence', '—'))}置信</span>
          </div>
          <h2>{esc(comp['label'])}</h2>
          <div class="metrics">
            <b>Avg {stats['avg_rank']:.2f}</b>
            <b>Top4 {render_pct(stats['top4_rate'])}</b>
            <b>吃鸡 {render_pct(stats['win_rate'])}</b>
            <b>raw {evidence.get('raw_n', stats.get('appearances', '—'))}</b>
            <b>weighted {evidence.get('weighted_n', stats.get('weighted_appearances', '—'))}</b>
            <b>n_eff {evidence.get('n_eff', stats.get('n_eff', '—'))}</b>
          </div>
        </header>
        <div class="detail-grid">
          <div class="detail-panel">
            <h3>阵容概览</h3>
            <p><strong>玩法原型：</strong>{esc(archetype)}</p>
            <p><strong>归类证据：</strong>{esc(format_archetype_signals_text(comp))}</p>
            {cluster_html}
            <p><strong>主C：</strong>{esc(carry_text or '样本不足')}</p>
            <p><strong>成型门槛：</strong>{esc(req_text or '样本不足')}</p>
            {expensive_html}
            <p><strong>成熟表现：</strong>{esc(mature_text)}</p>
            <p><strong>过渡表现：</strong>{esc(transition_text)}</p>
            <p><strong>成熟/过渡倒挂：</strong>{esc(inversion_text)}</p>
            <p><strong>版本趋势：</strong>{esc(format_trend_detail_text(comp.get('trend')))}</p>
            <p><strong>三星压力：</strong>{esc(star_pressure)}</p>
            <p><strong>热门程度：</strong>{esc(pop_text)}</p>
            <p><strong>路线：</strong>{esc(unique_route_bonds(comp))}</p>
            <p><strong>类型样本：</strong>{esc(breakdown or style)}</p>
            <p><strong>常见羁绊：</strong>{esc(bonds or '无稳定羁绊')}</p>
          </div>
          <div class="detail-panel">
            <h3>置信与评分</h3>
            <p><strong>置信证据：</strong>{esc(format_confidence_criteria_text(evidence))}</p>
            {f'<p><strong>低置信提示：</strong>{esc(failure_text)}</p>' if not evidence.get("recommendation_eligible", True) else ""}
            <p><strong>评分分解：</strong>{esc(format_score_breakdown_text(comp.get('score_breakdown')))}</p>
            <h3>关键装备与啾啾</h3>
            <p><strong>主C关键装备：</strong>{esc('；'.join(equip_parts) or '样本不足')}</p>
            <p><strong>啾啾成型：</strong>{esc('；'.join(jiujiu_parts) or '无明确依赖')}</p>
          </div>
        </div>
        <div class="board-section">
          <h3>7 / 8 / 9 级推荐阵容</h3>
          {html_variant_board_cards(comp)}
        </div>
      </div>
    </section>
    """


def render_composition_recommendations_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    recommendations = data["rankings"].get("composition_recommendations", {})
    pages: list[str] = []
    page_index = 0
    style_counts = {style: len(recommendations.get(style, [])) for style in PLAY_STYLES}
    recommend_count = sum(style_counts.values())
    for style in PLAY_STYLES:
        for comp in recommendations.get(style, []):
            pages.append(
                html_comp_detail_page(
                    comp,
                    style=style,
                    page_index=page_index,
                )
            )
            page_index += 1
    pages_html = "\n".join(pages) or '<section class="comp-page active"><p class="empty">当前样本不足。</p></section>'
    total_pages = max(page_index, 1)
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    style_filter = "".join(
        f'<button type="button" class="style-filter" data-style="{esc(style)}">{esc(style)} ({count})</button>'
        for style, count in style_counts.items()
    )
    return f"""
  <div class="panel-section comp-panel" id="{esc(panel_id)}">
    <header class="panel-header">
      <h2>阵容推荐详情</h2>
      <div class="sub">{esc(subtitle)} · 共 {recommend_count} 套 · 每页一套阵容，按 7/8/9 级推荐</div>
    </header>
    <div class="pager-bar">
      <div class="pager-controls">
        <button type="button" class="pager-btn" data-role="prev-page">上一页</button>
        <button type="button" class="pager-btn" data-role="next-page">下一页</button>
        <span class="page-status" data-role="page-status">第 1 / {total_pages} 页</span>
      </div>
      <div class="pager-controls">
        <span class="filter-label">类型</span>
        <button type="button" class="style-filter active" data-style="all">全部</button>
        {style_filter}
      </div>
    </div>
    <div class="comp-pages" data-role="comp-pages">
      {pages_html}
    </div>
  </div>
  <script>
(function() {{
  const panel = document.getElementById({json.dumps(panel_id)});
  if (!panel) return;
  const pages = Array.from(panel.querySelectorAll(".comp-page"));
  const pageStatus = panel.querySelector('[data-role="page-status"]');
  const prevBtn = panel.querySelector('[data-role="prev-page"]');
  const nextBtn = panel.querySelector('[data-role="next-page"]');
  let activeStyle = "all";
  let activeIndex = 0;

  function visiblePages() {{
    return pages.filter((page) => {{
      if (activeStyle !== "all" && page.dataset.style !== activeStyle) return false;
      return true;
    }});
  }}

  function showPage(index) {{
    const filtered = visiblePages();
    if (!filtered.length) {{
      pages.forEach((page) => page.classList.remove("active"));
      if (pageStatus) pageStatus.textContent = "当前筛选暂无阵容";
      if (prevBtn) prevBtn.disabled = true;
      if (nextBtn) nextBtn.disabled = true;
      return;
    }}
    activeIndex = Math.max(0, Math.min(index, filtered.length - 1));
    pages.forEach((page) => page.classList.remove("active"));
    filtered[activeIndex].classList.add("active");
    const current = filtered[activeIndex];
    if (pageStatus) {{
      pageStatus.textContent =
        `第 ${{activeIndex + 1}} / ${{filtered.length}} 页 · ${{current.dataset.style}}`;
    }}
    if (prevBtn) prevBtn.disabled = activeIndex <= 0;
    if (nextBtn) nextBtn.disabled = activeIndex >= filtered.length - 1;
  }}

  panel.querySelectorAll(".style-filter").forEach((button) => {{
    button.addEventListener("click", () => {{
      panel.querySelectorAll(".style-filter").forEach((item) => item.classList.remove("active"));
      button.classList.add("active");
      activeStyle = button.dataset.style;
      showPage(0);
    }});
  }});

  if (prevBtn) prevBtn.addEventListener("click", () => showPage(activeIndex - 1));
  if (nextBtn) nextBtn.addEventListener("click", () => showPage(activeIndex + 1));
  if (pages.length) {{
    pages[0].classList.add("active");
    showPage(0);
  }}
}})();
  </script>
"""


def build_jiujiu_comp_table_rows(data: dict[str, Any]) -> list[list[dict[str, Any]]]:
    rows: list[list[dict[str, Any]]] = []
    seen: set[tuple[str, str]] = set()
    recommendations = data["rankings"].get("composition_recommendations", {})
    for style in PLAY_STYLES:
        for comp in recommendations.get(style, []):
            if not comp.get("jiujiu_requirements"):
                continue
            stats = comp["stats"]
            for req in comp["jiujiu_requirements"]:
                key = (comp["label"], req.get("recommended_jiujiu", ""))
                if key in seen:
                    continue
                seen.add(key)
                wearers = "、".join(
                    f"{item['hero_name']}({render_pct(item['share'])})"
                    for item in req.get("recommended_wearers", [])[:3]
                ) or "待观察"
                target_bond = f"{req.get('trait', '—')}-{req.get('target_tier', '—')}"
                rows.append(
                    [
                        {
                            "text": comp["label"],
                            "sort": comp["label"],
                            "html": html_strategy_cell(comp["label"]),
                        },
                        _html_table_cell(style, sort_value=style),
                        _html_table_cell(req.get("recommended_jiujiu", "—"), sort_value=req.get("recommended_jiujiu", "")),
                        _html_table_cell(target_bond, sort_value=target_bond),
                        _html_table_cell(
                            render_pct(req.get("dependency_rate", 0)),
                            sort_value=req.get("dependency_rate", 0),
                        ),
                        _html_table_cell(wearers, sort_value=wearers),
                        _html_table_cell(f"{stats['avg_rank']:.2f}", sort_value=stats["avg_rank"]),
                        _html_table_cell(render_pct(stats["top4_rate"]), sort_value=stats["top4_rate"]),
                        _html_table_cell(str(stats["appearances"]), sort_value=stats["appearances"]),
                        _html_table_cell(comp.get("confidence", "—"), sort_value=comp.get("confidence", "")),
                    ]
                )
    return rows


def render_jiujiu_comps_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = "仅展示推荐阵容中存在明确啾啾成型依赖的策略；点击表头可排序。"
    headers = [
        ("阵容", "text"),
        ("类型", "text"),
        ("啾啾", "text"),
        ("目标羁绊", "text"),
        ("依赖率", "numeric"),
        ("推荐穿戴", "text"),
        ("平均名次", "numeric"),
        ("前四率", "numeric"),
        ("样本", "numeric"),
        ("置信度", "text"),
    ]
    return render_sortable_table_panel(
        panel_id=panel_id,
        title="带啾啾阵容推荐",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=build_jiujiu_comp_table_rows(data),
    )


def build_jiujiu_wearer_table_rows(data: dict[str, Any]) -> list[list[dict[str, Any]]]:
    rows: list[list[dict[str, Any]]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for item in data["rankings"].get("jiujiu", {}).get("jiujiu_rankings", []):
        item_name = item["equipment_name"]
        for comp in item.get("recommended_comps", []):
            for wearer in comp.get("recommended_wearers", []):
                key = (item_name, wearer["hero_name"], "阵容绑定", comp.get("family_label", ""))
                if key in seen:
                    continue
                seen.add(key)
                rows.append(
                    [
                        _html_table_cell(item_name, sort_value=item_name),
                        _html_table_cell(wearer["hero_name"], sort_value=wearer["hero_name"]),
                        _html_table_cell("阵容绑定", sort_value="阵容绑定"),
                        {
                            "text": comp.get("family_label", "—"),
                            "sort": comp.get("family_label", ""),
                            "html": html_strategy_cell(comp.get("family_label", "—")),
                        },
                        _html_table_cell(str(wearer["appearances"]), sort_value=wearer["appearances"]),
                        _html_table_cell(render_pct(wearer.get("share", 0)), sort_value=wearer.get("share", 0)),
                        _html_table_cell(
                            f"{comp['avg_rank']:.2f}" if comp.get("avg_rank") is not None else "—",
                            sort_value=comp.get("avg_rank", 999),
                        ),
                        _html_table_cell(
                            render_pct(comp.get("top4_rate", 0)),
                            sort_value=comp.get("top4_rate", -1),
                        ),
                    ]
                )
        for hero in item.get("recommended_heroes", []):
            key = (item_name, hero["hero_name"], "棋子增益", "")
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                [
                    _html_table_cell(item_name, sort_value=item_name),
                    _html_table_cell(hero["hero_name"], sort_value=hero["hero_name"]),
                    _html_table_cell("棋子增益", sort_value="棋子增益"),
                    _html_table_cell("—", sort_value=""),
                    _html_table_cell(str(hero["appearances"]), sort_value=hero["appearances"]),
                    _html_table_cell("—", sort_value=-1),
                    _html_table_cell(f"{hero['avg_rank']:.2f}", sort_value=hero["avg_rank"]),
                    _html_table_cell(render_pct(hero.get("top4_rate", 0)), sort_value=hero.get("top4_rate", -1)),
                ]
            )
    return rows


def render_jiujiu_wearers_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = "汇总啾啾在阵容绑定与棋子增益两类证据下的推荐穿戴棋子；点击表头可排序。"
    headers = [
        ("啾啾", "text"),
        ("棋子", "text"),
        ("证据类型", "text"),
        ("关联阵容", "text"),
        ("样本", "numeric"),
        ("占比", "numeric"),
        ("平均名次", "numeric"),
        ("前四率", "numeric"),
    ]
    return render_sortable_table_panel(
        panel_id=panel_id,
        title="佩戴啾啾棋子推荐",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=build_jiujiu_wearer_table_rows(data),
    )


def render_equipment_panel(
    *,
    panel_id: str,
    title: str,
    subtitle: str,
    note: str,
    headers: list[tuple[str, str]],
    rows: list[dict[str, Any]],
) -> str:
    header_parts: list[str] = []
    for index, (label, sort_type) in enumerate(headers):
        th_class = "sortable sort-asc" if index == 0 else "sortable"
        data_dir = ' data-dir="asc"' if index == 0 else ""
        header_parts.append(
            f'<th class="{th_class}" data-sort="{esc(sort_type)}"{data_dir}>{esc(label)}</th>'
        )
    header_html = "\n".join(header_parts)
    body_rows: list[str] = []
    trait_options: set[str] = set()
    for row in rows:
        traits = row.get("traits", [])
        trait_options.update(traits)
        trait_attr = esc(",".join(traits))
        cells = "\n".join(
            (
                f'<td data-sort="{esc(cell["sort"])}">{cell["html"]}</td>'
                if cell.get("html")
                else f'<td data-sort="{esc(cell["sort"])}">{esc(cell["text"])}</td>'
            )
            for cell in row["cells"]
        )
        body_rows.append(
            f'<tr data-tier="{esc(str(row.get("tier") or ""))}" '
            f'data-traits="{trait_attr}" '
            f'data-search="{esc(row.get("search_text", ""))}">{cells}</tr>'
        )
    body_html = "\n".join(body_rows) or f'<tr><td colspan="{len(headers)}">样本不足</td></tr>'
    trait_buttons = "".join(
        f'<button type="button" class="trait-filter" data-trait="{esc(trait)}">{esc(trait)}</button>'
        for trait in sorted(trait_options)
    )
    initial_sort_label = esc(headers[0][0]) if headers else ""
    note_html = f'<div class="note">{esc(note)}</div>' if note.strip() else ""

    return f"""
  <div class="panel-section equipment-panel" id="{esc(panel_id)}">
    <header class="panel-header">
      <div class="title-row">
        <h2>{esc(title)}</h2>
        <span class="sort-status" data-role="sort-status">当前按 {initial_sort_label} 升序</span>
      </div>
      <div class="sub">{esc(subtitle)}</div>
      {note_html}
    </header>
    <div class="filter-bar">
      <div class="filter-group">
        <span class="filter-label">费用</span>
        <button type="button" class="filter-btn active" data-tier="all">全部</button>
        <button type="button" class="filter-btn" data-tier="1">1费</button>
        <button type="button" class="filter-btn" data-tier="2">2费</button>
        <button type="button" class="filter-btn" data-tier="3">3费</button>
        <button type="button" class="filter-btn" data-tier="4">4费</button>
        <button type="button" class="filter-btn" data-tier="5">5费</button>
      </div>
      <div class="filter-group">
        <span class="filter-label">羁绊</span>
        <button type="button" class="trait-filter active" data-trait="all">全部</button>
        {trait_buttons}
      </div>
      <input type="search" class="search-input" data-role="search-input" placeholder="搜索棋子、装备、三件套…">
      <span class="filter-status" data-role="filter-status">显示全部</span>
    </div>
    <div class="table-wrap">
      <table class="sortable-table" data-role="equipment-table">
        <thead><tr>{header_html}</tr></thead>
        <tbody>{body_html}</tbody>
      </table>
    </div>
  </div>
  <script>
(function() {{
  const panel = document.getElementById({json.dumps(panel_id)});
  if (!panel) return;
  const sortStatusEl = panel.querySelector('[data-role="sort-status"]');
  const filterStatusEl = panel.querySelector('[data-role="filter-status"]');
  const searchInput = panel.querySelector('[data-role="search-input"]');
  const table = panel.querySelector('[data-role="equipment-table"]');
  const tbody = table.querySelector("tbody");
  let activeTier = "all";
  let activeTrait = "all";

  function updateSortStatus(th, dir) {{
    if (!sortStatusEl || !th) return;
    const label = th.textContent.replace(/\\s*[▲▼]\\s*$/, "").trim();
    sortStatusEl.textContent = `当前按 ${{label}} ${{dir === "desc" ? "降序" : "升序"}}`;
  }}

  function applyFilters() {{
    const keyword = (searchInput.value || "").trim().toLowerCase();
    let visible = 0;
    tbody.querySelectorAll("tr").forEach((row) => {{
      const tierMatch = activeTier === "all" || row.dataset.tier === activeTier;
      const traits = (row.dataset.traits || "").split(",").filter(Boolean);
      const traitMatch = activeTrait === "all" || traits.includes(activeTrait);
      const searchMatch = !keyword || (row.dataset.search || "").toLowerCase().includes(keyword);
      const show = tierMatch && traitMatch && searchMatch;
      row.classList.toggle("hidden", !show);
      if (show) visible += 1;
    }});
    if (filterStatusEl) {{
      if (activeTier === "all" && activeTrait === "all" && !keyword) {{
        filterStatusEl.textContent = "显示全部";
      }} else if (keyword) {{
        filterStatusEl.textContent = `筛选后 ${{visible}} 条 · 关键词「${{keyword}}」`;
      }} else {{
        filterStatusEl.textContent = `筛选后 ${{visible}} 条`;
      }}
    }}
  }}

  panel.querySelectorAll(".filter-btn").forEach((button) => {{
    button.addEventListener("click", () => {{
      panel.querySelectorAll(".filter-btn").forEach((item) => item.classList.remove("active"));
      button.classList.add("active");
      activeTier = button.dataset.tier;
      applyFilters();
    }});
  }});
  panel.querySelectorAll(".trait-filter").forEach((button) => {{
    button.addEventListener("click", () => {{
      panel.querySelectorAll(".trait-filter").forEach((item) => item.classList.remove("active"));
      button.classList.add("active");
      activeTrait = button.dataset.trait;
      applyFilters();
    }});
  }});
  if (searchInput) searchInput.addEventListener("input", applyFilters);

  panel.querySelectorAll("th.sortable").forEach((th, colIndex) => {{
    th.addEventListener("click", () => {{
      const visibleRows = Array.from(tbody.querySelectorAll("tr:not(.hidden)"));
      const hiddenRows = Array.from(tbody.querySelectorAll("tr.hidden"));
      const sortType = th.dataset.sort || "text";
      const isActive = th.classList.contains("sort-asc") || th.classList.contains("sort-desc");
      const newDir = isActive && th.dataset.dir === "asc" ? "desc" : "asc";
      table.querySelectorAll("th.sortable").forEach((header) => {{
        header.dataset.dir = "";
        header.classList.remove("sort-asc", "sort-desc");
      }});
      th.dataset.dir = newDir;
      th.classList.add(newDir === "asc" ? "sort-asc" : "sort-desc");
      updateSortStatus(th, newDir);
      visibleRows.sort((left, right) => {{
        const leftVal = left.cells[colIndex]?.dataset.sort ?? "";
        const rightVal = right.cells[colIndex]?.dataset.sort ?? "";
        if (sortType === "numeric") {{
          const leftNum = parseFloat(leftVal);
          const rightNum = parseFloat(rightVal);
          const safeLeft = Number.isFinite(leftNum) ? leftNum : Number.MAX_VALUE;
          const safeRight = Number.isFinite(rightNum) ? rightNum : Number.MAX_VALUE;
          return newDir === "asc" ? safeLeft - safeRight : safeRight - safeLeft;
        }}
        const cmp = String(leftVal).localeCompare(String(rightVal), "zh-CN");
        return newDir === "asc" ? cmp : -cmp;
      }});
      visibleRows.forEach((row) => tbody.appendChild(row));
      hiddenRows.forEach((row) => tbody.appendChild(row));
    }});
  }});

  const initialSortHeader = panel.querySelector("th.sort-asc, th.sort-desc");
  if (initialSortHeader) {{
    updateSortStatus(initialSortHeader, initialSortHeader.dataset.dir || "asc");
  }}
  applyFilters();
}})();
  </script>
"""


def _format_equipment_item_chips(items: list[dict[str, Any]], limit: int = 4) -> tuple[str, str]:
    if not items:
        return "—", ""
    chips = "".join(
        f'<span class="item-chip">{esc(item["equipment_name"])} '
        f'({render_pct(item.get("selected_rate", 0))}核选, n={item["appearances"]})</span>'
        for item in items[:limit]
    )
    text = " ".join(item["equipment_name"] for item in items[:limit])
    return f'<div class="item-list">{chips}</div>', text


def html_hero_equipment_detail_rows(items: list[dict[str, Any]]) -> str:
    if not items:
        return (
            '<p class="muted">当前棋子没有 raw 样本量 &gt;10 的单装记录。'
            "摘要推荐仍可参考上方概览列。</p>"
        )
    body = []
    for item in items:
        weighted_n = item.get("weighted_appearances")
        n_eff = item.get("n_eff")
        avg_rank = item.get("avg_rank")
        adjusted = item.get("adjusted_avg_rank")
        weighted_n_text = f"{weighted_n:.2f}" if weighted_n is not None else "—"
        n_eff_text = f"{n_eff:.2f}" if n_eff is not None else "—"
        avg_rank_text = f"{avg_rank:.2f}" if avg_rank is not None else "—"
        adjusted_text = f"{adjusted:.2f}" if adjusted is not None else "—"
        body.append(
            "<tr>"
            f'<td>{esc(item["equipment_name"])}</td>'
            f'<td>{esc(equipment_kind_label(item.get("equipment_kind", "normal")))}</td>'
            f'<td>{esc(str(item.get("appearances", 0)))}</td>'
            f'<td>{esc(weighted_n_text)}</td>'
            f'<td>{esc(n_eff_text)}</td>'
            f'<td>{esc(avg_rank_text)}</td>'
            f'<td>{render_pct(item.get("top4_rate"))}</td>'
            f'<td>{render_pct(item.get("top2_rate"))}</td>'
            f'<td>{render_pct(item.get("win_rate"))}</td>'
            f'<td>{esc(adjusted_text)}</td>'
            f'<td>{render_pct(item.get("selected_rate", 0))}</td>'
            f'<td>{esc(item.get("sample_quality", "—"))}</td>'
            "</tr>"
        )
    return f"""
    <div class="table-wrap">
      <table class="detail-equip-table">
        <thead>
          <tr>
            <th>装备</th><th>类型</th><th>raw n</th><th>加权 n</th><th>n_eff</th>
            <th>加权平均名次</th><th>加权前四率</th><th>加权前二率</th><th>加权吃鸡率</th>
            <th>修正名次</th><th>核选率</th><th>样本质量</th>
          </tr>
        </thead>
        <tbody>
          {"".join(body)}
        </tbody>
      </table>
    </div>
    """


def standalone_hero_equipment_css() -> str:
    return """
    :root {
      color-scheme: dark;
      --bg: #0f172a;
      --card: rgba(15, 23, 42, .88);
      --line: rgba(255,255,255,.12);
      --text: #e2e8f0;
      --muted: #94a3b8;
      --accent: #fde68a;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background:
        radial-gradient(circle at top left, rgba(56,189,248,.18), transparent 28%),
        radial-gradient(circle at top right, rgba(251,191,36,.16), transparent 24%),
        var(--bg);
      color: var(--text);
    }
    main {
      max-width: 1100px;
      margin: 0 auto;
      padding: 28px 18px 48px;
    }
    .card {
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 22px;
      padding: 22px;
      box-shadow: 0 18px 50px rgba(0,0,0,.24);
    }
    .eyebrow { color: var(--accent); font-size: 12px; letter-spacing: .12em; font-weight: 700; }
    h1 { margin: 8px 0 6px; font-size: 30px; }
    .sub, .muted { color: var(--muted); }
    .metrics {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin: 14px 0 18px;
    }
    .metrics b {
      background: rgba(255,255,255,.06);
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 13px;
    }
    .item-list { display: flex; flex-wrap: wrap; gap: 8px; }
    .item-chip {
      display: inline-flex;
      gap: 6px;
      background: rgba(253,230,138,.12);
      border: 1px solid rgba(253,230,138,.28);
      border-radius: 999px;
      padding: 4px 10px;
      font-size: 13px;
    }
    .detail-equip-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
      margin-top: 12px;
    }
    .detail-equip-table th, .detail-equip-table td {
      border-bottom: 1px solid rgba(255,255,255,.08);
      padding: 8px 6px;
      text-align: left;
      vertical-align: top;
    }
    .detail-equip-table th { color: var(--accent); white-space: nowrap; }
    .hero-equipment-back {
      display: inline-block;
      margin-top: 18px;
      color: #93c5fd;
      text-decoration: none;
      border-bottom: 1px dashed rgba(147,197,253,.55);
    }
    h3 { color: var(--accent); margin: 18px 0 8px; }
    """


def render_hero_equipment_detail_page(
    rec: dict[str, Any],
    *,
    dashboard_href: str = "../环境分析详情.html#equipment",
) -> str:
    hero_name = rec["hero_name"]
    hero_stats = rec.get("hero_stats") or {}
    tier = hero_stats.get("tier")
    traits = "、".join(rec.get("hero_traits") or []) or "—"
    detail_items = rec.get("detail_items") or []
    normal_html, _ = _format_equipment_item_chips(rec.get("recommended_items") or [])
    super_html, _ = _format_equipment_item_chips(rec.get("recommended_super_items") or [])
    food_html, _ = _format_equipment_item_chips(rec.get("recommended_food_items") or [])
    avg_rank = hero_stats.get("avg_rank")
    avg_rank_text = f"{avg_rank:.2f}" if avg_rank is not None else "—"
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(hero_name)} · 装备详情</title>
  <style>{standalone_hero_equipment_css()}</style>
</head>
<body>
<main>
  <article class="card">
    <div class="eyebrow">HERO EQUIPMENT DETAIL</div>
    <h1>{esc(hero_name)}</h1>
    <div class="sub">{esc(str(tier) + "费" if tier is not None else "费用未知")} · {esc(traits)}</div>
    <div class="metrics">
      <b>主C率 {render_pct(hero_stats.get("carry_rate", 0))}</b>
      <b>加权平均名次 {esc(avg_rank_text)}</b>
      <b>加权前四率 {render_pct(hero_stats.get("top4_rate"))}</b>
      <b>棋子样本 {esc(str(hero_stats.get("appearances", 0)))}</b>
      <b>详情装备 {esc(str(len(detail_items)))} 件</b>
    </div>
    <h3>摘要推荐</h3>
    <p><strong>普通装备：</strong>{normal_html or '<span class="muted">样本不足</span>'}</p>
    <p><strong>超级装备：</strong>{super_html or '<span class="muted">样本不足</span>'}</p>
    <p><strong>美食社装备：</strong>{food_html or '<span class="muted">样本不足</span>'}</p>
    <h3>单装明细</h3>
    <p class="muted">仅展示该棋子单装 raw 样本量 &gt;10 的装备；平均名次/前四率等为批次加权口径。</p>
    {html_hero_equipment_detail_rows(detail_items)}
    <p><a class="hero-equipment-back" href="{esc(dashboard_href)}">返回装备概览</a></p>
  </article>
</main>
</body>
</html>
"""


def render_equipment_recommendations_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    note = (
        "可按费用、羁绊和关键词筛选；点击表头排序。"
        "点击棋子名会在新标签页打开独立装备详情（仅 raw 样本 >10）。"
        "推荐装备不含超级装备与美食社装备（见独立分栏与独立页）。完整明细仍见 Excel。"
    )
    recommendations = [
        row
        for row in data["rankings"]["heroes_and_equipment"]["carry_equipment_recommendations"]
        if row.get("has_equipment_data")
    ]
    headers = [
        ("棋子", "text"),
        ("费用", "numeric"),
        ("羁绊", "text"),
        ("主C样本", "numeric"),
        ("主C率", "numeric"),
        ("推荐装备", "text"),
        ("超级装备", "text"),
        ("美食社装备", "text"),
        ("常见三件套", "text"),
        ("低样本观察", "text"),
    ]
    table_rows: list[dict[str, Any]] = []
    for rec in recommendations:
        hero_stats = rec["hero_stats"]
        tier = hero_stats.get("tier")
        traits = rec.get("hero_traits", [])
        items = rec.get("recommended_items", [])
        super_items = rec.get("recommended_super_items", [])
        food_items = rec.get("recommended_food_items", [])
        item_html, item_text = _format_equipment_item_chips(items)
        super_html, super_text = _format_equipment_item_chips(super_items)
        food_html, food_text = _format_equipment_item_chips(food_items)
        sets = rec.get("recommended_sets", [])
        set_text = "；".join(set_row["equipment_set"] for set_row in sets[:2]) or "—"
        low_sample = rec.get("low_sample_observations", [])
        low_text = "；".join(item["equipment_name"] for item in low_sample[:2]) or "—"
        detail_href = hero_equipment_detail_href_from_dashboard(
            rec["hero_name"],
            slug=rec.get("detail_slug"),
        )
        hero_link = (
            f'<a class="hero-equipment-link" href="{esc(detail_href)}" '
            f'target="_blank" rel="noopener noreferrer">{esc(rec["hero_name"])}</a>'
        )
        search_text = " ".join(
            [
                rec["hero_name"],
                " ".join(traits),
                item_text,
                super_text,
                food_text,
                set_text,
                low_text,
            ]
        )
        table_rows.append(
            {
                "tier": tier,
                "traits": traits,
                "search_text": search_text,
                "cells": [
                    {
                        "text": rec["hero_name"],
                        "sort": rec["hero_name"],
                        "html": hero_link,
                    },
                    _html_table_cell(str(tier or "—"), sort_value=tier or 99),
                    _html_table_cell("、".join(traits) or "—", sort_value="、".join(traits)),
                    _html_table_cell(
                        str(hero_stats.get("carry_appearances", 0)),
                        sort_value=hero_stats.get("carry_appearances", 0),
                    ),
                    _html_table_cell(
                        render_pct(hero_stats.get("carry_rate", 0)),
                        sort_value=hero_stats.get("carry_rate", 0),
                    ),
                    {"text": item_text or "—", "sort": item_text, "html": item_html},
                    {"text": super_text or "—", "sort": super_text, "html": super_html},
                    {"text": food_text or "—", "sort": food_text, "html": food_html},
                    _html_table_cell(set_text, sort_value=set_text),
                    _html_table_cell(low_text, sort_value=low_text),
                ],
            }
        )
    return render_equipment_panel(
        panel_id=panel_id,
        title="棋子装备推荐",
        subtitle=subtitle,
        note=note,
        headers=headers,
        rows=table_rows,
    )


def build_special_equipment_table_rows(data: dict[str, Any], section_key: str) -> list[list[dict[str, Any]]]:
    rows: list[list[dict[str, Any]]] = []
    for item in data["rankings"].get(section_key, {}).get("rankings", []):
        wearers = "、".join(
            f"{wearer['hero_name']}(n={wearer['appearances']})"
            for wearer in item.get("recommended_wearers", [])[:3]
        ) or "待观察"
        avg_rank = item.get("adjusted_avg_rank")
        top4 = item.get("top4_rate")
        win = item.get("win_rate")
        note = item.get("note") or "—"
        rows.append(
            [
                _html_table_cell(str(item.get("strength_rank", "—")), sort_value=item.get("strength_rank", 999)),
                _html_table_cell(item["equipment_name"], sort_value=item["equipment_name"]),
                _html_table_cell(str(item.get("appearances", 0)), sort_value=item.get("appearances", 0)),
                _html_table_cell(
                    f"{item.get('weighted_appearances', 0):.2f}"
                    if item.get("weighted_appearances") is not None
                    else "—",
                    sort_value=item.get("weighted_appearances", 0) or 0,
                ),
                _html_table_cell(
                    f"{item.get('n_eff', 0):.2f}" if item.get("n_eff") is not None else "—",
                    sort_value=item.get("n_eff", 0) or 0,
                ),
                _html_table_cell(
                    f"{avg_rank:.2f}" if avg_rank is not None else "—",
                    sort_value=avg_rank if avg_rank is not None else 999,
                ),
                _html_table_cell(
                    render_pct(top4) if top4 is not None else "—",
                    sort_value=top4 if top4 is not None else -1,
                ),
                _html_table_cell(
                    render_pct(win) if win is not None else "—",
                    sort_value=win if win is not None else -1,
                ),
                _html_table_cell(item.get("confidence", "低"), sort_value=item.get("confidence", "")),
                _html_table_cell(item.get("sample_quality", "—"), sort_value=item.get("sample_quality", "")),
                _html_table_cell(wearers, sort_value=wearers),
                _html_table_cell(note, sort_value=note),
            ]
        )
    return rows


def render_special_equipment_table_panel(
    data: dict[str, Any],
    *,
    panel_id: str,
    section_key: str,
    title: str,
    note: str,
) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    definition = data["rankings"].get(section_key, {}).get("definition", "")
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    full_note = f"{definition} {note}".strip()
    headers = [
        ("强度排名", "numeric"),
        ("装备", "text"),
        ("样本", "numeric"),
        ("加权样本", "numeric"),
        ("n_eff", "numeric"),
        ("修正名次", "numeric"),
        ("前四率", "numeric"),
        ("吃鸡率", "numeric"),
        ("置信度", "text"),
        ("样本质量", "text"),
        ("推荐佩戴", "text"),
        ("备注", "text"),
    ]
    return render_sortable_table_panel(
        panel_id=panel_id,
        title=title,
        subtitle=subtitle,
        note=full_note,
        headers=headers,
        rows=build_special_equipment_table_rows(data, section_key),
    )


def render_super_equipment_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    return render_special_equipment_table_panel(
        data,
        panel_id=panel_id,
        section_key="super_equipment",
        title="超级装备强度排行",
        note="点击表头可排序；低样本仅标为观察，不作为高置信推荐。",
    )


def render_food_equipment_table_panel(data: dict[str, Any], *, panel_id: str) -> str:
    return render_special_equipment_table_panel(
        data,
        panel_id=panel_id,
        section_key="food_equipment",
        title="美食社装备强度排行",
        note="含美味/绝味/暗黑前缀装备及杏仁豆腐、椒盐酥糖、岛好锅；岛好锅若样本极少会强制低置信提示。",
    )


def html_trap_comp_card(comp: dict[str, Any]) -> str:
    stats = comp["stats"]
    popularity = comp.get("popularity", {})
    return f"""
    <article class="trap-card">
      <header class="trap-head">
        <h2>{esc(comp['label'])}</h2>
        <span class="trap-badge">版本陷阱</span>
      </header>
      <p class="trap-reason">{esc(comp.get('trap_reason', '策略整体表现偏弱'))}</p>
      <div class="metrics">
        <b>Avg {stats['avg_rank']:.2f}</b>
        <b>Top4 {render_pct(stats['top4_rate'])}</b>
        <b>n={stats['appearances']}</b>
        <b>热度 {render_pct(popularity.get('pick_rate', 0))}</b>
        <b>{esc(comp.get('confidence', '—'))}置信</b>
      </div>
      <p><strong>类型：</strong>{esc(comp.get('play_style', '高费'))}</p>
      <p><strong>路线：</strong>{esc(unique_route_bonds(comp))}</p>
      <div class="board-section">
        <h3>7 / 8 / 9 级观察阵容</h3>
        {html_variant_board_cards(comp)}
      </div>
    </article>
    """


def render_trap_compositions_panel(data: dict[str, Any], *, panel_id: str) -> str:
    quality = data["overview"]["quality"]
    generated = data["generated_at"].split("T")[0]
    traps = data["rankings"].get("traps", {}).get("compositions", [])
    cards_html = "".join(html_trap_comp_card(comp) for comp in traps) or '<p class="empty">暂无稳定陷阱阵容。</p>'
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    return f"""
  <div class="panel-section trap-panel" id="{esc(panel_id)}">
    <header class="panel-header">
      <h2>版本陷阱阵容</h2>
      <div class="sub">{esc(subtitle)} · 每个陷阱阵容展示 7/8/9 级观察阵容</div>
    </header>
    <div class="trap-list">
      {cards_html}
    </div>
  </div>
"""


def render_html_panels(data: dict[str, Any]) -> dict[str, str]:
    return {
        "composition_recommendations": render_composition_recommendations_panel(
            data, panel_id="panel-compositions"
        ),
        "primary_bond_strength": render_primary_bond_strength_table_panel(
            data, panel_id="panel-primary-bond"
        ),
        "equipment": render_equipment_recommendations_panel(data, panel_id="panel-equipment"),
        "super_equipment": render_super_equipment_table_panel(
            data, panel_id="panel-super-equipment"
        ),
        "food_equipment": render_food_equipment_table_panel(
            data, panel_id="panel-food-equipment"
        ),
        "cards_cai": render_card_prefix_table_panel(data, "彩", panel_id="panel-cards-cai"),
        "cards_yellow": render_card_prefix_table_panel(data, "黄", panel_id="panel-cards-yellow"),
        "cards_blue": render_card_prefix_table_panel(data, "蓝", panel_id="panel-cards-blue"),
        "cards_white": render_card_prefix_table_panel(data, "白", panel_id="panel-cards-white"),
        "duo_compositions": render_duo_composition_table_panel(data, panel_id="panel-duo"),
        "low_cost_carries": render_low_cost_carry_table_panel(data, panel_id="panel-low-cost"),
        "jiujiu_comps": render_jiujiu_comps_table_panel(data, panel_id="panel-jiujiu-comps"),
        "jiujiu_wearers": render_jiujiu_wearers_table_panel(data, panel_id="panel-jiujiu-wearers"),
        "trap_compositions": render_trap_compositions_panel(data, panel_id="panel-traps"),
    }


def interactive_dashboard_css() -> str:
    return """
    :root { color-scheme: dark; }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: #10131f;
      font-family: "Microsoft YaHei", "PingFang SC", "Noto Sans CJK SC", sans-serif;
      color: #e2e8f0;
    }
    .dashboard {
      width: 1200px;
      max-width: 100%;
      margin: 0 auto;
      padding: 28px 24px 42px;
      background:
        radial-gradient(circle at 10% 0%, rgba(91,141,239,.28), transparent 28%),
        radial-gradient(circle at 90% 4%, rgba(255,189,89,.18), transparent 24%),
        linear-gradient(145deg, #151a2d 0%, #0c1020 100%);
      min-height: 100vh;
    }
    .dashboard-header { margin-bottom: 16px; }
    .eyebrow { color: #fbbf24; font-weight: 800; letter-spacing: 4px; font-size: 14px; }
    .dashboard-header h1 { font-size: 36px; margin: 8px 0; line-height: 1.1; }
    .sub { color: #94a3b8; font-size: 17px; line-height: 1.45; }
    .tab-bar {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin: 18px 0;
      padding: 12px;
      border-radius: 18px;
      background: rgba(15,23,42,.72);
      border: 1px solid rgba(255,255,255,.12);
      position: sticky;
      top: 0;
      z-index: 20;
      backdrop-filter: blur(8px);
    }
    .tab-btn {
      border: 1px solid rgba(255,255,255,.14);
      background: rgba(255,255,255,.05);
      color: #cbd5e1;
      border-radius: 999px;
      padding: 8px 14px;
      cursor: pointer;
      font-size: 13px;
    }
    .tab-btn:hover { background: rgba(255,255,255,.1); }
    .tab-btn.active {
      background: rgba(251,191,36,.22);
      color: #fde68a;
      border-color: rgba(251,191,36,.45);
    }
    .dashboard-panel { display: none; }
    .dashboard-panel.active { display: block; }
    .panel-header { margin-bottom: 14px; }
    .panel-header h2 { font-size: 28px; margin: 0 0 8px; line-height: 1.15; }
    .title-row {
      display: flex;
      flex-wrap: wrap;
      align-items: baseline;
      gap: 12px 20px;
      margin: 0 0 8px;
    }
    .sort-status {
      color: #94a3b8;
      font-size: 14px;
      font-weight: 600;
      line-height: 1.2;
      white-space: nowrap;
    }
    .note { color: #94a3b8; font-size: 15px; margin: 8px 0 14px; line-height: 1.45; }
    .table-wrap {
      overflow-x: auto;
      border: 1px solid rgba(255,255,255,.12);
      border-radius: 20px;
      background: rgba(15,23,42,.55);
    }
    table { width: 100%; border-collapse: collapse; font-size: 14px; }
    th, td {
      padding: 10px 12px;
      border-bottom: 1px solid rgba(255,255,255,.08);
      text-align: left;
      vertical-align: top;
    }
    th {
      position: sticky;
      top: 64px;
      z-index: 10;
      background: rgba(15,23,42,.96);
      color: #cbd5e1;
      cursor: pointer;
      user-select: none;
      white-space: nowrap;
    }
    th.sort-asc, th.sort-desc { color: #fde68a; }
    th.sort-asc::after { content: " ▲"; color: #93c5fd; }
    th.sort-desc::after { content: " ▼"; color: #93c5fd; }
    tr:hover td { background: rgba(255,255,255,.03); }
    tr.hidden { display: none; }
    td { color: #cbd5e1; line-height: 1.35; overflow-wrap: anywhere; }
    .strategy-cell .bond { color: #e2e8f0; font-weight: 600; }
    .strategy-cell .carries {
      color: #94a3b8;
      font-size: 13px;
      margin-top: 4px;
      line-height: 1.3;
      overflow-wrap: anywhere;
    }
    .strategy-list {
      display: flex;
      flex-direction: column;
      gap: 4px;
      min-width: 120px;
      max-width: 220px;
    }
    .strategy-brief {
      display: inline-block;
      background: rgba(255,255,255,.05);
      border-radius: 8px;
      padding: 3px 8px;
      font-size: 13px;
      color: #cbd5e1;
      overflow-wrap: anywhere;
    }
    .filter-bar {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
      margin-bottom: 14px;
      padding: 14px 16px;
      border-radius: 18px;
      background: rgba(15,23,42,.55);
      border: 1px solid rgba(255,255,255,.12);
    }
    .filter-group { display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }
    .filter-label { color: #fde68a; font-size: 14px; font-weight: 700; margin-right: 4px; }
    .filter-btn, .trait-filter, .pager-btn, .style-filter {
      border: 1px solid rgba(255,255,255,.14);
      background: rgba(255,255,255,.05);
      color: #cbd5e1;
      border-radius: 999px;
      padding: 7px 12px;
      cursor: pointer;
      font-size: 13px;
    }
    .filter-btn.active:not([data-tier="all"]),
    .trait-filter.active:not([data-trait="all"]),
    .style-filter.active:not([data-style="all"]) {
      background: rgba(251,191,36,.22);
      color: #fde68a;
      border-color: rgba(251,191,36,.45);
    }
    .filter-btn.active[data-tier="all"],
    .trait-filter.active[data-trait="all"],
    .style-filter.active[data-style="all"] {
      background: rgba(255,255,255,.06);
      color: #94a3b8;
      border-color: rgba(255,255,255,.22);
    }
    .pager-btn:hover, .style-filter:hover, .filter-btn:hover, .trait-filter:hover {
      background: rgba(255,255,255,.1);
    }
    .pager-btn:disabled { opacity: .55; cursor: default; }
    .search-input {
      min-width: 220px;
      flex: 1 1 220px;
      border: 1px solid rgba(255,255,255,.14);
      background: rgba(15,23,42,.7);
      color: #e2e8f0;
      border-radius: 12px;
      padding: 8px 12px;
      font-size: 14px;
    }
    .filter-status, .page-status { color: #94a3b8; font-size: 14px; font-weight: 600; }
    .item-list { display: flex; flex-direction: column; gap: 4px; }
    .item-chip {
      display: inline-block;
      background: rgba(255,255,255,.04);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 8px;
      padding: 3px 8px;
      font-size: 13px;
      color: #cbd5e1;
    }
    .pager-bar {
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      align-items: center;
      justify-content: space-between;
      margin: 0 0 16px;
      padding: 14px 16px;
      border-radius: 18px;
      background: rgba(15,23,42,.55);
      border: 1px solid rgba(255,255,255,.12);
    }
    .pager-controls { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    .comp-page { display: none; }
    .comp-page.active { display: block; }
    .comp-page-inner {
      background: rgba(15,23,42,.6);
      border: 1px solid rgba(255,255,255,.12);
      border-radius: 28px;
      padding: 24px;
      box-shadow: 0 18px 50px rgba(0,0,0,.24);
    }
    .comp-head { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    .badge { background: #fbbf24; color: #111827; border-radius: 999px; padding: 6px 12px; font-weight: 900; }
    .archetype-badge {
      background: rgba(167,139,250,.22);
      color: #ddd6fe;
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 13px;
      font-weight: 700;
    }
    .zone-badge {
      background: rgba(52,211,153,.18);
      color: #a7f3d0;
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 13px;
      font-weight: 700;
    }
    .zone-badge.observation {
      background: rgba(248,113,113,.18);
      color: #fecaca;
    }
    .zone-badge.ceiling {
      background: rgba(251,191,36,.18);
      color: #fde68a;
    }
    .trend-badge {
      background: rgba(96,165,250,.18);
      color: #bfdbfe;
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 13px;
      font-weight: 700;
    }
    .conf-pill {
      background: rgba(147,197,253,.18);
      color: #bfdbfe;
      border-radius: 999px;
      padding: 6px 12px;
      font-size: 13px;
    }
    .comp-page-inner h2 { margin: 10px 0 0; font-size: 26px; line-height: 1.2; overflow-wrap: anywhere; }
    .metrics { display: flex; flex-wrap: wrap; gap: 10px; margin: 12px 0; color: #94a3b8; }
    .detail-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-top: 16px; }
    .detail-panel {
      background: rgba(255,255,255,.04);
      border: 1px solid rgba(255,255,255,.1);
      border-radius: 20px;
      padding: 16px;
    }
    .detail-panel h3, .board-section h3 { margin: 0 0 10px; color: #fde68a; font-size: 20px; }
    p { margin: 7px 0; color: #cbd5e1; font-size: 15px; line-height: 1.45; overflow-wrap: anywhere; }
    .board-section { margin-top: 18px; }
    .board-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 14px; }
    .board-card {
      background: rgba(255,255,255,.04);
      border: 1px solid rgba(255,255,255,.1);
      border-radius: 18px;
      padding: 14px;
      min-height: 210px;
    }
    .board-head { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 8px; }
    .level-badge, .source-badge, .conf-badge, .sample-badge {
      border-radius: 999px;
      padding: 4px 8px;
      font-size: 12px;
      font-weight: 700;
    }
    .level-badge { background: rgba(251,191,36,.22); color: #fde68a; }
    .source-badge { background: rgba(147,197,253,.18); color: #bfdbfe; }
    .conf-badge { background: rgba(255,255,255,.06); color: #94a3b8; }
    .sample-badge { background: rgba(248,113,113,.18); color: #fecaca; }
    .bond-note { color: #94a3b8; font-size: 13px; line-height: 1.4; min-height: 38px; }
    .hero-chips { display: flex; flex-wrap: wrap; gap: 6px; }
    .hero-chip {
      background: rgba(255,255,255,.05);
      border-radius: 10px;
      padding: 4px 8px;
      font-size: 13px;
      color: #e2e8f0;
      line-height: 1.3;
    }
    .trap-list { display: flex; flex-direction: column; gap: 18px; }
    .trap-card {
      background: rgba(15,23,42,.6);
      border: 1px solid rgba(248,113,113,.24);
      border-radius: 28px;
      padding: 22px;
      box-shadow: 0 18px 50px rgba(0,0,0,.24);
    }
    .trap-head { display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; flex-wrap: wrap; }
    .trap-card h2 { margin: 0; font-size: 24px; line-height: 1.2; overflow-wrap: anywhere; }
    .trap-badge {
      background: rgba(248,113,113,.22);
      color: #fecaca;
      border-radius: 999px;
      padding: 6px 12px;
      font-weight: 800;
      white-space: nowrap;
    }
    .trap-reason { color: #fecaca; margin: 10px 0; line-height: 1.45; }
    .muted, .empty { color: #94a3b8; }
    .hero-equipment-link {
      color: #fde68a;
      font-weight: 700;
      text-decoration: none;
      border-bottom: 1px dashed rgba(253,230,138,.45);
    }
    .hero-equipment-link:hover { color: #fff7c2; }
    footer { margin-top: 22px; color: #94a3b8; font-size: 14px; text-align: center; }
    @media (max-width: 900px) {
      .detail-grid, .board-grid { grid-template-columns: 1fr; }
      th { top: 0; }
    }
"""


def render_interactive_html(data: dict[str, Any]) -> str:
    quality = data["overview"]["quality"]
    generated = esc(data["generated_at"].split("T")[0])
    subtitle = (
        f"基于 {quality['matches']} 局 / {data['overview']['filtered_players']} 条过滤后玩家记录 · {generated}"
    )
    panels = render_html_panels(data)
    tab_buttons = "".join(
        f'<button type="button" class="tab-btn{" active" if index == 0 else ""}" '
        f'data-hash="{esc(hash_key)}">{esc(label)}</button>'
        for index, (hash_key, label, panel_key) in enumerate(INTERACTIVE_PANELS)
    )
    panel_sections = "".join(
        f'<section class="dashboard-panel{" active" if index == 0 else ""}" '
        f'data-hash="{esc(hash_key)}" id="dashboard-{esc(hash_key)}">'
        f'{panels[panel_key]}'
        f"</section>"
        for index, (hash_key, _label, panel_key) in enumerate(INTERACTIVE_PANELS)
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DZPPQ 环境分析详情</title>
  <style>{interactive_dashboard_css()}</style>
</head>
<body>
<main class="dashboard">
  <header class="dashboard-header">
    <div class="eyebrow">DZPPQ META DASHBOARD</div>
    <h1>环境分析详情</h1>
    <div class="sub">{subtitle} · 点击标签切换表格与详情</div>
  </header>
  <nav class="tab-bar" aria-label="环境分析视图切换">
    {tab_buttons}
  </nav>
  {panel_sections}
  <footer>完整数据见 latest_meta_analysis_report.md / latest_meta_analysis.json / latest_meta_analysis_equipment.xlsx</footer>
</main>
<script>
(function() {{
  const panels = Array.from(document.querySelectorAll(".dashboard-panel"));
  const tabs = Array.from(document.querySelectorAll(".tab-btn"));

  function activatePanel(hashKey, updateHash) {{
    const target = hashKey || (panels[0] && panels[0].dataset.hash) || "";
    panels.forEach((panel) => {{
      panel.classList.toggle("active", panel.dataset.hash === target);
    }});
    tabs.forEach((tab) => {{
      tab.classList.toggle("active", tab.dataset.hash === target);
    }});
    if (updateHash && target) {{
      history.replaceState(null, "", `#${{target}}`);
    }}
  }}

  function routeHash(rawHash, updateHash) {{
    const hashKey = (rawHash || "").replace(/^#/, "");
    if (!hashKey) {{
      activatePanel(panels[0] && panels[0].dataset.hash, updateHash);
      return;
    }}
    if (panels.some((panel) => panel.dataset.hash === hashKey)) {{
      activatePanel(hashKey, updateHash);
      return;
    }}
    activatePanel(panels[0] && panels[0].dataset.hash, updateHash);
  }}

  tabs.forEach((tab) => {{
    tab.addEventListener("click", () => activatePanel(tab.dataset.hash, true));
  }});

  routeHash(location.hash, false);

  window.addEventListener("hashchange", () => {{
    routeHash(location.hash, false);
  }});
}})();
</script>
</body>
</html>
"""


def write_html(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _xlsx_header_style():
    try:
        from openpyxl.styles import Font, PatternFill

        return Font(bold=True), PatternFill("solid", fgColor="E8EEF7")
    except ImportError:
        return None, None


def _write_xlsx_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    header_font, header_fill = _xlsx_header_style()
    ws.append(headers)
    if header_font and header_fill:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
    for row in rows:
        ws.append(row)
    for column in ws.columns:
        max_len = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 48)


def render_xlsx(data: dict[str, Any], xlsx_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    wb = Workbook()
    hero_rows: list[list[Any]] = []
    hero_super_rows: list[list[Any]] = []
    hero_food_rows: list[list[Any]] = []
    set_rows: list[list[Any]] = []
    low_sample_rows: list[list[Any]] = []
    recommendations = data["rankings"]["heroes_and_equipment"]["carry_equipment_recommendations"]

    def append_hero_item_rows(
        target: list[list[Any]],
        rec: dict[str, Any],
        items: list[dict[str, Any]],
        *,
        include_kind: bool = False,
    ) -> None:
        hero = rec["hero_stats"]
        for rank, item in enumerate(items, start=1):
            row = [
                rec["hero_name"],
                hero.get("tier"),
                hero.get("carry_rate"),
                hero.get("avg_rank"),
                hero.get("appearances"),
                rank,
                item["equipment_name"],
            ]
            if include_kind:
                row.append(item.get("equipment_kind", "normal"))
            row.extend(
                [
                    item.get("adjusted_avg_rank"),
                    item.get("avg_rank"),
                    item.get("top4_rate"),
                    item.get("appearances"),
                    item.get("selected_rate"),
                    item.get("selected_priority"),
                    item.get("sample_quality"),
                ]
            )
            target.append(row)

    for rec in recommendations:
        hero = rec["hero_stats"]
        append_hero_item_rows(hero_rows, rec, rec.get("recommended_items", []), include_kind=True)
        append_hero_item_rows(hero_super_rows, rec, rec.get("recommended_super_items", []))
        append_hero_item_rows(hero_food_rows, rec, rec.get("recommended_food_items", []))
        for rank, item in enumerate(rec.get("low_sample_observations", []), start=1):
            low_sample_rows.append(
                [
                    rec["hero_name"],
                    hero.get("tier"),
                    rank,
                    item["equipment_name"],
                    item.get("equipment_kind", "normal"),
                    item.get("adjusted_avg_rank"),
                    item.get("avg_rank"),
                    item.get("top4_rate"),
                    item.get("appearances"),
                    item.get("selected_rate"),
                    item.get("selected_priority"),
                ]
            )
        for rank, item in enumerate(rec.get("recommended_sets", []), start=1):
            set_rows.append(
                [
                    rec["hero_name"],
                    hero.get("tier"),
                    rank,
                    item.get("equipment_set"),
                    item.get("adjusted_avg_rank"),
                    item.get("avg_rank"),
                    item.get("top4_rate"),
                    item.get("appearances"),
                ]
            )

    hero_headers = [
        "英雄",
        "费用",
        "主C率(%)",
        "英雄平均名次",
        "英雄样本",
        "装备顺位",
        "装备",
        "装备类型",
        "修正名次",
        "平均名次",
        "前四率(%)",
        "样本",
        "核选占比(%)",
        "核选优先级",
        "样本质量",
    ]
    special_hero_headers = [
        "英雄",
        "费用",
        "主C率(%)",
        "英雄平均名次",
        "英雄样本",
        "装备顺位",
        "装备",
        "修正名次",
        "平均名次",
        "前四率(%)",
        "样本",
        "核选占比(%)",
        "核选优先级",
        "样本质量",
    ]

    ws_hero = wb.active
    ws_hero.title = "全英雄出装"
    _write_xlsx_sheet(ws_hero, hero_headers, hero_rows)

    ws_hero_super = wb.create_sheet("英雄超级装备推荐")
    _write_xlsx_sheet(ws_hero_super, special_hero_headers, hero_super_rows)

    ws_hero_food = wb.create_sheet("英雄美食社装备推荐")
    _write_xlsx_sheet(ws_hero_food, special_hero_headers, hero_food_rows)

    ws_comp = wb.create_sheet("阵容主C关键装备")
    comp_rows: list[list[Any]] = []
    for comp in data["rankings"].get("compositions", []):
        for note in comp.get("carry_equipment_notes", []):
            for rank, item in enumerate(note.get("items", []), start=1):
                comp_rows.append(
                    [
                        comp.get("label"),
                        comp.get("play_style"),
                        note.get("hero_name"),
                        rank,
                        item.get("equipment_name"),
                        item.get("label"),
                        item.get("appearances"),
                        item.get("use_rate"),
                        item.get("with_avg_rank"),
                        item.get("without_avg_rank"),
                        item.get("without_item_penalty"),
                        item.get("with_top4_rate"),
                        item.get("selected_rate"),
                    ]
                )
    _write_xlsx_sheet(
        ws_comp,
        [
            "阵容",
            "类型",
            "主C",
            "装备顺位",
            "装备",
            "标签",
            "样本",
            "使用率(%)",
            "带装平均名次",
            "不带平均名次",
            "不带惩罚",
            "带装前四率(%)",
            "核选占比(%)",
        ],
        comp_rows,
    )

    ws_sets = wb.create_sheet("常见三件套")
    _write_xlsx_sheet(
        ws_sets,
        ["英雄", "费用", "组合顺位", "三件套", "修正名次", "平均名次", "前四率(%)", "样本"],
        set_rows,
    )

    ws_low = wb.create_sheet("低样本观察")
    _write_xlsx_sheet(
        ws_low,
        [
            "英雄",
            "费用",
            "观察顺位",
            "装备",
            "装备类型",
            "修正名次",
            "平均名次",
            "前四率(%)",
            "样本",
            "核选占比(%)",
            "核选优先级",
        ],
        low_sample_rows,
    )

    def special_rank_rows(section_key: str) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for item in data["rankings"].get(section_key, {}).get("rankings", []):
            wearers = "、".join(
                f"{wearer['hero_name']}(n={wearer['appearances']})"
                for wearer in item.get("recommended_wearers", [])[:5]
            )
            rows.append(
                [
                    item.get("strength_rank"),
                    item.get("equipment_name"),
                    item.get("appearances"),
                    item.get("weighted_appearances"),
                    item.get("n_eff"),
                    item.get("adjusted_avg_rank"),
                    item.get("avg_rank"),
                    item.get("top4_rate"),
                    item.get("win_rate"),
                    item.get("confidence"),
                    item.get("sample_quality"),
                    wearers,
                    item.get("note"),
                ]
            )
        return rows

    special_headers = [
        "强度排名",
        "装备",
        "样本",
        "加权样本",
        "n_eff",
        "修正名次",
        "平均名次",
        "前四率(%)",
        "吃鸡率(%)",
        "置信度",
        "样本质量",
        "推荐佩戴",
        "备注",
    ]
    ws_super = wb.create_sheet("超级装备排行")
    _write_xlsx_sheet(ws_super, special_headers, special_rank_rows("super_equipment"))
    ws_food = wb.create_sheet("美食社装备排行")
    _write_xlsx_sheet(ws_food, special_headers, special_rank_rows("food_equipment"))

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def cleanup_legacy_html_outputs(data_dir: Path) -> None:
    for filename in LEGACY_HTML_FILENAMES:
        legacy_path = data_dir / filename
        if legacy_path.exists():
            legacy_path.unlink()


def cleanup_hero_equipment_dir(hero_equipment_dir: Path) -> None:
    if not hero_equipment_dir.exists():
        return
    for path in hero_equipment_dir.glob("*.html"):
        path.unlink()


def write_hero_equipment_pages(
    data: dict[str, Any],
    hero_equipment_dir: Path,
    *,
    dashboard_href: str = "../环境分析详情.html#equipment",
) -> list[dict[str, str]]:
    cleanup_hero_equipment_dir(hero_equipment_dir)
    hero_equipment_dir.mkdir(parents=True, exist_ok=True)
    written: list[dict[str, str]] = []
    recommendations = data["rankings"]["heroes_and_equipment"]["carry_equipment_recommendations"]
    for rec in recommendations:
        if not rec.get("has_equipment_data"):
            continue
        hero_name = rec["hero_name"]
        filename = hero_equipment_detail_filename(hero_name)
        path = hero_equipment_dir / filename
        write_html(
            path,
            render_hero_equipment_detail_page(rec, dashboard_href=dashboard_href),
        )
        written.append(
            {
                "hero_name": hero_name,
                "detail_slug": rec.get("detail_slug") or hero_equipment_detail_slug(hero_name),
                "path": rel(path),
            }
        )
    return written


def write_outputs(
    data: dict[str, Any],
    json_path: Path,
    md_path: Path,
    interactive_html_path: Path,
    xlsx_path: Path,
    hero_equipment_dir: Path | None = None,
) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    interactive_html_path.parent.mkdir(parents=True, exist_ok=True)
    hero_dir = hero_equipment_dir or DEFAULT_HERO_EQUIPMENT_DIR

    hero_pages = write_hero_equipment_pages(data, hero_dir)
    data["outputs"] = {
        "equipment_xlsx": rel(xlsx_path),
        "json": rel(json_path),
        "markdown": rel(md_path),
        "interactive_html": rel(interactive_html_path),
        "hero_equipment_dir": rel(hero_dir),
        "hero_equipment_pages": hero_pages,
    }
    md_path.write_text(render_md(data), encoding="utf-8")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_html(interactive_html_path, render_interactive_html(data))
    render_xlsx(data, xlsx_path)
    cleanup_legacy_html_outputs(interactive_html_path.parent)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", type=Path, default=None, help="SQLite match DB path")
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON, help="JSON output path")
    parser.add_argument("--md", type=Path, default=DEFAULT_MD, help="Markdown output path")
    parser.add_argument(
        "--html",
        type=Path,
        default=DEFAULT_INTERACTIVE_HTML,
        help="Interactive HTML dashboard output path",
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Excel equipment output path")
    parser.add_argument("--balance-notes", type=Path, default=None, help="Optional balance notes file")
    parser.add_argument(
        "--recency-half-life-days",
        type=float,
        default=DEFAULT_RECENCY_HALF_LIFE_DAYS,
        help="Half-life in days for batch recency weighting",
    )
    parser.add_argument("--min-comp-apps", type=int, default=5)
    parser.add_argument("--min-entity-apps", type=int, default=10)
    parser.add_argument("--min-card-apps", type=int, default=12)
    return parser


def resolve_output_path(path: Path) -> Path:
    return path if path.is_absolute() else ROOT / path


def main() -> None:
    args = build_parser().parse_args()
    data = build_analysis(args)
    json_path = resolve_output_path(args.json)
    md_path = resolve_output_path(args.md)
    interactive_html_path = resolve_output_path(args.html)
    xlsx_path = resolve_output_path(args.xlsx)
    write_outputs(
        data,
        json_path,
        md_path,
        interactive_html_path,
        xlsx_path,
    )
    print(f"Wrote {rel(json_path)}")
    print(f"Wrote {rel(md_path)}")
    print(f"Wrote {rel(interactive_html_path)}")
    print(f"Wrote {rel(xlsx_path)}")
    print(f"Wrote {data['outputs'].get('hero_equipment_dir', rel(DEFAULT_HERO_EQUIPMENT_DIR))}/")


if __name__ == "__main__":
    main()
