# -*- coding: utf-8 -*-
"""Tests for card catalog and card-details workbook synchronization."""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.init_card_details_workbook import sync_card_details_workbook  # noqa: E402
from src.card_catalog import build_card_catalog  # noqa: E402
from src.card_details import (  # noqa: E402
    CardDetailsValidationError,
    load_asset_candidate_map,
    load_card_details,
    read_card_details,
)


ASSETS = (
    "白·普通卡",
    "白·中坚力量",
    "白·小而美",
    "蓝·普通卡",
    "蓝·半步满级+满级玩家",
    "蓝·重质拍档支援",
    "蓝·一起刷刷刷+天降揪揪pro",
    "黄·吸吸宝pro快速成型",
    "黄·巨神兵",
    "黄·迅迅迅捷双剑",
    "黄·摇盒高手",
    "黄·终极反击",
    "黄·蛋商银行",
    "彩·普通卡",
)


class CardDetailsWorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.template_dir = self.root / "cards"
        self.template_dir.mkdir()
        for stem in ASSETS:
            (self.template_dir / f"{stem}.jpg").touch()
        self.workbook_path = self.root / "card_details.xlsx"

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def _sync(self, *, force: bool = False) -> dict[str, int]:
        return sync_card_details_workbook(
            self.workbook_path,
            template_dir=self.template_dir,
            force=force,
        )

    def test_initializes_real_cards_and_all_known_groups(self) -> None:
        counts = self._sync()
        loaded = load_card_details(
            self.workbook_path,
            template_dir=self.template_dir,
        )
        self.assertEqual(counts, {"白": 6, "蓝": 9, "黄": 10, "彩": 1, "同模板组合": 11})
        self.assertEqual(
            loaded.same_template_candidates["白·中坚力量"],
            ("白·中坚力量", "白·威力代价", "白·后院"),
        )
        self.assertEqual(
            loaded.same_template_candidates["白·小而美"],
            ("白·小而美", "白·法力专注"),
        )
        self.assertEqual(
            loaded.same_template_candidates["蓝·半步满级+满级玩家"],
            ("蓝·半步满级", "蓝·满级玩家"),
        )
        self.assertEqual(
            loaded.same_template_candidates["蓝·重质拍档支援"],
            ("蓝·重质也重量pro", "蓝·最佳拍档", "蓝·最强支援"),
        )
        self.assertEqual(
            loaded.same_template_candidates["蓝·一起刷刷刷+天降揪揪pro"],
            ("蓝·我们全都要", "蓝·一起刷刷刷", "蓝·天降揪揪pro"),
        )
        self.assertEqual(
            loaded.same_template_candidates["黄·巨神兵"],
            loaded.same_template_candidates["黄·迅迅迅捷双剑"],
        )
        self.assertEqual(
            loaded.same_template_candidates["黄·终极反击"],
            ("黄·终极反击", "黄·学术反击"),
        )
        self.assertEqual(
            loaded.same_template_candidates["黄·蛋商银行"],
            ("黄·大亨", "黄·蛋商银行"),
        )
        # Candidate names remain real names; CARD_LABEL_ALIASES is not applied.
        self.assertIn("蓝·最佳拍档", loaded.by_color["蓝"])
        self.assertIn("蓝·最强支援", loaded.by_color["蓝"])
        self.assertNotIn("蓝·拍档支援", loaded.by_color["蓝"])

    def test_public_read_and_raw_asset_map_apis(self) -> None:
        self._sync()
        details = read_card_details(
            self.workbook_path,
            template_dir=self.template_dir,
        )
        mapping = load_asset_candidate_map(
            self.workbook_path,
            template_dir=self.template_dir,
        )
        self.assertEqual(details["白"]["白·普通卡"], "")
        self.assertEqual(mapping["白·普通卡"], ("白·普通卡",))
        self.assertEqual(
            mapping["黄·摇盒高手"],
            ("黄·死亡摇滚", "黄·摇盒高手"),
        )
        self.assertEqual(
            mapping["黄·蛋商银行"],
            ("黄·大亨", "黄·蛋商银行"),
        )
        self.assertEqual(
            mapping["白·中坚力量"],
            ("白·中坚力量", "白·威力代价", "白·后院"),
        )
        self.assertEqual(
            mapping["白·小而美"],
            ("白·小而美", "白·法力专注"),
        )

    def test_sync_preserves_details_and_manual_group_until_force(self) -> None:
        self._sync()
        workbook = load_workbook(self.workbook_path)
        blue = workbook["蓝"]
        target_row = next(
            row for row in range(2, blue.max_row + 1)
            if blue.cell(row, 1).value == "蓝·最佳拍档"
        )
        blue.cell(target_row, 2, "手工详情")
        group = workbook["同模板组合"]
        group.append(
            (
                "蓝·普通卡",
                json.dumps(["蓝·手工候选A", "蓝·手工候选B"], ensure_ascii=False),
            )
        )
        workbook.save(self.workbook_path)
        workbook.close()

        self._sync()
        loaded = load_card_details(self.workbook_path, template_dir=self.template_dir)
        self.assertEqual(loaded.by_color["蓝"]["蓝·最佳拍档"], "手工详情")
        self.assertEqual(
            loaded.same_template_candidates["蓝·普通卡"],
            ("蓝·手工候选A", "蓝·手工候选B"),
        )
        self.assertIn("蓝·手工候选A", loaded.by_color["蓝"])

        self._sync(force=True)
        forced = load_card_details(self.workbook_path, template_dir=self.template_dir)
        self.assertEqual(forced.by_color["蓝"]["蓝·最佳拍档"], "手工详情")
        self.assertNotIn("蓝·普通卡", forced.same_template_candidates)
        self.assertNotIn("蓝·手工候选A", forced.by_color["蓝"])

    def test_failed_validation_keeps_formal_workbook_bytes_unchanged(self) -> None:
        self._sync()
        workbook = load_workbook(self.workbook_path)
        workbook["同模板组合"].append(
            (
                "蓝·普通卡",
                json.dumps(["蓝·只有一个候选"], ensure_ascii=False),
            )
        )
        workbook.save(self.workbook_path)
        workbook.close()
        original_bytes = self.workbook_path.read_bytes()
        temporary_path = self.workbook_path.with_name(
            f".{self.workbook_path.name}.tmp.xlsx"
        )

        with self.assertRaisesRegex(
            CardDetailsValidationError,
            "candidate JSON must contain at least two cards",
        ):
            self._sync()

        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)
        self.assertFalse(temporary_path.exists())

    def test_strict_validation_reports_group_sheet_row(self) -> None:
        self._sync()
        workbook = load_workbook(self.workbook_path)
        sheet = workbook["同模板组合"]
        row_number = sheet.max_row + 1
        sheet.append(("不存在的模板", '["蓝·普通卡", "蓝·普通卡"]'))
        workbook.save(self.workbook_path)
        workbook.close()
        with self.assertRaisesRegex(
            CardDetailsValidationError,
            rf"同模板组合 sheet row {row_number}: template asset",
        ):
            load_card_details(self.workbook_path, template_dir=self.template_dir)

    def test_catalog_is_direct_assets_plus_candidate_union(self) -> None:
        groups = {
            "蓝·普通卡": ("蓝·候选甲", "蓝·候选乙"),
        }
        catalog = build_card_catalog(self.template_dir, groups)
        self.assertIn("白·普通卡", catalog["白"])
        self.assertIn("蓝·候选甲", catalog["蓝"])
        self.assertIn("蓝·候选乙", catalog["蓝"])
        self.assertNotIn("蓝·普通卡", catalog["蓝"])

    def test_formal_workbook_has_guardian_detail(self) -> None:
        loaded = load_card_details()
        self.assertEqual(
            loaded.by_color["黄"]["黄·守护"],
            "获得1个守护头盔，己方蛋仔获得120生命值",
        )
        self.assertEqual(
            loaded.same_template_candidates["黄·巨神兵"],
            ("黄·巨神兵", "黄·迅迅迅捷双剑"),
        )
        self.assertEqual(
            loaded.same_template_candidates["蓝·半步满级+满级玩家"],
            ("蓝·半步满级", "蓝·满级玩家"),
        )
        self.assertNotIn("蓝·半步满级+满级玩家", loaded.by_color["蓝"])
        self.assertIn("蓝·半步满级", loaded.by_color["蓝"])
        self.assertIn("蓝·满级玩家", loaded.by_color["蓝"])
        self.assertIn("黄·迅迅迅捷双剑", loaded.by_color["黄"])
        self.assertIn("白·威力代价", loaded.by_color["白"])
        self.assertIn("白·后院", loaded.by_color["白"])
        self.assertIn("白·法力专注", loaded.by_color["白"])


if __name__ == "__main__":
    unittest.main()
