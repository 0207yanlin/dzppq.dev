# -*- coding: utf-8 -*-
"""Create or synchronize data/card_details.xlsx.

Shared-template groups come from src.card_catalog defaults plus preserved custom
rows. See README "共享卡牌模板开发" before adding or replacing a group.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.card_catalog import (  # noqa: E402
    CARD_COLORS,
    build_card_catalog,
    default_same_template_groups,
)
from src.card_details import (  # noqa: E402
    CARD_DETAILS_PATH,
    DETAIL_HEADERS,
    GROUP_HEADERS,
    GROUP_SHEET,
    load_card_details,
)
from src.layout import CARD_TEMPLATE_DIR  # noqa: E402


def _read_existing(path: Path) -> tuple[dict[str, str], list[tuple[str, str]]]:
    details: dict[str, str] = {}
    groups: list[tuple[str, str]] = []
    if not path.is_file():
        return details, groups
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for color in CARD_COLORS:
            if color not in workbook.sheetnames:
                continue
            sheet = workbook[color]
            for name, text, *_ in sheet.iter_rows(min_row=2, values_only=True):
                if isinstance(name, str) and name.strip():
                    details[name.strip()] = text if isinstance(text, str) else ""
        if GROUP_SHEET in workbook.sheetnames:
            sheet = workbook[GROUP_SHEET]
            for stem, candidates, *_ in sheet.iter_rows(min_row=2, values_only=True):
                if isinstance(stem, str) and stem.strip():
                    value = candidates if isinstance(candidates, str) else ""
                    groups.append((stem.strip(), value))
    finally:
        workbook.close()
    return details, groups


def _groups_for_sync(
    existing_rows: list[tuple[str, str]],
    *,
    template_dir: Path,
    force: bool,
) -> list[tuple[str, str]]:
    defaults = default_same_template_groups(template_dir)
    default_rows = [
        (stem, json.dumps(list(candidates), ensure_ascii=False))
        for stem, candidates in defaults.items()
    ]
    if force:
        return default_rows
    rows = list(existing_rows)
    existing_stems = {stem for stem, _ in existing_rows}
    rows.extend(row for row in default_rows if row[0] not in existing_stems)
    return rows


def _parse_groups(rows: list[tuple[str, str]]) -> dict[str, tuple[str, ...]]:
    groups: dict[str, tuple[str, ...]] = {}
    for stem, raw_candidates in rows:
        try:
            parsed = json.loads(raw_candidates)
        except (json.JSONDecodeError, TypeError):
            continue
        if isinstance(parsed, list) and all(isinstance(item, str) for item in parsed):
            groups[stem] = tuple(parsed)
    return groups


def sync_card_details_workbook(
    output_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
    force: bool = False,
) -> dict[str, int]:
    """Synchronize the workbook while preserving details and custom groups."""
    output_path = Path(output_path)
    existing_details, existing_group_rows = _read_existing(output_path)
    group_rows = _groups_for_sync(
        existing_group_rows,
        template_dir=Path(template_dir),
        force=force,
    )
    groups = _parse_groups(group_rows)
    catalog = build_card_catalog(Path(template_dir), groups)

    workbook = Workbook()
    workbook.remove(workbook.active)
    counts: dict[str, int] = {}
    for color in CARD_COLORS:
        sheet = workbook.create_sheet(color)
        sheet.append(DETAIL_HEADERS)
        for name in catalog[color]:
            sheet.append((name, existing_details.get(name, "")))
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 80
        counts[color] = len(catalog[color])

    group_sheet = workbook.create_sheet(GROUP_SHEET)
    group_sheet.append(GROUP_HEADERS)
    for row in group_rows:
        group_sheet.append(row)
    group_sheet.freeze_panes = "A2"
    group_sheet.column_dimensions["A"].width = 38
    group_sheet.column_dimensions["B"].width = 100
    counts[GROUP_SHEET] = len(group_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.name}.tmp.xlsx")
    try:
        try:
            workbook.save(temporary_path)
        finally:
            workbook.close()
        # Validate the complete candidate before touching the formal workbook.
        load_card_details(temporary_path, template_dir=Path(template_dir))
        temporary_path.replace(output_path)
    except BaseException:
        temporary_path.unlink(missing_ok=True)
        raise
    return counts


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=CARD_DETAILS_PATH)
    parser.add_argument("--template-dir", type=Path, default=CARD_TEMPLATE_DIR)
    parser.add_argument(
        "--force",
        action="store_true",
        help=(
            "replace 同模板组合 with built-in rows; custom-only candidate rows "
            "may disappear even though retained card details are preserved"
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    options = build_parser().parse_args(argv)
    counts = sync_card_details_workbook(
        options.output,
        template_dir=options.template_dir,
        force=options.force,
    )
    summary = ", ".join(f"{sheet}={count}" for sheet, count in counts.items())
    print(f"wrote {options.output}: {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
