# -*- coding: utf-8 -*-
"""Runtime loader and validation for the card-details workbook."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.card_catalog import (
    CARD_COLORS,
    build_asset_candidate_map,
    card_color,
    list_card_asset_stems,
)
from src.layout import CARD_TEMPLATE_DIR, ROOT

CARD_DETAILS_PATH = ROOT / "data" / "card_details.xlsx"
DETAIL_HEADERS = ("卡牌名称", "文字详情")
GROUP_SHEET = "同模板组合"
GROUP_HEADERS = ("模板名称", "卡牌列表")


class CardDetailsValidationError(ValueError):
    """Raised when card_details.xlsx cannot safely be used at runtime."""


@dataclass(frozen=True)
class CardDetails:
    """Validated workbook data used by runtime consumers."""

    by_color: dict[str, dict[str, str]]
    same_template_candidates: dict[str, tuple[str, ...]]
    asset_candidates: dict[str, tuple[str, ...]]


def _fail(sheet: str, row: int, message: str) -> None:
    raise CardDetailsValidationError(f"{sheet} sheet row {row}: {message}")


def _validate_headers(sheet, expected: tuple[str, str]) -> None:
    values = tuple(sheet.cell(1, column).value for column in range(1, 3))
    if values != expected:
        _fail(sheet.title, 1, f"expected columns {list(expected)!r}, got {list(values)!r}")
    for row in range(1, sheet.max_row + 1):
        extra_values = [
            sheet.cell(row, column).value
            for column in range(3, sheet.max_column + 1)
            if sheet.cell(row, column).value not in (None, "")
        ]
        if extra_values:
            _fail(sheet.title, row, f"unexpected extra columns: {extra_values!r}")


def _read_color_sheets(workbook) -> dict[str, dict[str, str]]:
    details: dict[str, dict[str, str]] = {}
    for color in CARD_COLORS:
        sheet = workbook[color]
        _validate_headers(sheet, DETAIL_HEADERS)
        rows: dict[str, str] = {}
        for row_number in range(2, sheet.max_row + 1):
            name_value = sheet.cell(row_number, 1).value
            detail_value = sheet.cell(row_number, 2).value
            if name_value in (None, "") and detail_value in (None, ""):
                continue
            if not isinstance(name_value, str) or not name_value.strip():
                _fail(color, row_number, "卡牌名称 must be a non-empty string")
            name = name_value.strip()
            if card_color(name) != color:
                _fail(color, row_number, f"card {name!r} does not have prefix {color}·")
            if name in rows:
                _fail(color, row_number, f"duplicate card name {name!r}")
            if detail_value is None:
                detail = ""
            elif isinstance(detail_value, str):
                detail = detail_value
            else:
                _fail(color, row_number, "文字详情 must be text or empty")
            rows[name] = detail
        details[color] = rows
    return details


def _read_same_template_sheet(
    workbook,
    *,
    asset_stems: set[str],
    details: dict[str, dict[str, str]],
) -> dict[str, tuple[str, ...]]:
    sheet = workbook[GROUP_SHEET]
    _validate_headers(sheet, GROUP_HEADERS)
    groups: dict[str, tuple[str, ...]] = {}
    for row_number in range(2, sheet.max_row + 1):
        stem_value = sheet.cell(row_number, 1).value
        candidates_value = sheet.cell(row_number, 2).value
        if stem_value in (None, "") and candidates_value in (None, ""):
            continue
        if not isinstance(stem_value, str) or not stem_value.strip():
            _fail(GROUP_SHEET, row_number, "模板名称 must be a non-empty string")
        stem = stem_value.strip()
        if stem in groups:
            _fail(GROUP_SHEET, row_number, f"duplicate template name {stem!r}")
        if stem not in asset_stems:
            _fail(GROUP_SHEET, row_number, f"template asset {stem!r}.jpg does not exist")
        if not isinstance(candidates_value, str):
            _fail(GROUP_SHEET, row_number, "卡牌列表 must be a JSON array string")
        try:
            parsed = json.loads(candidates_value)
        except json.JSONDecodeError as exc:
            _fail(GROUP_SHEET, row_number, f"invalid candidate JSON: {exc.msg}")
        if not isinstance(parsed, list) or len(parsed) < 2:
            _fail(GROUP_SHEET, row_number, "candidate JSON must contain at least two cards")
        if any(not isinstance(name, str) or not name.strip() for name in parsed):
            _fail(GROUP_SHEET, row_number, "every candidate must be a non-empty string")
        candidates = tuple(name.strip() for name in parsed)
        if len(set(candidates)) != len(candidates):
            _fail(GROUP_SHEET, row_number, "candidate card names must not repeat")
        for candidate in candidates:
            color = card_color(candidate)
            if color is None:
                _fail(
                    GROUP_SHEET,
                    row_number,
                    f"candidate {candidate!r} must include a color prefix",
                )
            if candidate not in details[color]:
                _fail(
                    GROUP_SHEET,
                    row_number,
                    f"candidate {candidate!r} is missing from sheet {color}",
                )
        groups[stem] = candidates
    return groups


def load_card_details(
    workbook_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
) -> CardDetails:
    """Load and strictly validate details before an ADB runtime starts."""
    path = Path(workbook_path)
    if not path.is_file():
        raise CardDetailsValidationError(f"card details workbook does not exist: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheets = [*CARD_COLORS, GROUP_SHEET]
        if workbook.sheetnames != expected_sheets:
            raise CardDetailsValidationError(
                f"workbook sheets must be exactly {expected_sheets!r}, "
                f"got {workbook.sheetnames!r}"
            )
        details = _read_color_sheets(workbook)
        asset_stems = set(list_card_asset_stems(template_dir))
        groups = _read_same_template_sheet(
            workbook,
            asset_stems=asset_stems,
            details=details,
        )
    finally:
        workbook.close()
    return CardDetails(
        by_color=details,
        same_template_candidates=groups,
        asset_candidates=build_asset_candidate_map(sorted(asset_stems), groups),
    )


def read_card_details(
    workbook_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
) -> dict[str, dict[str, str]]:
    """Read the four color detail maps after strict validation."""
    return load_card_details(workbook_path, template_dir=template_dir).by_color


def load_asset_candidate_map(
    workbook_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
) -> dict[str, tuple[str, ...]]:
    """Build raw asset stem -> real candidate card names."""
    return load_card_details(workbook_path, template_dir=template_dir).asset_candidates


def load_same_template_candidates(
    workbook_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
) -> dict[str, tuple[str, ...]]:
    """Read validated raw template stem -> grouped real card names."""
    return load_card_details(
        workbook_path,
        template_dir=template_dir,
    ).same_template_candidates


def validate_card_details_workbook(
    workbook_path: Path = CARD_DETAILS_PATH,
    *,
    template_dir: Path = CARD_TEMPLATE_DIR,
) -> CardDetails:
    """Validate the workbook and return its runtime-ready representation."""
    return load_card_details(workbook_path, template_dir=template_dir)
